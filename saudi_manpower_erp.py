#!/usr/bin/env python3
"""
Saudi Manpower ERP Platform
============================
A comprehensive, production-ready Enterprise Resource Planning system
specifically designed for Saudi Arabian manpower and staffing companies.

Features:
- Complete HR & Payroll Management with Saudi Labor Law compliance
- GOSI calculation (Saudi and non-Saudi rates)
- EOSB (End of Service Benefit) calculation
- WPS (Wage Protection System) SIF file generation
- ZATCA-compliant e-invoicing with TLV QR codes
- Full Accounting with Chart of Accounts
- Contract and Client management
- Leave management
- Asset management
- Financial reporting (P&L, Balance Sheet, Trial Balance)
- PySide6 modern GUI with Arabic support

Author: Saudi Manpower ERP Team
Version: 2.0.0
License: Commercial
"""

# =============================================================================
# SECTION 1: IMPORTS AND CONSTANTS
# =============================================================================

import sys
import os
import re
import json
import math
import base64
import struct
import hashlib
import logging
import sqlite3
import datetime
import calendar
import threading
import traceback
import io
import csv
import time
import uuid
import copy
import textwrap
import string
import random
import tempfile
import shutil
import subprocess
import platform
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple, Union, Callable
from functools import partial, lru_cache
from contextlib import contextmanager
from dataclasses import dataclass, field, asdict
from enum import Enum, IntEnum
from collections import defaultdict, OrderedDict
from decimal import Decimal, ROUND_HALF_UP

# Third-party imports
try:
    import bcrypt
    HAS_BCRYPT = True
except ImportError:
    HAS_BCRYPT = False
    logging.warning("bcrypt not available, using SHA-256 fallback")

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import numpy as np
    HAS_NUMPY = True
except ImportError:
    HAS_NUMPY = False

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    from matplotlib.figure import Figure
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False

try:
    import qrcode
    HAS_QRCODE = True
except ImportError:
    HAS_QRCODE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm, inch
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        Image as RLImage, PageBreak, HRFlowable, KeepTogether
    )
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.graphics.shapes import Drawing, Rect, String
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    from dateutil.relativedelta import relativedelta
    from dateutil.parser import parse as date_parse
    HAS_DATEUTIL = True
except ImportError:
    HAS_DATEUTIL = False

# PySide6 imports
try:
    from PySide6.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QGridLayout, QFormLayout, QStackedWidget, QSplitter,
        QLabel, QPushButton, QLineEdit, QTextEdit, QPlainTextEdit,
        QComboBox, QSpinBox, QDoubleSpinBox, QCheckBox, QRadioButton,
        QTabWidget, QTableWidget, QTableWidgetItem, QTreeWidget,
        QTreeWidgetItem, QListWidget, QListWidgetItem, QHeaderView,
        QScrollArea, QFrame, QGroupBox, QProgressBar, QSlider,
        QDateEdit, QTimeEdit, QDateTimeEdit, QCalendarWidget,
        QDialog, QDialogButtonBox, QFileDialog, QMessageBox,
        QInputDialog, QColorDialog, QFontDialog, QProgressDialog,
        QSizePolicy, QToolBar, QStatusBar, QMenuBar, QMenu,
        QSystemTrayIcon, QDockWidget, QAction, QToolButton,
        QAbstractItemView, QButtonGroup, QStackedLayout,
        QScrollBar, QSplashScreen
    )
    from PySide6.QtCore import (
        Qt, QTimer, QThread, QObject, Signal, Slot, QRunnable,
        QThreadPool, QSize, QPoint, QRect, QDate, QDateTime,
        QTime, QUrl, QSettings, QTranslator, QLocale,
        QAbstractTableModel, QModelIndex, QSortFilterProxyModel,
        QRegularExpression, QPropertyAnimation, QEasingCurve,
        QParallelAnimationGroup, QSequentialAnimationGroup,
        QEvent, QMimeData
    )
    from PySide6.QtGui import (
        QFont, QFontMetrics, QIcon, QPixmap, QImage, QColor,
        QPainter, QPen, QBrush, QLinearGradient, QRadialGradient,
        QPainterPath, QKeySequence, QAction as QGuiAction,
        QCursor, QTextDocument, QTextCursor, QValidator,
        QIntValidator, QDoubleValidator, QRegularExpressionValidator,
        QPalette, QShortcut, QStandardItemModel, QStandardItem
    )
    HAS_PYSIDE6 = True
except ImportError:
    HAS_PYSIDE6 = False
    print("ERROR: PySide6 not installed. Run: pip install PySide6")
    sys.exit(1)

# =============================================================================
# CONSTANTS
# =============================================================================

APP_NAME = "Saudi Manpower ERP"
APP_VERSION = "2.0.0"
APP_AUTHOR = "WorkforcePro GCC"
APP_COPYRIGHT = f"© {datetime.datetime.now().year} {APP_AUTHOR}"
APP_WEBSITE = "https://workforcepro.com.sa"

# Database
DB_PATH = Path.home() / ".saudi_manpower_erp" / "erp.db"
DB_VERSION = 10

# Company defaults
DEFAULT_COMPANY_NAME = "Saudi Manpower Company"
DEFAULT_COMPANY_NAME_AR = "شركة القوى العاملة السعودية"
DEFAULT_COUNTRY = "Saudi Arabia"
DEFAULT_CURRENCY = "SAR"
DEFAULT_VAT_RATE = 0.15  # 15% VAT in Saudi Arabia
DEFAULT_TIMEZONE = "Asia/Riyadh"

# GOSI Rates
GOSI_SAUDI_EMPLOYER_RATE = 0.0975    # 9.75%
GOSI_SAUDI_EMPLOYEE_RATE = 0.0975    # 9.75%
GOSI_NON_SAUDI_EMPLOYER_RATE = 0.02  # 2.00% (work hazard only)
GOSI_NON_SAUDI_EMPLOYEE_RATE = 0.00  # 0%
GOSI_SALARY_CEILING = 45000          # SAR 45,000 max GOSI base

# Annual leave per Saudi Labor Law
ANNUAL_LEAVE_DAYS = 21  # First 5 years
ANNUAL_LEAVE_DAYS_SENIOR = 30  # After 5 years

# EOSB per Saudi Labor Law
EOSB_HALF_MONTH_YEARS = 5   # First 5 years: 0.5 month per year
EOSB_FULL_MONTH_START = 5   # After 5 years: 1 month per year

# Logging configuration
LOG_DIR = Path.home() / ".saudi_manpower_erp" / "logs"
LOG_FILE = LOG_DIR / f"erp_{datetime.date.today().strftime('%Y%m%d')}.log"

# UI Constants
SIDEBAR_WIDTH = 240
HEADER_HEIGHT = 60
STATUS_BAR_HEIGHT = 25
ICON_SIZE = 20
TABLE_ROW_HEIGHT = 36
ANIMATION_DURATION = 200

# Pagination
DEFAULT_PAGE_SIZE = 50
MAX_PAGE_SIZE = 200

# File paths
ASSETS_DIR = Path(__file__).parent / "assets"
TEMP_DIR = Path(tempfile.gettempdir()) / "saudi_erp"

# Ensure directories exist
DB_PATH.parent.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(parents=True, exist_ok=True)
TEMP_DIR.mkdir(parents=True, exist_ok=True)

# =============================================================================
# LOGGING SETUP
# =============================================================================

def setup_logging():
    """Configure application-wide logging."""
    formatter = logging.Formatter(
        '%(asctime)s [%(levelname)s] %(name)s:%(lineno)d - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    # File handler
    file_handler = logging.FileHandler(LOG_FILE, encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)

    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    # Root logger
    root_logger = logging.getLogger()
    root_logger.setLevel(logging.DEBUG)
    root_logger.addHandler(file_handler)
    root_logger.addHandler(console_handler)

    return logging.getLogger(APP_NAME)

logger = setup_logging()

# =============================================================================
# SECTION 2: CORE FRAMEWORK LAYER
# =============================================================================

class Config:
    """
    Configuration manager for the ERP system.
    Manages settings with defaults, persistence, and type conversion.
    """

    DEFAULTS = {
        "app.language": "en",
        "app.theme": "light",
        "app.font_size": 10,
        "app.show_toolbar": True,
        "app.show_statusbar": True,
        "db.path": str(DB_PATH),
        "db.backup_enabled": True,
        "db.backup_interval_days": 7,
        "company.name": DEFAULT_COMPANY_NAME,
        "company.name_ar": DEFAULT_COMPANY_NAME_AR,
        "company.country": DEFAULT_COUNTRY,
        "company.currency": DEFAULT_CURRENCY,
        "company.vat_rate": DEFAULT_VAT_RATE,
        "payroll.gosi_saudi_employer_rate": GOSI_SAUDI_EMPLOYER_RATE,
        "payroll.gosi_saudi_employee_rate": GOSI_SAUDI_EMPLOYEE_RATE,
        "payroll.gosi_non_saudi_employer_rate": GOSI_NON_SAUDI_EMPLOYER_RATE,
        "payroll.gosi_salary_ceiling": GOSI_SALARY_CEILING,
        "payroll.payment_day": 25,
        "payroll.wps_bank_code": "RJHI",
        "leave.annual_days": ANNUAL_LEAVE_DAYS,
        "leave.sick_days": 30,
        "notifications.email_enabled": False,
        "notifications.check_interval_minutes": 5,
        "ui.records_per_page": DEFAULT_PAGE_SIZE,
        "ui.confirm_delete": True,
        "ui.auto_save": True,
    }

    def __init__(self):
        self._settings: Dict[str, Any] = dict(self.DEFAULTS)
        self._qt_settings = QSettings(APP_AUTHOR, APP_NAME)
        self._load_from_qt()

    def _load_from_qt(self):
        """Load settings from Qt persistent storage."""
        for key in self.DEFAULTS:
            qt_val = self._qt_settings.value(key)
            if qt_val is not None:
                self._settings[key] = qt_val

    def get(self, key: str, default: Any = None) -> Any:
        """Get a configuration value."""
        return self._settings.get(key, default if default is not None else self.DEFAULTS.get(key))

    def set(self, key: str, value: Any):
        """Set a configuration value and persist it."""
        self._settings[key] = value
        self._qt_settings.setValue(key, value)
        self._qt_settings.sync()

    def get_bool(self, key: str, default: bool = False) -> bool:
        """Get a boolean configuration value."""
        val = self.get(key, default)
        if isinstance(val, str):
            return val.lower() in ('true', '1', 'yes')
        return bool(val)

    def get_int(self, key: str, default: int = 0) -> int:
        """Get an integer configuration value."""
        try:
            return int(self.get(key, default))
        except (ValueError, TypeError):
            return default

    def get_float(self, key: str, default: float = 0.0) -> float:
        """Get a float configuration value."""
        try:
            return float(self.get(key, default))
        except (ValueError, TypeError):
            return default

    def reset_to_defaults(self):
        """Reset all settings to default values."""
        self._settings = dict(self.DEFAULTS)
        self._qt_settings.clear()

    def export_settings(self) -> dict:
        """Export current settings as dictionary."""
        return dict(self._settings)

    def import_settings(self, settings: dict):
        """Import settings from dictionary."""
        for key, value in settings.items():
            if key in self.DEFAULTS:
                self.set(key, value)


class ServiceContainer:
    """
    Dependency injection container for the ERP system.
    Manages service instances and their dependencies.
    """

    def __init__(self):
        self._services: Dict[str, Any] = {}
        self._factories: Dict[str, Callable] = {}
        self._singletons: Dict[str, Any] = {}

    def register(self, name: str, factory: Callable, singleton: bool = True):
        """Register a service factory."""
        self._factories[name] = factory
        if not singleton:
            return
        # Eager singleton registration
        logger.debug(f"Registered service: {name}")

    def register_instance(self, name: str, instance: Any):
        """Register an existing instance as a service."""
        self._singletons[name] = instance
        logger.debug(f"Registered instance: {name}")

    def get(self, name: str) -> Any:
        """Get a service by name."""
        # Check singleton cache
        if name in self._singletons:
            return self._singletons[name]

        # Create from factory
        if name in self._factories:
            instance = self._factories[name](self)
            self._singletons[name] = instance
            return instance

        raise KeyError(f"Service not found: {name}")

    def has(self, name: str) -> bool:
        """Check if a service is registered."""
        return name in self._singletons or name in self._factories


class ModuleRegistry:
    """
    Registry for ERP modules.
    Manages module registration, dependencies, and lifecycle.
    """

    def __init__(self):
        self._modules: Dict[str, dict] = {}
        self._enabled: set = set()

    def register_module(self, name: str, display_name: str, icon: str,
                        view_class: type, requires_permission: str = None,
                        order: int = 0):
        """Register an ERP module."""
        self._modules[name] = {
            "name": name,
            "display_name": display_name,
            "icon": icon,
            "view_class": view_class,
            "requires_permission": requires_permission,
            "order": order,
            "enabled": True,
        }
        self._enabled.add(name)
        logger.debug(f"Registered module: {name}")

    def get_module(self, name: str) -> dict:
        """Get module configuration."""
        return self._modules.get(name)

    def get_all_modules(self) -> List[dict]:
        """Get all modules sorted by order."""
        return sorted(
            [m for m in self._modules.values() if m["enabled"]],
            key=lambda x: x["order"]
        )

    def enable_module(self, name: str):
        """Enable a module."""
        if name in self._modules:
            self._modules[name]["enabled"] = True
            self._enabled.add(name)

    def disable_module(self, name: str):
        """Disable a module."""
        if name in self._modules:
            self._modules[name]["enabled"] = False
            self._enabled.discard(name)

    def is_enabled(self, name: str) -> bool:
        """Check if a module is enabled."""
        return name in self._enabled


class EventBus:
    """
    Simple event system for inter-module communication.
    Allows modules to communicate without direct coupling.
    """

    def __init__(self):
        self._listeners: Dict[str, List[Callable]] = defaultdict(list)
        self._once_listeners: Dict[str, List[Callable]] = defaultdict(list)

    def on(self, event: str, callback: Callable):
        """Subscribe to an event."""
        self._listeners[event].append(callback)

    def once(self, event: str, callback: Callable):
        """Subscribe to an event, but only trigger once."""
        self._once_listeners[event].append(callback)

    def off(self, event: str, callback: Callable = None):
        """Unsubscribe from an event."""
        if callback is None:
            self._listeners.pop(event, None)
            self._once_listeners.pop(event, None)
        else:
            if event in self._listeners:
                self._listeners[event] = [
                    cb for cb in self._listeners[event] if cb != callback
                ]

    def emit(self, event: str, *args, **kwargs):
        """Emit an event to all subscribers."""
        # Regular listeners
        for callback in self._listeners.get(event, []):
            try:
                callback(*args, **kwargs)
            except Exception as e:
                logger.error(f"Error in event handler for {event}: {e}")

        # Once listeners
        once_callbacks = self._once_listeners.pop(event, [])
        for callback in once_callbacks:
            try:
                callback(*args, **kwargs)
            except Exception as e:
                logger.error(f"Error in once-event handler for {event}: {e}")

    def emit_async(self, event: str, *args, **kwargs):
        """Emit an event asynchronously in a separate thread."""
        thread = threading.Thread(
            target=self.emit, args=(event, *args), kwargs=kwargs, daemon=True
        )
        thread.start()


# Global instances
config = Config()
event_bus = EventBus()
module_registry = ModuleRegistry()
container = ServiceContainer()

# =============================================================================
# SECTION 3: DATABASE ENGINE
# =============================================================================

class DatabaseManager:
    """
    SQLite database manager with WAL mode, migrations, and helper methods.
    Handles all database operations with connection pooling simulation.
    """

    def __init__(self, db_path: str = None):
        self.db_path = str(db_path or DB_PATH)
        self._local = threading.local()
        self._lock = threading.Lock()
        self._version = DB_VERSION
        logger.info(f"Database path: {self.db_path}")
        self._ensure_directory()

    def _ensure_directory(self):
        """Ensure the database directory exists."""
        Path(self.db_path).parent.mkdir(parents=True, exist_ok=True)

    @property
    def conn(self) -> sqlite3.Connection:
        """Get thread-local database connection."""
        if not hasattr(self._local, 'connection') or self._local.connection is None:
            self._local.connection = self._create_connection()
        return self._local.connection

    def _create_connection(self) -> sqlite3.Connection:
        """Create a new database connection with optimal settings."""
        conn = sqlite3.connect(
            self.db_path,
            detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES,
            timeout=30,
            check_same_thread=False
        )
        conn.row_factory = sqlite3.Row

        # Enable WAL mode for better concurrency
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA cache_size=10000")
        conn.execute("PRAGMA temp_store=MEMORY")
        conn.execute("PRAGMA mmap_size=268435456")  # 256MB
        conn.execute("PRAGMA foreign_keys=ON")
        conn.execute("PRAGMA auto_vacuum=INCREMENTAL")

        logger.debug(f"Created database connection (thread: {threading.current_thread().name})")
        return conn

    def close(self):
        """Close the thread-local connection."""
        if hasattr(self._local, 'connection') and self._local.connection:
            self._local.connection.close()
            self._local.connection = None

    @contextmanager
    def transaction(self):
        """Context manager for database transactions."""
        conn = self.conn
        try:
            conn.execute("BEGIN")
            yield conn
            conn.execute("COMMIT")
        except Exception as e:
            conn.execute("ROLLBACK")
            logger.error(f"Transaction rolled back: {e}")
            raise

    def execute(self, sql: str, params: tuple = ()) -> sqlite3.Cursor:
        """Execute a SQL statement."""
        try:
            cursor = self.conn.cursor()
            cursor.execute(sql, params)
            self.conn.commit()
            return cursor
        except sqlite3.Error as e:
            logger.error(f"SQL Error: {e}\nSQL: {sql}\nParams: {params}")
            raise

    def execute_many(self, sql: str, params_list: List[tuple]) -> int:
        """Execute a SQL statement with multiple parameter sets."""
        try:
            cursor = self.conn.cursor()
            cursor.executemany(sql, params_list)
            self.conn.commit()
            return cursor.rowcount
        except sqlite3.Error as e:
            logger.error(f"SQL Error (executemany): {e}\nSQL: {sql}")
            raise

    def fetchall(self, sql: str, params: tuple = ()) -> List[dict]:
        """Execute a SELECT and return all rows as dictionaries."""
        try:
            cursor = self.conn.cursor()
            cursor.execute(sql, params)
            rows = cursor.fetchall()
            return [dict(row) for row in rows]
        except sqlite3.Error as e:
            logger.error(f"Fetchall Error: {e}\nSQL: {sql}\nParams: {params}")
            raise

    def fetchone(self, sql: str, params: tuple = ()) -> Optional[dict]:
        """Execute a SELECT and return one row as dictionary."""
        try:
            cursor = self.conn.cursor()
            cursor.execute(sql, params)
            row = cursor.fetchone()
            return dict(row) if row else None
        except sqlite3.Error as e:
            logger.error(f"Fetchone Error: {e}\nSQL: {sql}\nParams: {params}")
            raise

    def fetchscalar(self, sql: str, params: tuple = ()) -> Any:
        """Execute a SELECT and return a single scalar value."""
        row = self.fetchone(sql, params)
        if row:
            return list(row.values())[0]
        return None

    def insert(self, table: str, data: dict) -> int:
        """Insert a record and return the new ID."""
        columns = ', '.join(data.keys())
        placeholders = ', '.join(['?'] * len(data))
        sql = f"INSERT INTO {table} ({columns}) VALUES ({placeholders})"
        cursor = self.execute(sql, tuple(data.values()))
        return cursor.lastrowid

    def update(self, table: str, data: dict, where: str, where_params: tuple = ()) -> int:
        """Update records and return number of affected rows."""
        set_clause = ', '.join([f"{k} = ?" for k in data.keys()])
        sql = f"UPDATE {table} SET {set_clause} WHERE {where}"
        cursor = self.execute(sql, tuple(data.values()) + where_params)
        return cursor.rowcount

    def delete(self, table: str, where: str, where_params: tuple = ()) -> int:
        """Delete records and return number of affected rows."""
        sql = f"DELETE FROM {table} WHERE {where}"
        cursor = self.execute(sql, where_params)
        return cursor.rowcount

    def table_exists(self, table_name: str) -> bool:
        """Check if a table exists."""
        result = self.fetchone(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (table_name,)
        )
        return result is not None

    def get_schema_version(self) -> int:
        """Get the current schema version."""
        try:
            result = self.fetchone("SELECT version FROM schema_version ORDER BY id DESC LIMIT 1")
            return result['version'] if result else 0
        except Exception:
            return 0

    def create_tables(self):
        """Create all database tables with proper schemas."""
        logger.info("Creating database tables...")

        tables_sql = [
            # Schema version tracking
            """
            CREATE TABLE IF NOT EXISTS schema_version (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                version INTEGER NOT NULL,
                applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                description TEXT
            )
            """,

            # =========================================================
            # COMPANY & ORGANIZATION
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                name_ar TEXT,
                cr_number TEXT UNIQUE,
                vat_number TEXT,
                address TEXT,
                address_ar TEXT,
                city TEXT,
                region TEXT,
                country TEXT DEFAULT 'Saudi Arabia',
                postal_code TEXT,
                phone TEXT,
                fax TEXT,
                email TEXT,
                website TEXT,
                logo_path TEXT,
                bank_name TEXT,
                bank_account TEXT,
                bank_iban TEXT,
                bank_swift TEXT,
                currency TEXT DEFAULT 'SAR',
                fiscal_year_start TEXT DEFAULT '01-01',
                default_vat_rate REAL DEFAULT 0.15,
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_companies_cr ON companies(cr_number)",

            """
            CREATE TABLE IF NOT EXISTS branches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                name TEXT NOT NULL,
                name_ar TEXT,
                code TEXT,
                manager_name TEXT,
                address TEXT,
                city TEXT,
                region TEXT,
                phone TEXT,
                email TEXT,
                is_active INTEGER DEFAULT 1,
                is_head_office INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_branches_company ON branches(company_id)",

            """
            CREATE TABLE IF NOT EXISTS departments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                branch_id INTEGER REFERENCES branches(id),
                parent_id INTEGER REFERENCES departments(id),
                name TEXT NOT NULL,
                name_ar TEXT,
                code TEXT,
                manager_id INTEGER,
                cost_center_id INTEGER,
                description TEXT,
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_departments_company ON departments(company_id)",

            """
            CREATE TABLE IF NOT EXISTS cost_centers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                code TEXT NOT NULL,
                name TEXT NOT NULL,
                name_ar TEXT,
                description TEXT,
                parent_id INTEGER REFERENCES cost_centers(id),
                budget REAL DEFAULT 0,
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_cost_centers_company ON cost_centers(company_id)",

            # =========================================================
            # USERS, ROLES, PERMISSIONS
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS roles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                display_name TEXT,
                description TEXT,
                is_system INTEGER DEFAULT 0,
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS permissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                module TEXT NOT NULL,
                action TEXT NOT NULL,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_permissions_module ON permissions(module)",

            """
            CREATE TABLE IF NOT EXISTS role_permissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                role_id INTEGER NOT NULL REFERENCES roles(id) ON DELETE CASCADE,
                permission_id INTEGER NOT NULL REFERENCES permissions(id) ON DELETE CASCADE,
                UNIQUE(role_id, permission_id)
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                email TEXT,
                full_name TEXT,
                full_name_ar TEXT,
                employee_id INTEGER,
                company_id INTEGER REFERENCES companies(id),
                branch_id INTEGER REFERENCES branches(id),
                department_id INTEGER REFERENCES departments(id),
                is_active INTEGER DEFAULT 1,
                is_admin INTEGER DEFAULT 0,
                must_change_password INTEGER DEFAULT 0,
                last_login TIMESTAMP,
                login_attempts INTEGER DEFAULT 0,
                locked_until TIMESTAMP,
                language TEXT DEFAULT 'en',
                theme TEXT DEFAULT 'light',
                avatar_path TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_users_username ON users(username)",

            """
            CREATE TABLE IF NOT EXISTS user_roles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                role_id INTEGER NOT NULL REFERENCES roles(id) ON DELETE CASCADE,
                UNIQUE(user_id, role_id)
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS user_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id),
                session_token TEXT NOT NULL,
                ip_address TEXT,
                user_agent TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                expires_at TIMESTAMP,
                is_active INTEGER DEFAULT 1
            )
            """,

            # =========================================================
            # CLIENTS
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS clients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                client_number TEXT UNIQUE,
                name TEXT NOT NULL,
                name_ar TEXT,
                cr_number TEXT,
                vat_number TEXT,
                industry TEXT,
                client_type TEXT DEFAULT 'corporate',
                address TEXT,
                address_ar TEXT,
                city TEXT,
                region TEXT,
                country TEXT DEFAULT 'Saudi Arabia',
                postal_code TEXT,
                phone TEXT,
                fax TEXT,
                email TEXT,
                website TEXT,
                payment_terms INTEGER DEFAULT 30,
                credit_limit REAL DEFAULT 0,
                account_id INTEGER,
                salesperson_id INTEGER REFERENCES users(id),
                status TEXT DEFAULT 'active',
                rating TEXT,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_clients_company ON clients(company_id)",
            "CREATE INDEX IF NOT EXISTS idx_clients_name ON clients(name)",
            "CREATE INDEX IF NOT EXISTS idx_clients_vat ON clients(vat_number)",

            """
            CREATE TABLE IF NOT EXISTS client_contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                client_id INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
                first_name TEXT NOT NULL,
                last_name TEXT,
                job_title TEXT,
                department TEXT,
                phone TEXT,
                mobile TEXT,
                email TEXT,
                is_primary INTEGER DEFAULT 0,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_client_contacts_client ON client_contacts(client_id)",

            """
            CREATE TABLE IF NOT EXISTS client_statements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                client_id INTEGER NOT NULL REFERENCES clients(id),
                statement_date DATE NOT NULL,
                opening_balance REAL DEFAULT 0,
                total_invoiced REAL DEFAULT 0,
                total_paid REAL DEFAULT 0,
                closing_balance REAL DEFAULT 0,
                generated_by INTEGER REFERENCES users(id),
                generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            # =========================================================
            # EMPLOYEES
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                emp_number TEXT UNIQUE,
                first_name TEXT NOT NULL,
                last_name TEXT,
                middle_name TEXT,
                arabic_name TEXT,
                nationality TEXT,
                id_type TEXT DEFAULT 'iqama',
                id_number TEXT,
                iqama_number TEXT UNIQUE,
                iqama_expiry DATE,
                iqama_issue_place TEXT,
                passport_number TEXT,
                passport_expiry DATE,
                passport_issue_country TEXT,
                visa_number TEXT,
                visa_type TEXT,
                visa_expiry DATE,
                visa_issue_date DATE,
                border_number TEXT,
                date_of_birth DATE,
                place_of_birth TEXT,
                gender TEXT DEFAULT 'male',
                marital_status TEXT DEFAULT 'single',
                religion TEXT,
                education TEXT,
                education_field TEXT,
                university TEXT,
                graduation_year INTEGER,
                mobile TEXT,
                mobile_2 TEXT,
                email TEXT,
                emergency_contact TEXT,
                emergency_phone TEXT,
                emergency_relation TEXT,
                current_address TEXT,
                permanent_address TEXT,
                hire_date DATE,
                probation_end_date DATE,
                termination_date DATE,
                termination_reason TEXT,
                rehire_eligible INTEGER DEFAULT 1,
                job_title TEXT,
                job_title_ar TEXT,
                department_id INTEGER REFERENCES departments(id),
                branch_id INTEGER REFERENCES branches(id),
                cost_center_id INTEGER REFERENCES cost_centers(id),
                company_id INTEGER NOT NULL REFERENCES companies(id),
                direct_manager_id INTEGER REFERENCES employees(id),
                employment_type TEXT DEFAULT 'full_time',
                contract_type TEXT DEFAULT 'indefinite',
                work_location TEXT,
                salary_type TEXT DEFAULT 'monthly',
                basic_salary REAL DEFAULT 0,
                housing_allowance REAL DEFAULT 0,
                transport_allowance REAL DEFAULT 0,
                food_allowance REAL DEFAULT 0,
                phone_allowance REAL DEFAULT 0,
                other_allowances REAL DEFAULT 0,
                total_salary REAL DEFAULT 0,
                bank_name TEXT,
                bank_account TEXT,
                bank_iban TEXT,
                bank_swift TEXT,
                payment_method TEXT DEFAULT 'bank_transfer',
                gosi_number TEXT,
                gosi_eligible INTEGER DEFAULT 1,
                gosi_registered_date DATE,
                tax_number TEXT,
                is_saudi INTEGER DEFAULT 0,
                is_active INTEGER DEFAULT 1,
                is_on_leave INTEGER DEFAULT 0,
                notes TEXT,
                photo_path TEXT,
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_employees_company ON employees(company_id)",
            "CREATE INDEX IF NOT EXISTS idx_employees_emp_number ON employees(emp_number)",
            "CREATE INDEX IF NOT EXISTS idx_employees_department ON employees(department_id)",
            "CREATE INDEX IF NOT EXISTS idx_employees_branch ON employees(branch_id)",
            "CREATE INDEX IF NOT EXISTS idx_employees_iqama ON employees(iqama_number)",
            "CREATE INDEX IF NOT EXISTS idx_employees_active ON employees(is_active)",
            "CREATE INDEX IF NOT EXISTS idx_employees_nationality ON employees(nationality)",

            """
            CREATE TABLE IF NOT EXISTS employee_documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
                document_type TEXT NOT NULL,
                document_number TEXT,
                issue_date DATE,
                expiry_date DATE,
                issuing_authority TEXT,
                file_path TEXT,
                file_name TEXT,
                file_size INTEGER,
                notes TEXT,
                is_verified INTEGER DEFAULT 0,
                verified_by INTEGER REFERENCES users(id),
                verified_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_emp_docs_employee ON employee_documents(employee_id)",
            "CREATE INDEX IF NOT EXISTS idx_emp_docs_expiry ON employee_documents(expiry_date)",

            """
            CREATE TABLE IF NOT EXISTS employee_dependents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
                full_name TEXT NOT NULL,
                relationship TEXT NOT NULL,
                date_of_birth DATE,
                nationality TEXT,
                id_number TEXT,
                gender TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS employee_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
                change_type TEXT NOT NULL,
                field_name TEXT,
                old_value TEXT,
                new_value TEXT,
                effective_date DATE,
                reason TEXT,
                approved_by INTEGER REFERENCES users(id),
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_emp_history_employee ON employee_history(employee_id)",

            """
            CREATE TABLE IF NOT EXISTS employee_assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                contract_id INTEGER REFERENCES contracts(id),
                client_id INTEGER REFERENCES clients(id),
                assignment_date DATE,
                end_date DATE,
                bill_rate REAL,
                pay_rate REAL,
                location TEXT,
                position TEXT,
                notes TEXT,
                is_active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_emp_assignments_employee ON employee_assignments(employee_id)",

            # =========================================================
            # CONTRACTS
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS contracts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                client_id INTEGER NOT NULL REFERENCES clients(id),
                contract_number TEXT UNIQUE,
                title TEXT NOT NULL,
                description TEXT,
                contract_type TEXT DEFAULT 'manpower_supply',
                status TEXT DEFAULT 'draft',
                start_date DATE,
                end_date DATE,
                auto_renew INTEGER DEFAULT 0,
                renewal_period_months INTEGER DEFAULT 12,
                notice_period_days INTEGER DEFAULT 30,
                payment_terms INTEGER DEFAULT 30,
                currency TEXT DEFAULT 'SAR',
                vat_applicable INTEGER DEFAULT 1,
                total_value REAL DEFAULT 0,
                signed_date DATE,
                signed_by TEXT,
                document_path TEXT,
                account_manager_id INTEGER REFERENCES users(id),
                notes TEXT,
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_contracts_company ON contracts(company_id)",
            "CREATE INDEX IF NOT EXISTS idx_contracts_client ON contracts(client_id)",
            "CREATE INDEX IF NOT EXISTS idx_contracts_status ON contracts(status)",

            """
            CREATE TABLE IF NOT EXISTS contract_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                contract_id INTEGER NOT NULL REFERENCES contracts(id) ON DELETE CASCADE,
                description TEXT NOT NULL,
                quantity REAL DEFAULT 1,
                unit TEXT DEFAULT 'person',
                unit_price REAL DEFAULT 0,
                discount_percent REAL DEFAULT 0,
                total_price REAL DEFAULT 0,
                vat_rate REAL DEFAULT 0.15,
                vat_amount REAL DEFAULT 0,
                line_total REAL DEFAULT 0,
                sort_order INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_contract_lines_contract ON contract_lines(contract_id)",

            """
            CREATE TABLE IF NOT EXISTS contract_amendments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                contract_id INTEGER NOT NULL REFERENCES contracts(id),
                amendment_number TEXT,
                effective_date DATE,
                description TEXT,
                changes_json TEXT,
                approved_by TEXT,
                document_path TEXT,
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            # =========================================================
            # PAYROLL
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS payroll_periods (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                period_name TEXT NOT NULL,
                period_year INTEGER NOT NULL,
                period_month INTEGER NOT NULL,
                start_date DATE NOT NULL,
                end_date DATE NOT NULL,
                payment_date DATE,
                status TEXT DEFAULT 'open',
                total_employees INTEGER DEFAULT 0,
                total_gross REAL DEFAULT 0,
                total_deductions REAL DEFAULT 0,
                total_net REAL DEFAULT 0,
                total_gosi_employee REAL DEFAULT 0,
                total_gosi_employer REAL DEFAULT 0,
                closed_by INTEGER REFERENCES users(id),
                closed_at TIMESTAMP,
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_payroll_periods_company ON payroll_periods(company_id)",
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_payroll_periods_unique ON payroll_periods(company_id, period_year, period_month)",

            """
            CREATE TABLE IF NOT EXISTS payroll_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                period_id INTEGER NOT NULL REFERENCES payroll_periods(id),
                run_number INTEGER DEFAULT 1,
                run_type TEXT DEFAULT 'regular',
                status TEXT DEFAULT 'draft',
                total_employees INTEGER DEFAULT 0,
                total_gross REAL DEFAULT 0,
                total_deductions REAL DEFAULT 0,
                total_net REAL DEFAULT 0,
                run_by INTEGER REFERENCES users(id),
                run_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                posted_by INTEGER REFERENCES users(id),
                posted_at TIMESTAMP,
                notes TEXT
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_payroll_runs_period ON payroll_runs(period_id)",

            """
            CREATE TABLE IF NOT EXISTS payroll_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                payroll_run_id INTEGER NOT NULL REFERENCES payroll_runs(id),
                period_id INTEGER NOT NULL REFERENCES payroll_periods(id),
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                period_year INTEGER NOT NULL,
                period_month INTEGER NOT NULL,
                working_days INTEGER DEFAULT 30,
                actual_days INTEGER DEFAULT 30,
                basic_salary REAL DEFAULT 0,
                housing_allowance REAL DEFAULT 0,
                transport_allowance REAL DEFAULT 0,
                food_allowance REAL DEFAULT 0,
                phone_allowance REAL DEFAULT 0,
                other_allowances REAL DEFAULT 0,
                overtime_amount REAL DEFAULT 0,
                bonus_amount REAL DEFAULT 0,
                commission_amount REAL DEFAULT 0,
                gross_salary REAL DEFAULT 0,
                gosi_employee REAL DEFAULT 0,
                gosi_employer REAL DEFAULT 0,
                gosi_base REAL DEFAULT 0,
                loan_deduction REAL DEFAULT 0,
                advance_deduction REAL DEFAULT 0,
                absence_deduction REAL DEFAULT 0,
                other_deductions REAL DEFAULT 0,
                total_deductions REAL DEFAULT 0,
                net_salary REAL DEFAULT 0,
                payment_method TEXT DEFAULT 'bank_transfer',
                bank_iban TEXT,
                status TEXT DEFAULT 'draft',
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_payroll_entries_run ON payroll_entries(payroll_run_id)",
            "CREATE INDEX IF NOT EXISTS idx_payroll_entries_employee ON payroll_entries(employee_id)",
            "CREATE INDEX IF NOT EXISTS idx_payroll_entries_period ON payroll_entries(period_id)",

            """
            CREATE TABLE IF NOT EXISTS payroll_adjustments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                period_id INTEGER NOT NULL REFERENCES payroll_periods(id),
                adjustment_type TEXT NOT NULL,
                description TEXT,
                amount REAL NOT NULL,
                is_addition INTEGER DEFAULT 1,
                is_recurring INTEGER DEFAULT 0,
                recurring_months INTEGER DEFAULT 0,
                start_period_id INTEGER REFERENCES payroll_periods(id),
                end_period_id INTEGER REFERENCES payroll_periods(id),
                approved_by INTEGER REFERENCES users(id),
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            # =========================================================
            # LOANS & ADVANCES
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS employee_loans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                loan_number TEXT UNIQUE,
                loan_type TEXT DEFAULT 'personal',
                amount REAL NOT NULL,
                outstanding_amount REAL,
                installment_amount REAL,
                number_of_installments INTEGER,
                start_period_id INTEGER REFERENCES payroll_periods(id),
                disbursement_date DATE,
                disbursement_method TEXT DEFAULT 'bank_transfer',
                interest_rate REAL DEFAULT 0,
                reason TEXT,
                status TEXT DEFAULT 'pending',
                approved_by INTEGER REFERENCES users(id),
                approved_at TIMESTAMP,
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_loans_employee ON employee_loans(employee_id)",

            """
            CREATE TABLE IF NOT EXISTS employee_advances (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                amount REAL NOT NULL,
                advance_date DATE DEFAULT CURRENT_DATE,
                reason TEXT,
                status TEXT DEFAULT 'pending',
                deduction_period_id INTEGER REFERENCES payroll_periods(id),
                approved_by INTEGER REFERENCES users(id),
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS loan_repayments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                loan_id INTEGER NOT NULL REFERENCES employee_loans(id),
                period_id INTEGER NOT NULL REFERENCES payroll_periods(id),
                amount REAL NOT NULL,
                payment_date DATE,
                payroll_entry_id INTEGER REFERENCES payroll_entries(id),
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            # =========================================================
            # LEAVE MANAGEMENT
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS leave_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                code TEXT NOT NULL,
                name TEXT NOT NULL,
                name_ar TEXT,
                days_per_year REAL DEFAULT 0,
                is_paid INTEGER DEFAULT 1,
                is_carry_forward INTEGER DEFAULT 0,
                max_carry_forward_days REAL DEFAULT 0,
                requires_approval INTEGER DEFAULT 1,
                min_advance_days INTEGER DEFAULT 0,
                max_consecutive_days INTEGER DEFAULT 0,
                applicable_gender TEXT DEFAULT 'all',
                is_active INTEGER DEFAULT 1,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS leave_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                leave_type_id INTEGER NOT NULL REFERENCES leave_types(id),
                start_date DATE NOT NULL,
                end_date DATE NOT NULL,
                days_requested REAL,
                reason TEXT,
                status TEXT DEFAULT 'pending',
                applied_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                approved_by INTEGER REFERENCES users(id),
                approved_at TIMESTAMP,
                rejection_reason TEXT,
                replacement_employee_id INTEGER REFERENCES employees(id),
                return_date DATE,
                actual_return_date DATE,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_leave_requests_employee ON leave_requests(employee_id)",
            "CREATE INDEX IF NOT EXISTS idx_leave_requests_status ON leave_requests(status)",

            """
            CREATE TABLE IF NOT EXISTS leave_balances (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                leave_type_id INTEGER NOT NULL REFERENCES leave_types(id),
                year INTEGER NOT NULL,
                opening_balance REAL DEFAULT 0,
                accrued REAL DEFAULT 0,
                used REAL DEFAULT 0,
                carried_forward REAL DEFAULT 0,
                closing_balance REAL DEFAULT 0,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(employee_id, leave_type_id, year)
            )
            """,

            # =========================================================
            # ATTENDANCE & TIMESHEETS
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS timesheets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                period_id INTEGER REFERENCES payroll_periods(id),
                week_start DATE,
                week_end DATE,
                total_hours REAL DEFAULT 0,
                overtime_hours REAL DEFAULT 0,
                status TEXT DEFAULT 'draft',
                submitted_at TIMESTAMP,
                approved_by INTEGER REFERENCES users(id),
                approved_at TIMESTAMP,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS attendance_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id),
                attendance_date DATE NOT NULL,
                check_in TIME,
                check_out TIME,
                break_minutes INTEGER DEFAULT 0,
                total_hours REAL DEFAULT 0,
                overtime_hours REAL DEFAULT 0,
                status TEXT DEFAULT 'present',
                location TEXT,
                notes TEXT,
                source TEXT DEFAULT 'manual',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(employee_id, attendance_date)
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_attendance_employee ON attendance_records(employee_id)",
            "CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance_records(attendance_date)",

            # =========================================================
            # ACCOUNTING
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS chart_of_accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                account_code TEXT NOT NULL,
                account_name TEXT NOT NULL,
                account_name_ar TEXT,
                account_type TEXT NOT NULL,
                account_subtype TEXT,
                parent_id INTEGER REFERENCES chart_of_accounts(id),
                level INTEGER DEFAULT 1,
                is_header INTEGER DEFAULT 0,
                is_active INTEGER DEFAULT 1,
                allow_posting INTEGER DEFAULT 1,
                normal_balance TEXT DEFAULT 'debit',
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(company_id, account_code)
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_coa_company ON chart_of_accounts(company_id)",
            "CREATE INDEX IF NOT EXISTS idx_coa_code ON chart_of_accounts(account_code)",
            "CREATE INDEX IF NOT EXISTS idx_coa_type ON chart_of_accounts(account_type)",

            """
            CREATE TABLE IF NOT EXISTS fiscal_periods (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                name TEXT NOT NULL,
                fiscal_year INTEGER NOT NULL,
                period_number INTEGER NOT NULL,
                start_date DATE NOT NULL,
                end_date DATE NOT NULL,
                status TEXT DEFAULT 'open',
                closed_by INTEGER REFERENCES users(id),
                closed_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(company_id, fiscal_year, period_number)
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS journal_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                entry_number TEXT UNIQUE,
                entry_date DATE NOT NULL,
                description TEXT,
                reference TEXT,
                source_type TEXT,
                source_id INTEGER,
                status TEXT DEFAULT 'draft',
                total_debit REAL DEFAULT 0,
                total_credit REAL DEFAULT 0,
                fiscal_period_id INTEGER REFERENCES fiscal_periods(id),
                is_recurring INTEGER DEFAULT 0,
                recurring_template_id INTEGER,
                posted_by INTEGER REFERENCES users(id),
                posted_at TIMESTAMP,
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_journal_entries_company ON journal_entries(company_id)",
            "CREATE INDEX IF NOT EXISTS idx_journal_entries_date ON journal_entries(entry_date)",
            "CREATE INDEX IF NOT EXISTS idx_journal_entries_status ON journal_entries(status)",

            """
            CREATE TABLE IF NOT EXISTS journal_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                journal_entry_id INTEGER NOT NULL REFERENCES journal_entries(id) ON DELETE CASCADE,
                account_id INTEGER NOT NULL REFERENCES chart_of_accounts(id),
                description TEXT,
                debit REAL DEFAULT 0,
                credit REAL DEFAULT 0,
                cost_center_id INTEGER REFERENCES cost_centers(id),
                department_id INTEGER REFERENCES departments(id),
                employee_id INTEGER REFERENCES employees(id),
                reference TEXT,
                sort_order INTEGER DEFAULT 0
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_journal_lines_entry ON journal_lines(journal_entry_id)",
            "CREATE INDEX IF NOT EXISTS idx_journal_lines_account ON journal_lines(account_id)",

            # =========================================================
            # BANKING
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS bank_accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                account_name TEXT NOT NULL,
                bank_name TEXT NOT NULL,
                account_number TEXT NOT NULL,
                iban TEXT UNIQUE,
                swift_code TEXT,
                currency TEXT DEFAULT 'SAR',
                account_type TEXT DEFAULT 'current',
                gl_account_id INTEGER REFERENCES chart_of_accounts(id),
                opening_balance REAL DEFAULT 0,
                current_balance REAL DEFAULT 0,
                overdraft_limit REAL DEFAULT 0,
                branch_name TEXT,
                is_active INTEGER DEFAULT 1,
                is_wps_account INTEGER DEFAULT 0,
                wps_bank_code TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS bank_transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bank_account_id INTEGER NOT NULL REFERENCES bank_accounts(id),
                transaction_date DATE NOT NULL,
                value_date DATE,
                description TEXT,
                reference TEXT,
                transaction_type TEXT,
                amount REAL NOT NULL,
                balance REAL,
                is_reconciled INTEGER DEFAULT 0,
                reconciliation_id INTEGER REFERENCES bank_reconciliation(id),
                journal_entry_id INTEGER REFERENCES journal_entries(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_bank_trans_account ON bank_transactions(bank_account_id)",
            "CREATE INDEX IF NOT EXISTS idx_bank_trans_date ON bank_transactions(transaction_date)",

            """
            CREATE TABLE IF NOT EXISTS bank_reconciliation (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                bank_account_id INTEGER NOT NULL REFERENCES bank_accounts(id),
                statement_date DATE NOT NULL,
                statement_balance REAL NOT NULL,
                book_balance REAL NOT NULL,
                outstanding_deposits REAL DEFAULT 0,
                outstanding_payments REAL DEFAULT 0,
                difference REAL DEFAULT 0,
                status TEXT DEFAULT 'in_progress',
                reconciled_by INTEGER REFERENCES users(id),
                reconciled_at TIMESTAMP,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            # =========================================================
            # INVOICING & PAYMENTS
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                invoice_number TEXT UNIQUE,
                invoice_type TEXT DEFAULT 'standard',
                client_id INTEGER NOT NULL REFERENCES clients(id),
                contract_id INTEGER REFERENCES contracts(id),
                invoice_date DATE NOT NULL,
                due_date DATE,
                supply_date DATE,
                billing_period_start DATE,
                billing_period_end DATE,
                currency TEXT DEFAULT 'SAR',
                subtotal REAL DEFAULT 0,
                discount_amount REAL DEFAULT 0,
                vat_rate REAL DEFAULT 0.15,
                vat_amount REAL DEFAULT 0,
                total_amount REAL DEFAULT 0,
                paid_amount REAL DEFAULT 0,
                outstanding_amount REAL DEFAULT 0,
                status TEXT DEFAULT 'draft',
                payment_terms INTEGER DEFAULT 30,
                bank_account_id INTEGER REFERENCES bank_accounts(id),
                zatca_qr_code TEXT,
                zatca_uuid TEXT,
                zatca_hash TEXT,
                notes TEXT,
                terms_conditions TEXT,
                journal_entry_id INTEGER REFERENCES journal_entries(id),
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_invoices_company ON invoices(company_id)",
            "CREATE INDEX IF NOT EXISTS idx_invoices_client ON invoices(client_id)",
            "CREATE INDEX IF NOT EXISTS idx_invoices_status ON invoices(status)",
            "CREATE INDEX IF NOT EXISTS idx_invoices_date ON invoices(invoice_date)",

            """
            CREATE TABLE IF NOT EXISTS invoice_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER NOT NULL REFERENCES invoices(id) ON DELETE CASCADE,
                description TEXT NOT NULL,
                quantity REAL DEFAULT 1,
                unit TEXT DEFAULT 'person_month',
                unit_price REAL DEFAULT 0,
                discount_percent REAL DEFAULT 0,
                discount_amount REAL DEFAULT 0,
                subtotal REAL DEFAULT 0,
                vat_rate REAL DEFAULT 0.15,
                vat_amount REAL DEFAULT 0,
                line_total REAL DEFAULT 0,
                account_id INTEGER REFERENCES chart_of_accounts(id),
                employee_id INTEGER REFERENCES employees(id),
                sort_order INTEGER DEFAULT 0
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_invoice_lines_invoice ON invoice_lines(invoice_id)",

            """
            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                payment_number TEXT UNIQUE,
                client_id INTEGER REFERENCES clients(id),
                payment_date DATE NOT NULL,
                amount REAL NOT NULL,
                currency TEXT DEFAULT 'SAR',
                payment_method TEXT DEFAULT 'bank_transfer',
                reference TEXT,
                bank_account_id INTEGER REFERENCES bank_accounts(id),
                invoice_id INTEGER REFERENCES invoices(id),
                notes TEXT,
                status TEXT DEFAULT 'posted',
                journal_entry_id INTEGER REFERENCES journal_entries(id),
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_payments_invoice ON payments(invoice_id)",
            "CREATE INDEX IF NOT EXISTS idx_payments_client ON payments(client_id)",

            """
            CREATE TABLE IF NOT EXISTS credit_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                credit_note_number TEXT UNIQUE,
                client_id INTEGER NOT NULL REFERENCES clients(id),
                original_invoice_id INTEGER REFERENCES invoices(id),
                credit_note_date DATE NOT NULL,
                reason TEXT,
                subtotal REAL DEFAULT 0,
                vat_amount REAL DEFAULT 0,
                total_amount REAL DEFAULT 0,
                status TEXT DEFAULT 'draft',
                journal_entry_id INTEGER REFERENCES journal_entries(id),
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS credit_note_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                credit_note_id INTEGER NOT NULL REFERENCES credit_notes(id) ON DELETE CASCADE,
                description TEXT NOT NULL,
                quantity REAL DEFAULT 1,
                unit_price REAL DEFAULT 0,
                subtotal REAL DEFAULT 0,
                vat_rate REAL DEFAULT 0.15,
                vat_amount REAL DEFAULT 0,
                line_total REAL DEFAULT 0,
                sort_order INTEGER DEFAULT 0
            )
            """,

            # =========================================================
            # ASSETS
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS assets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                asset_number TEXT UNIQUE,
                asset_name TEXT NOT NULL,
                asset_name_ar TEXT,
                category TEXT,
                subcategory TEXT,
                description TEXT,
                serial_number TEXT,
                model TEXT,
                manufacturer TEXT,
                supplier TEXT,
                purchase_date DATE,
                purchase_price REAL DEFAULT 0,
                salvage_value REAL DEFAULT 0,
                useful_life_months INTEGER DEFAULT 60,
                depreciation_method TEXT DEFAULT 'straight_line',
                depreciation_rate REAL DEFAULT 0,
                accumulated_depreciation REAL DEFAULT 0,
                net_book_value REAL DEFAULT 0,
                department_id INTEGER REFERENCES departments(id),
                branch_id INTEGER REFERENCES branches(id),
                assigned_to INTEGER REFERENCES employees(id),
                location TEXT,
                condition TEXT DEFAULT 'good',
                is_active INTEGER DEFAULT 1,
                disposal_date DATE,
                disposal_reason TEXT,
                disposal_amount REAL DEFAULT 0,
                asset_account_id INTEGER REFERENCES chart_of_accounts(id),
                depreciation_account_id INTEGER REFERENCES chart_of_accounts(id),
                accumulated_dep_account_id INTEGER REFERENCES chart_of_accounts(id),
                warranty_expiry DATE,
                notes TEXT,
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_assets_company ON assets(company_id)",
            "CREATE INDEX IF NOT EXISTS idx_assets_category ON assets(category)",

            """
            CREATE TABLE IF NOT EXISTS asset_depreciation (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL REFERENCES assets(id),
                period TEXT NOT NULL,
                depreciation_date DATE NOT NULL,
                amount REAL NOT NULL,
                accumulated_amount REAL NOT NULL,
                net_book_value REAL NOT NULL,
                journal_entry_id INTEGER REFERENCES journal_entries(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(asset_id, period)
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS asset_maintenance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_id INTEGER NOT NULL REFERENCES assets(id),
                maintenance_type TEXT DEFAULT 'preventive',
                maintenance_date DATE NOT NULL,
                description TEXT,
                cost REAL DEFAULT 0,
                vendor TEXT,
                performed_by TEXT,
                next_maintenance_date DATE,
                documents_path TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            # =========================================================
            # COMPLIANCE (GOSI, WPS, ZATCA, VAT)
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS gosi_submissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                period_id INTEGER NOT NULL REFERENCES payroll_periods(id),
                submission_date DATE,
                total_employees INTEGER DEFAULT 0,
                total_saudi INTEGER DEFAULT 0,
                total_non_saudi INTEGER DEFAULT 0,
                total_employee_contribution REAL DEFAULT 0,
                total_employer_contribution REAL DEFAULT 0,
                total_amount REAL DEFAULT 0,
                status TEXT DEFAULT 'pending',
                reference_number TEXT,
                file_path TEXT,
                submitted_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS wps_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                payroll_run_id INTEGER NOT NULL REFERENCES payroll_runs(id),
                file_name TEXT,
                file_content TEXT,
                file_path TEXT,
                employer_id TEXT,
                pay_date DATE,
                total_records INTEGER DEFAULT 0,
                total_amount REAL DEFAULT 0,
                status TEXT DEFAULT 'generated',
                submission_date DATE,
                reference_number TEXT,
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS zatca_invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER NOT NULL REFERENCES invoices(id),
                invoice_uuid TEXT UNIQUE,
                invoice_hash TEXT,
                qr_code TEXT,
                xml_content TEXT,
                submission_id TEXT,
                status TEXT DEFAULT 'pending',
                submitted_at TIMESTAMP,
                clearance_status TEXT,
                reporting_status TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS vat_returns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL REFERENCES companies(id),
                period_start DATE NOT NULL,
                period_end DATE NOT NULL,
                return_type TEXT DEFAULT 'quarterly',
                total_taxable_sales REAL DEFAULT 0,
                total_vat_on_sales REAL DEFAULT 0,
                total_taxable_purchases REAL DEFAULT 0,
                total_vat_on_purchases REAL DEFAULT 0,
                net_vat_due REAL DEFAULT 0,
                status TEXT DEFAULT 'draft',
                filing_date DATE,
                payment_date DATE,
                reference_number TEXT,
                notes TEXT,
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS vat_return_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vat_return_id INTEGER NOT NULL REFERENCES vat_returns(id),
                line_type TEXT NOT NULL,
                description TEXT,
                taxable_amount REAL DEFAULT 0,
                vat_amount REAL DEFAULT 0,
                adjustment_amount REAL DEFAULT 0,
                sort_order INTEGER DEFAULT 0
            )
            """,

            # =========================================================
            # SYSTEM TABLES
            # =========================================================
            """
            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id),
                title TEXT NOT NULL,
                message TEXT,
                notification_type TEXT DEFAULT 'info',
                link TEXT,
                is_read INTEGER DEFAULT 0,
                read_at TIMESTAMP,
                expires_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_notifications_user ON notifications(user_id)",
            "CREATE INDEX IF NOT EXISTS idx_notifications_read ON notifications(is_read)",

            """
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER REFERENCES users(id),
                username TEXT,
                action TEXT NOT NULL,
                table_name TEXT,
                record_id INTEGER,
                old_values TEXT,
                new_values TEXT,
                ip_address TEXT,
                user_agent TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            "CREATE INDEX IF NOT EXISTS idx_audit_logs_user ON audit_logs(user_id)",
            "CREATE INDEX IF NOT EXISTS idx_audit_logs_table ON audit_logs(table_name)",
            "CREATE INDEX IF NOT EXISTS idx_audit_logs_date ON audit_logs(created_at)",

            """
            CREATE TABLE IF NOT EXISTS sequences (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                sequence_name TEXT NOT NULL,
                prefix TEXT,
                suffix TEXT,
                current_value INTEGER DEFAULT 0,
                step INTEGER DEFAULT 1,
                padding INTEGER DEFAULT 5,
                UNIQUE(company_id, sequence_name)
            )
            """,

            """
            CREATE TABLE IF NOT EXISTS settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER REFERENCES companies(id),
                setting_key TEXT NOT NULL,
                setting_value TEXT,
                setting_type TEXT DEFAULT 'string',
                description TEXT,
                is_sensitive INTEGER DEFAULT 0,
                updated_by INTEGER REFERENCES users(id),
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(company_id, setting_key)
            )
            """,

            # Triggers for updated_at
            """
            CREATE TRIGGER IF NOT EXISTS update_companies_timestamp
            AFTER UPDATE ON companies
            BEGIN
                UPDATE companies SET updated_at = CURRENT_TIMESTAMP WHERE id = NEW.id;
            END
            """,

            """
            CREATE TRIGGER IF NOT EXISTS update_employees_timestamp
            AFTER UPDATE ON employees
            BEGIN
                UPDATE employees SET updated_at = CURRENT_TIMESTAMP WHERE id = NEW.id;
            END
            """,

            """
            CREATE TRIGGER IF NOT EXISTS update_invoices_timestamp
            AFTER UPDATE ON invoices
            BEGIN
                UPDATE invoices SET updated_at = CURRENT_TIMESTAMP WHERE id = NEW.id;
            END
            """,
        ]

        for sql in tables_sql:
            try:
                self.conn.execute(sql)
            except sqlite3.Error as e:
                logger.error(f"Error creating table/index: {e}\nSQL: {sql[:100]}")
                raise

        self.conn.commit()
        logger.info("Database tables created successfully")

        # Insert schema version
        current_version = self.get_schema_version()
        if current_version < self._version:
            self.execute(
                "INSERT INTO schema_version (version, description) VALUES (?, ?)",
                (self._version, "Initial schema")
            )

    def get_next_sequence(self, company_id: int, sequence_name: str) -> str:
        """Get the next value in a sequence."""
        with self._lock:
            seq = self.fetchone(
                "SELECT * FROM sequences WHERE company_id = ? AND sequence_name = ?",
                (company_id, sequence_name)
            )
            if not seq:
                # Create sequence
                defaults = {
                    'employee': ('EMP', '', 1, 1, 5),
                    'contract': ('CNT', '', 1, 1, 5),
                    'invoice': ('INV', '', 1, 1, 6),
                    'payment': ('PAY', '', 1, 1, 6),
                    'journal': ('JE', '', 1, 1, 6),
                    'loan': ('LN', '', 1, 1, 5),
                    'asset': ('AST', '', 1, 1, 5),
                    'client': ('CLI', '', 1, 1, 5),
                    'credit_note': ('CN', '', 1, 1, 6),
                    'payroll_run': ('PR', '', 1, 1, 6),
                }
                prefix, suffix, current, step, padding = defaults.get(
                    sequence_name, ('', '', 1, 1, 6)
                )
                self.execute(
                    """INSERT INTO sequences
                       (company_id, sequence_name, prefix, suffix, current_value, step, padding)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (company_id, sequence_name, prefix, suffix, current, step, padding)
                )
                seq = self.fetchone(
                    "SELECT * FROM sequences WHERE company_id = ? AND sequence_name = ?",
                    (company_id, sequence_name)
                )

            next_val = seq['current_value'] + seq['step']
            self.execute(
                "UPDATE sequences SET current_value = ? WHERE id = ?",
                (next_val, seq['id'])
            )

            number = str(next_val).zfill(seq['padding'])
            year = datetime.date.today().year
            return f"{seq['prefix']}{year}{number}{seq['suffix']}"

    def backup(self, backup_path: str = None) -> str:
        """Create a database backup."""
        if not backup_path:
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_dir = Path(self.db_path).parent / "backups"
            backup_dir.mkdir(exist_ok=True)
            backup_path = str(backup_dir / f"erp_backup_{timestamp}.db")

        import shutil
        shutil.copy2(self.db_path, backup_path)
        logger.info(f"Database backed up to: {backup_path}")
        return backup_path

    def vacuum(self):
        """Run VACUUM to optimize the database."""
        self.conn.execute("PRAGMA incremental_vacuum(100)")
        logger.info("Database vacuumed")


# =============================================================================
# SECTION 4: ORM-STYLE MODEL CLASSES
# =============================================================================

class BaseModel:
    """
    Base model class providing common functionality for all ERP models.
    Supports serialization, deserialization, and basic validation.
    """
    _table = ""
    _fields: List[str] = []

    def __init__(self, **kwargs):
        for key, value in kwargs.items():
            setattr(self, key, value)

    def to_dict(self) -> dict:
        """Convert model to dictionary."""
        result = {}
        for key, value in self.__dict__.items():
            if not key.startswith('_'):
                if isinstance(value, (datetime.date, datetime.datetime)):
                    result[key] = value.isoformat()
                elif isinstance(value, Decimal):
                    result[key] = float(value)
                else:
                    result[key] = value
        return result

    @classmethod
    def from_dict(cls, data: dict):
        """Create model from dictionary."""
        instance = cls()
        for key, value in data.items():
            setattr(instance, key, value)
        return instance

    def validate(self) -> List[str]:
        """Validate model and return list of errors."""
        return []

    def is_valid(self) -> bool:
        """Check if model is valid."""
        return len(self.validate()) == 0

    def __repr__(self):
        return f"{self.__class__.__name__}({self.to_dict()})"


class Company(BaseModel):
    """Company model."""
    _table = "companies"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.name: str = ""
        self.name_ar: str = ""
        self.cr_number: str = ""
        self.vat_number: str = ""
        self.address: str = ""
        self.city: str = ""
        self.country: str = "Saudi Arabia"
        self.phone: str = ""
        self.email: str = ""
        self.currency: str = "SAR"
        self.default_vat_rate: float = 0.15
        self.is_active: int = 1
        super().__init__(**kwargs)

    def validate(self) -> List[str]:
        errors = []
        if not self.name:
            errors.append("Company name is required")
        if not self.cr_number:
            errors.append("CR number is required")
        return errors


class Branch(BaseModel):
    """Branch model."""
    _table = "branches"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.company_id: Optional[int] = None
        self.name: str = ""
        self.name_ar: str = ""
        self.code: str = ""
        self.manager_name: str = ""
        self.address: str = ""
        self.city: str = ""
        self.phone: str = ""
        self.is_active: int = 1
        self.is_head_office: int = 0
        super().__init__(**kwargs)


class Department(BaseModel):
    """Department model."""
    _table = "departments"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.company_id: Optional[int] = None
        self.branch_id: Optional[int] = None
        self.parent_id: Optional[int] = None
        self.name: str = ""
        self.name_ar: str = ""
        self.code: str = ""
        self.manager_id: Optional[int] = None
        self.is_active: int = 1
        super().__init__(**kwargs)


class CostCenter(BaseModel):
    """Cost Center model."""
    _table = "cost_centers"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.company_id: Optional[int] = None
        self.code: str = ""
        self.name: str = ""
        self.budget: float = 0.0
        self.is_active: int = 1
        super().__init__(**kwargs)


class User(BaseModel):
    """User model."""
    _table = "users"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.username: str = ""
        self.password_hash: str = ""
        self.email: str = ""
        self.full_name: str = ""
        self.company_id: Optional[int] = None
        self.is_active: int = 1
        self.is_admin: int = 0
        self.language: str = "en"
        super().__init__(**kwargs)

    def validate(self) -> List[str]:
        errors = []
        if not self.username:
            errors.append("Username is required")
        if len(self.username) < 3:
            errors.append("Username must be at least 3 characters")
        return errors


class Role(BaseModel):
    """Role model."""
    _table = "roles"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.name: str = ""
        self.display_name: str = ""
        self.description: str = ""
        self.is_system: int = 0
        self.is_active: int = 1
        super().__init__(**kwargs)


class Client(BaseModel):
    """Client model."""
    _table = "clients"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.company_id: Optional[int] = None
        self.client_number: str = ""
        self.name: str = ""
        self.name_ar: str = ""
        self.cr_number: str = ""
        self.vat_number: str = ""
        self.industry: str = ""
        self.client_type: str = "corporate"
        self.address: str = ""
        self.city: str = ""
        self.country: str = "Saudi Arabia"
        self.phone: str = ""
        self.email: str = ""
        self.payment_terms: int = 30
        self.credit_limit: float = 0.0
        self.status: str = "active"
        self.notes: str = ""
        super().__init__(**kwargs)

    def validate(self) -> List[str]:
        errors = []
        if not self.name:
            errors.append("Client name is required")
        return errors


class Employee(BaseModel):
    """
    Employee model with all 50+ fields per Saudi labor requirements.
    """
    _table = "employees"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.emp_number: str = ""
        self.first_name: str = ""
        self.last_name: str = ""
        self.middle_name: str = ""
        self.arabic_name: str = ""
        self.nationality: str = ""
        self.id_type: str = "iqama"
        self.id_number: str = ""
        self.iqama_number: str = ""
        self.iqama_expiry: Optional[str] = None
        self.iqama_issue_place: str = ""
        self.passport_number: str = ""
        self.passport_expiry: Optional[str] = None
        self.passport_issue_country: str = ""
        self.visa_number: str = ""
        self.visa_type: str = ""
        self.visa_expiry: Optional[str] = None
        self.border_number: str = ""
        self.date_of_birth: Optional[str] = None
        self.gender: str = "male"
        self.marital_status: str = "single"
        self.religion: str = ""
        self.education: str = ""
        self.mobile: str = ""
        self.email: str = ""
        self.emergency_contact: str = ""
        self.emergency_phone: str = ""
        self.current_address: str = ""
        self.hire_date: Optional[str] = None
        self.termination_date: Optional[str] = None
        self.termination_reason: str = ""
        self.job_title: str = ""
        self.job_title_ar: str = ""
        self.department_id: Optional[int] = None
        self.branch_id: Optional[int] = None
        self.cost_center_id: Optional[int] = None
        self.company_id: Optional[int] = None
        self.direct_manager_id: Optional[int] = None
        self.employment_type: str = "full_time"
        self.contract_type: str = "indefinite"
        self.salary_type: str = "monthly"
        self.basic_salary: float = 0.0
        self.housing_allowance: float = 0.0
        self.transport_allowance: float = 0.0
        self.food_allowance: float = 0.0
        self.phone_allowance: float = 0.0
        self.other_allowances: float = 0.0
        self.total_salary: float = 0.0
        self.bank_name: str = ""
        self.bank_account: str = ""
        self.bank_iban: str = ""
        self.payment_method: str = "bank_transfer"
        self.gosi_number: str = ""
        self.gosi_eligible: int = 1
        self.tax_number: str = ""
        self.is_saudi: int = 0
        self.is_active: int = 1
        self.is_on_leave: int = 0
        self.notes: str = ""
        self.photo_path: str = ""
        super().__init__(**kwargs)

    def validate(self) -> List[str]:
        errors = []
        if not self.first_name:
            errors.append("First name is required")
        if not self.nationality:
            errors.append("Nationality is required")
        if not self.hire_date:
            errors.append("Hire date is required")
        if self.basic_salary < 0:
            errors.append("Basic salary cannot be negative")
        return errors

    def get_full_name(self) -> str:
        """Get employee full name."""
        parts = [self.first_name]
        if self.middle_name:
            parts.append(self.middle_name)
        if self.last_name:
            parts.append(self.last_name)
        return " ".join(parts)

    def calculate_total_salary(self) -> float:
        """Calculate total monthly salary."""
        return (
            self.basic_salary + self.housing_allowance +
            self.transport_allowance + self.food_allowance +
            self.phone_allowance + self.other_allowances
        )

    def get_years_of_service(self, as_of_date: str = None) -> float:
        """Calculate years of service."""
        if not self.hire_date:
            return 0.0
        hire = datetime.date.fromisoformat(self.hire_date)
        if as_of_date:
            end = datetime.date.fromisoformat(as_of_date)
        elif self.termination_date:
            end = datetime.date.fromisoformat(self.termination_date)
        else:
            end = datetime.date.today()
        delta = end - hire
        return delta.days / 365.25


class Contract(BaseModel):
    """Contract model."""
    _table = "contracts"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.company_id: Optional[int] = None
        self.client_id: Optional[int] = None
        self.contract_number: str = ""
        self.title: str = ""
        self.description: str = ""
        self.contract_type: str = "manpower_supply"
        self.status: str = "draft"
        self.start_date: Optional[str] = None
        self.end_date: Optional[str] = None
        self.payment_terms: int = 30
        self.currency: str = "SAR"
        self.vat_applicable: int = 1
        self.total_value: float = 0.0
        self.notes: str = ""
        super().__init__(**kwargs)


class ContractLine(BaseModel):
    """Contract line model."""
    _table = "contract_lines"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.contract_id: Optional[int] = None
        self.description: str = ""
        self.quantity: float = 1.0
        self.unit: str = "person"
        self.unit_price: float = 0.0
        self.discount_percent: float = 0.0
        self.total_price: float = 0.0
        self.vat_rate: float = 0.15
        self.vat_amount: float = 0.0
        self.line_total: float = 0.0
        super().__init__(**kwargs)

    def calculate(self):
        """Calculate line totals."""
        self.total_price = self.quantity * self.unit_price
        discount = self.total_price * (self.discount_percent / 100)
        net = self.total_price - discount
        self.vat_amount = net * self.vat_rate
        self.line_total = net + self.vat_amount


class PayrollPeriod(BaseModel):
    """Payroll period model."""
    _table = "payroll_periods"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.company_id: Optional[int] = None
        self.period_name: str = ""
        self.period_year: int = datetime.date.today().year
        self.period_month: int = datetime.date.today().month
        self.start_date: Optional[str] = None
        self.end_date: Optional[str] = None
        self.payment_date: Optional[str] = None
        self.status: str = "open"
        self.total_employees: int = 0
        self.total_gross: float = 0.0
        self.total_net: float = 0.0
        super().__init__(**kwargs)


class PayrollRun(BaseModel):
    """Payroll run model."""
    _table = "payroll_runs"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.period_id: Optional[int] = None
        self.run_number: int = 1
        self.run_type: str = "regular"
        self.status: str = "draft"
        self.total_employees: int = 0
        self.total_gross: float = 0.0
        self.total_net: float = 0.0
        super().__init__(**kwargs)


class PayrollEntry(BaseModel):
    """Payroll entry model."""
    _table = "payroll_entries"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.payroll_run_id: Optional[int] = None
        self.period_id: Optional[int] = None
        self.employee_id: Optional[int] = None
        self.period_year: int = 0
        self.period_month: int = 0
        self.working_days: int = 30
        self.actual_days: int = 30
        self.basic_salary: float = 0.0
        self.housing_allowance: float = 0.0
        self.transport_allowance: float = 0.0
        self.food_allowance: float = 0.0
        self.other_allowances: float = 0.0
        self.overtime_amount: float = 0.0
        self.bonus_amount: float = 0.0
        self.gross_salary: float = 0.0
        self.gosi_employee: float = 0.0
        self.gosi_employer: float = 0.0
        self.gosi_base: float = 0.0
        self.loan_deduction: float = 0.0
        self.advance_deduction: float = 0.0
        self.absence_deduction: float = 0.0
        self.other_deductions: float = 0.0
        self.total_deductions: float = 0.0
        self.net_salary: float = 0.0
        self.status: str = "draft"
        super().__init__(**kwargs)


class EmployeeLoan(BaseModel):
    """Employee loan model."""
    _table = "employee_loans"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.employee_id: Optional[int] = None
        self.loan_number: str = ""
        self.loan_type: str = "personal"
        self.amount: float = 0.0
        self.outstanding_amount: float = 0.0
        self.installment_amount: float = 0.0
        self.number_of_installments: int = 0
        self.reason: str = ""
        self.status: str = "pending"
        super().__init__(**kwargs)


class LeaveType(BaseModel):
    """Leave type model."""
    _table = "leave_types"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.company_id: Optional[int] = None
        self.code: str = ""
        self.name: str = ""
        self.name_ar: str = ""
        self.days_per_year: float = 0.0
        self.is_paid: int = 1
        self.is_carry_forward: int = 0
        self.max_carry_forward_days: float = 0.0
        self.requires_approval: int = 1
        self.applicable_gender: str = "all"
        self.is_active: int = 1
        super().__init__(**kwargs)


class LeaveRequest(BaseModel):
    """Leave request model."""
    _table = "leave_requests"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.employee_id: Optional[int] = None
        self.leave_type_id: Optional[int] = None
        self.start_date: Optional[str] = None
        self.end_date: Optional[str] = None
        self.days_requested: float = 0.0
        self.reason: str = ""
        self.status: str = "pending"
        super().__init__(**kwargs)

    def calculate_days(self) -> float:
        """Calculate number of leave days."""
        if self.start_date and self.end_date:
            start = datetime.date.fromisoformat(self.start_date)
            end = datetime.date.fromisoformat(self.end_date)
            return (end - start).days + 1
        return 0.0


class Timesheet(BaseModel):
    """Timesheet model."""
    _table = "timesheets"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.employee_id: Optional[int] = None
        self.period_id: Optional[int] = None
        self.week_start: Optional[str] = None
        self.week_end: Optional[str] = None
        self.total_hours: float = 0.0
        self.overtime_hours: float = 0.0
        self.status: str = "draft"
        super().__init__(**kwargs)


class ChartOfAccount(BaseModel):
    """Chart of Account model."""
    _table = "chart_of_accounts"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.company_id: Optional[int] = None
        self.account_code: str = ""
        self.account_name: str = ""
        self.account_name_ar: str = ""
        self.account_type: str = ""
        self.account_subtype: str = ""
        self.parent_id: Optional[int] = None
        self.level: int = 1
        self.is_header: int = 0
        self.is_active: int = 1
        self.allow_posting: int = 1
        self.normal_balance: str = "debit"
        super().__init__(**kwargs)


class JournalEntry(BaseModel):
    """Journal entry model."""
    _table = "journal_entries"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.company_id: Optional[int] = None
        self.entry_number: str = ""
        self.entry_date: Optional[str] = None
        self.description: str = ""
        self.reference: str = ""
        self.source_type: str = ""
        self.source_id: Optional[int] = None
        self.status: str = "draft"
        self.total_debit: float = 0.0
        self.total_credit: float = 0.0
        self.lines: List[dict] = []
        super().__init__(**kwargs)

    def is_balanced(self) -> bool:
        """Check if journal entry is balanced."""
        return abs(self.total_debit - self.total_credit) < 0.01


class Invoice(BaseModel):
    """Invoice model."""
    _table = "invoices"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.company_id: Optional[int] = None
        self.invoice_number: str = ""
        self.invoice_type: str = "standard"
        self.client_id: Optional[int] = None
        self.contract_id: Optional[int] = None
        self.invoice_date: Optional[str] = None
        self.due_date: Optional[str] = None
        self.currency: str = "SAR"
        self.subtotal: float = 0.0
        self.discount_amount: float = 0.0
        self.vat_rate: float = 0.15
        self.vat_amount: float = 0.0
        self.total_amount: float = 0.0
        self.paid_amount: float = 0.0
        self.outstanding_amount: float = 0.0
        self.status: str = "draft"
        self.payment_terms: int = 30
        self.zatca_qr_code: str = ""
        self.notes: str = ""
        self.lines: List[dict] = []
        super().__init__(**kwargs)

    def validate(self) -> List[str]:
        errors = []
        if not self.client_id:
            errors.append("Client is required")
        if not self.invoice_date:
            errors.append("Invoice date is required")
        if not self.lines:
            errors.append("Invoice must have at least one line")
        return errors

    def calculate_totals(self):
        """Recalculate invoice totals from lines."""
        self.subtotal = sum(
            line.get('subtotal', line.get('quantity', 1) * line.get('unit_price', 0))
            for line in self.lines
        )
        self.vat_amount = sum(line.get('vat_amount', 0) for line in self.lines)
        self.total_amount = self.subtotal + self.vat_amount - self.discount_amount
        self.outstanding_amount = self.total_amount - self.paid_amount


class Payment(BaseModel):
    """Payment model."""
    _table = "payments"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.company_id: Optional[int] = None
        self.payment_number: str = ""
        self.client_id: Optional[int] = None
        self.payment_date: Optional[str] = None
        self.amount: float = 0.0
        self.currency: str = "SAR"
        self.payment_method: str = "bank_transfer"
        self.reference: str = ""
        self.invoice_id: Optional[int] = None
        self.status: str = "posted"
        super().__init__(**kwargs)


class Asset(BaseModel):
    """Asset model."""
    _table = "assets"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.company_id: Optional[int] = None
        self.asset_number: str = ""
        self.asset_name: str = ""
        self.category: str = ""
        self.purchase_date: Optional[str] = None
        self.purchase_price: float = 0.0
        self.salvage_value: float = 0.0
        self.useful_life_months: int = 60
        self.depreciation_method: str = "straight_line"
        self.accumulated_depreciation: float = 0.0
        self.net_book_value: float = 0.0
        self.is_active: int = 1
        super().__init__(**kwargs)

    def calculate_monthly_depreciation(self) -> float:
        """Calculate monthly depreciation amount."""
        depreciable_amount = self.purchase_price - self.salvage_value
        if self.useful_life_months <= 0:
            return 0.0
        if self.depreciation_method == "straight_line":
            return depreciable_amount / self.useful_life_months
        elif self.depreciation_method == "declining_balance":
            rate = self.depreciation_method / self.useful_life_months
            return self.net_book_value * rate / 12
        return depreciable_amount / self.useful_life_months


class AuditLog(BaseModel):
    """Audit log model."""
    _table = "audit_logs"

    def __init__(self, **kwargs):
        self.id: Optional[int] = None
        self.user_id: Optional[int] = None
        self.username: str = ""
        self.action: str = ""
        self.table_name: str = ""
        self.record_id: Optional[int] = None
        self.old_values: str = ""
        self.new_values: str = ""
        self.created_at: Optional[str] = None
        super().__init__(**kwargs)


# =============================================================================
# SECTION 5: REPOSITORY LAYER
# =============================================================================

class BaseRepository:
    """
    Base repository providing common CRUD operations.
    All specific repositories inherit from this class.
    """

    table: str = ""
    model_class = BaseModel

    def __init__(self, db: DatabaseManager):
        self.db = db
        self._logger = logging.getLogger(f"repo.{self.__class__.__name__}")

    def create(self, data: dict) -> int:
        """Create a new record and return its ID."""
        # Remove None values for cleaner inserts
        clean_data = {k: v for k, v in data.items() if v is not None and k != 'id'}
        try:
            row_id = self.db.insert(self.table, clean_data)
            self._logger.debug(f"Created {self.table} record with ID: {row_id}")
            return row_id
        except Exception as e:
            self._logger.error(f"Error creating {self.table}: {e}")
            raise

    def get_by_id(self, record_id: int) -> Optional[dict]:
        """Get a record by ID."""
        return self.db.fetchone(
            f"SELECT * FROM {self.table} WHERE id = ?",
            (record_id,)
        )

    def update(self, record_id: int, data: dict) -> bool:
        """Update a record by ID."""
        clean_data = {k: v for k, v in data.items() if k != 'id'}
        if not clean_data:
            return False
        rows = self.db.update(self.table, clean_data, "id = ?", (record_id,))
        return rows > 0

    def delete(self, record_id: int) -> bool:
        """Delete a record by ID."""
        rows = self.db.delete(self.table, "id = ?", (record_id,))
        return rows > 0

    def list_all(self, filters: dict = None, page: int = 1,
                 per_page: int = DEFAULT_PAGE_SIZE,
                 order_by: str = "id DESC") -> List[dict]:
        """List all records with optional filtering and pagination."""
        where_clauses = []
        params = []

        if filters:
            for key, value in filters.items():
                if value is not None and value != "":
                    if isinstance(value, str) and '%' in value:
                        where_clauses.append(f"{key} LIKE ?")
                    else:
                        where_clauses.append(f"{key} = ?")
                    params.append(value)

        where_sql = ""
        if where_clauses:
            where_sql = "WHERE " + " AND ".join(where_clauses)

        offset = (page - 1) * per_page
        sql = f"""
            SELECT * FROM {self.table}
            {where_sql}
            ORDER BY {order_by}
            LIMIT ? OFFSET ?
        """
        params.extend([per_page, offset])
        return self.db.fetchall(sql, tuple(params))

    def count(self, filters: dict = None) -> int:
        """Count records with optional filtering."""
        where_clauses = []
        params = []

        if filters:
            for key, value in filters.items():
                if value is not None and value != "":
                    where_clauses.append(f"{key} = ?")
                    params.append(value)

        where_sql = ""
        if where_clauses:
            where_sql = "WHERE " + " AND ".join(where_clauses)

        result = self.db.fetchscalar(
            f"SELECT COUNT(*) FROM {self.table} {where_sql}",
            tuple(params)
        )
        return result or 0

    def search(self, query: str, fields: List[str]) -> List[dict]:
        """Search records across specified fields."""
        if not fields or not query:
            return []
        conditions = " OR ".join([f"{f} LIKE ?" for f in fields])
        params = tuple([f"%{query}%"] * len(fields))
        return self.db.fetchall(
            f"SELECT * FROM {self.table} WHERE {conditions} LIMIT 100",
            params
        )

    def exists(self, field: str, value: Any, exclude_id: int = None) -> bool:
        """Check if a value exists in a field."""
        sql = f"SELECT 1 FROM {self.table} WHERE {field} = ?"
        params = [value]
        if exclude_id:
            sql += " AND id != ?"
            params.append(exclude_id)
        return self.db.fetchone(sql, tuple(params)) is not None


class CompanyRepository(BaseRepository):
    """Repository for company operations."""
    table = "companies"

    def get_active(self) -> List[dict]:
        """Get all active companies."""
        return self.db.fetchall("SELECT * FROM companies WHERE is_active = 1")

    def get_default(self) -> Optional[dict]:
        """Get the first active company (default)."""
        return self.db.fetchone(
            "SELECT * FROM companies WHERE is_active = 1 ORDER BY id LIMIT 1"
        )


class ClientRepository(BaseRepository):
    """Repository for client operations."""
    table = "clients"

    def get_by_vat(self, vat: str) -> Optional[dict]:
        """Get client by VAT number."""
        return self.db.fetchone(
            "SELECT * FROM clients WHERE vat_number = ?", (vat,)
        )

    def get_by_cr(self, cr: str) -> Optional[dict]:
        """Get client by CR number."""
        return self.db.fetchone(
            "SELECT * FROM clients WHERE cr_number = ?", (cr,)
        )

    def get_active(self, company_id: int) -> List[dict]:
        """Get active clients for a company."""
        return self.db.fetchall(
            "SELECT * FROM clients WHERE company_id = ? AND status = 'active' ORDER BY name",
            (company_id,)
        )

    def get_statement(self, client_id: int) -> dict:
        """Get client account statement."""
        client = self.get_by_id(client_id)
        if not client:
            return {}

        invoices = self.db.fetchall(
            """SELECT i.*, 'invoice' as type FROM invoices i
               WHERE i.client_id = ? AND i.status IN ('posted','partial','paid')
               ORDER BY i.invoice_date""",
            (client_id,)
        )
        payments = self.db.fetchall(
            """SELECT p.*, 'payment' as type FROM payments p
               WHERE p.client_id = ? ORDER BY p.payment_date""",
            (client_id,)
        )

        total_invoiced = sum(inv['total_amount'] for inv in invoices)
        total_paid = sum(pay['amount'] for pay in payments)

        return {
            'client': client,
            'invoices': invoices,
            'payments': payments,
            'total_invoiced': total_invoiced,
            'total_paid': total_paid,
            'balance': total_invoiced - total_paid
        }

    def search_clients(self, query: str) -> List[dict]:
        """Search clients by name, CR, or VAT."""
        return self.search(query, ['name', 'name_ar', 'cr_number', 'vat_number', 'phone'])


class EmployeeRepository(BaseRepository):
    """Repository for employee operations."""
    table = "employees"

    def get_active(self, company_id: int = None) -> List[dict]:
        """Get all active employees."""
        if company_id:
            return self.db.fetchall(
                """SELECT e.*, d.name as department_name, b.name as branch_name
                   FROM employees e
                   LEFT JOIN departments d ON e.department_id = d.id
                   LEFT JOIN branches b ON e.branch_id = b.id
                   WHERE e.is_active = 1 AND e.company_id = ?
                   ORDER BY e.emp_number""",
                (company_id,)
            )
        return self.db.fetchall(
            """SELECT e.*, d.name as department_name, b.name as branch_name
               FROM employees e
               LEFT JOIN departments d ON e.department_id = d.id
               LEFT JOIN branches b ON e.branch_id = b.id
               WHERE e.is_active = 1 ORDER BY e.emp_number"""
        )

    def get_by_department(self, dept_id: int) -> List[dict]:
        """Get employees by department."""
        return self.db.fetchall(
            "SELECT * FROM employees WHERE department_id = ? AND is_active = 1",
            (dept_id,)
        )

    def search_employees(self, query: str) -> List[dict]:
        """Search employees by multiple fields."""
        return self.db.fetchall(
            """SELECT e.*, d.name as department_name FROM employees e
               LEFT JOIN departments d ON e.department_id = d.id
               WHERE e.first_name LIKE ? OR e.last_name LIKE ?
               OR e.arabic_name LIKE ? OR e.emp_number LIKE ?
               OR e.iqama_number LIKE ? OR e.passport_number LIKE ?
               OR e.mobile LIKE ? OR e.email LIKE ?
               LIMIT 50""",
            tuple([f"%{query}%"] * 8)
        )

    def get_expiring_documents(self, days: int = 30) -> List[dict]:
        """Get employees with expiring documents."""
        cutoff = (datetime.date.today() + datetime.timedelta(days=days)).isoformat()
        today = datetime.date.today().isoformat()

        return self.db.fetchall(
            """SELECT e.id, e.emp_number, e.first_name, e.last_name, e.arabic_name,
                      e.iqama_number, e.iqama_expiry,
                      e.passport_number, e.passport_expiry,
                      e.visa_number, e.visa_expiry
               FROM employees e
               WHERE e.is_active = 1 AND (
                   (e.iqama_expiry IS NOT NULL AND e.iqama_expiry BETWEEN ? AND ?)
                   OR (e.passport_expiry IS NOT NULL AND e.passport_expiry BETWEEN ? AND ?)
                   OR (e.visa_expiry IS NOT NULL AND e.visa_expiry BETWEEN ? AND ?)
               )
               ORDER BY e.iqama_expiry, e.passport_expiry, e.visa_expiry""",
            (today, cutoff, today, cutoff, today, cutoff)
        )

    def get_with_details(self, employee_id: int) -> Optional[dict]:
        """Get employee with all related details."""
        emp = self.db.fetchone(
            """SELECT e.*, d.name as department_name, b.name as branch_name,
                      cc.name as cost_center_name
               FROM employees e
               LEFT JOIN departments d ON e.department_id = d.id
               LEFT JOIN branches b ON e.branch_id = b.id
               LEFT JOIN cost_centers cc ON e.cost_center_id = cc.id
               WHERE e.id = ?""",
            (employee_id,)
        )
        if emp:
            emp['documents'] = self.db.fetchall(
                "SELECT * FROM employee_documents WHERE employee_id = ? ORDER BY expiry_date",
                (employee_id,)
            )
            emp['dependents'] = self.db.fetchall(
                "SELECT * FROM employee_dependents WHERE employee_id = ?",
                (employee_id,)
            )
            emp['history'] = self.db.fetchall(
                "SELECT * FROM employee_history WHERE employee_id = ? ORDER BY created_at DESC LIMIT 20",
                (employee_id,)
            )
        return emp

    def get_nationality_stats(self) -> List[dict]:
        """Get employee count by nationality."""
        return self.db.fetchall(
            """SELECT nationality, COUNT(*) as count
               FROM employees WHERE is_active = 1
               GROUP BY nationality ORDER BY count DESC"""
        )

    def get_department_stats(self) -> List[dict]:
        """Get employee count by department."""
        return self.db.fetchall(
            """SELECT d.name as department, COUNT(e.id) as count
               FROM employees e
               JOIN departments d ON e.department_id = d.id
               WHERE e.is_active = 1
               GROUP BY d.id, d.name ORDER BY count DESC"""
        )


class ContractRepository(BaseRepository):
    """Repository for contract operations."""
    table = "contracts"

    def get_active_contracts(self, company_id: int = None) -> List[dict]:
        """Get active contracts."""
        sql = """SELECT c.*, cl.name as client_name
                 FROM contracts c
                 JOIN clients cl ON c.client_id = cl.id
                 WHERE c.status = 'active'"""
        if company_id:
            sql += " AND c.company_id = ?"
            return self.db.fetchall(sql, (company_id,))
        return self.db.fetchall(sql)

    def get_with_lines(self, contract_id: int) -> Optional[dict]:
        """Get contract with all its lines."""
        contract = self.get_by_id(contract_id)
        if contract:
            contract['lines'] = self.db.fetchall(
                "SELECT * FROM contract_lines WHERE contract_id = ? ORDER BY sort_order",
                (contract_id,)
            )
            contract['assignments'] = self.db.fetchall(
                """SELECT ea.*, e.first_name, e.last_name, e.emp_number
                   FROM employee_assignments ea
                   JOIN employees e ON ea.employee_id = e.id
                   WHERE ea.contract_id = ? AND ea.is_active = 1""",
                (contract_id,)
            )
        return contract

    def get_expiring(self, days: int = 30) -> List[dict]:
        """Get contracts expiring soon."""
        cutoff = (datetime.date.today() + datetime.timedelta(days=days)).isoformat()
        today = datetime.date.today().isoformat()
        return self.db.fetchall(
            """SELECT c.*, cl.name as client_name
               FROM contracts c JOIN clients cl ON c.client_id = cl.id
               WHERE c.status = 'active'
               AND c.end_date BETWEEN ? AND ?
               ORDER BY c.end_date""",
            (today, cutoff)
        )


class PayrollRepository(BaseRepository):
    """Repository for payroll operations."""
    table = "payroll_periods"

    def get_period_entries(self, period_id: int) -> List[dict]:
        """Get all payroll entries for a period."""
        return self.db.fetchall(
            """SELECT pe.*, e.first_name, e.last_name, e.emp_number,
                      e.bank_iban, e.bank_name, e.is_saudi, e.nationality
               FROM payroll_entries pe
               JOIN employees e ON pe.employee_id = e.id
               WHERE pe.period_id = ?
               ORDER BY e.emp_number""",
            (period_id,)
        )

    def get_run_entries(self, run_id: int) -> List[dict]:
        """Get entries for a payroll run."""
        return self.db.fetchall(
            """SELECT pe.*, e.first_name, e.last_name, e.emp_number,
                      e.bank_iban, e.bank_name, e.is_saudi, e.nationality,
                      e.department_id, d.name as department_name
               FROM payroll_entries pe
               JOIN employees e ON pe.employee_id = e.id
               LEFT JOIN departments d ON e.department_id = d.id
               WHERE pe.payroll_run_id = ?
               ORDER BY e.emp_number""",
            (run_id,)
        )

    def get_employee_history(self, employee_id: int, months: int = 12) -> List[dict]:
        """Get payroll history for an employee."""
        return self.db.fetchall(
            """SELECT pe.*, pp.period_name, pp.period_year, pp.period_month
               FROM payroll_entries pe
               JOIN payroll_periods pp ON pe.period_id = pp.id
               WHERE pe.employee_id = ?
               ORDER BY pp.period_year DESC, pp.period_month DESC
               LIMIT ?""",
            (employee_id, months)
        )

    def get_active_loans(self, employee_id: int) -> List[dict]:
        """Get active loans for an employee."""
        return self.db.fetchall(
            """SELECT * FROM employee_loans
               WHERE employee_id = ? AND status = 'approved'
               AND outstanding_amount > 0""",
            (employee_id,)
        )


class InvoiceRepository(BaseRepository):
    """Repository for invoice operations."""
    table = "invoices"

    def get_outstanding(self, company_id: int = None) -> List[dict]:
        """Get outstanding invoices."""
        sql = """SELECT i.*, c.name as client_name
                 FROM invoices i JOIN clients c ON i.client_id = c.id
                 WHERE i.status IN ('posted', 'partial') AND i.outstanding_amount > 0"""
        if company_id:
            sql += " AND i.company_id = ?"
            sql += " ORDER BY i.due_date"
            return self.db.fetchall(sql, (company_id,))
        sql += " ORDER BY i.due_date"
        return self.db.fetchall(sql)

    def get_by_client(self, client_id: int) -> List[dict]:
        """Get invoices for a client."""
        return self.db.fetchall(
            """SELECT * FROM invoices WHERE client_id = ?
               ORDER BY invoice_date DESC""",
            (client_id,)
        )

    def get_with_lines(self, invoice_id: int) -> Optional[dict]:
        """Get invoice with all lines."""
        invoice = self.get_by_id(invoice_id)
        if invoice:
            invoice['lines'] = self.db.fetchall(
                "SELECT * FROM invoice_lines WHERE invoice_id = ? ORDER BY sort_order",
                (invoice_id,)
            )
            invoice['payments'] = self.db.fetchall(
                "SELECT * FROM payments WHERE invoice_id = ? ORDER BY payment_date",
                (invoice_id,)
            )
        return invoice

    def get_overdue(self, company_id: int) -> List[dict]:
        """Get overdue invoices."""
        today = datetime.date.today().isoformat()
        return self.db.fetchall(
            """SELECT i.*, c.name as client_name,
                      julianday('now') - julianday(i.due_date) as days_overdue
               FROM invoices i JOIN clients c ON i.client_id = c.id
               WHERE i.company_id = ? AND i.status IN ('posted','partial')
               AND i.due_date < ? AND i.outstanding_amount > 0
               ORDER BY i.due_date""",
            (company_id, today)
        )

    def get_ar_aging(self, company_id: int) -> dict:
        """Get Accounts Receivable aging report."""
        today = datetime.date.today().isoformat()
        rows = self.db.fetchall(
            """SELECT c.name as client_name,
                      SUM(CASE WHEN julianday(?) - julianday(i.due_date) <= 0 THEN i.outstanding_amount ELSE 0 END) as current_amount,
                      SUM(CASE WHEN julianday(?) - julianday(i.due_date) BETWEEN 1 AND 30 THEN i.outstanding_amount ELSE 0 END) as d1_30,
                      SUM(CASE WHEN julianday(?) - julianday(i.due_date) BETWEEN 31 AND 60 THEN i.outstanding_amount ELSE 0 END) as d31_60,
                      SUM(CASE WHEN julianday(?) - julianday(i.due_date) BETWEEN 61 AND 90 THEN i.outstanding_amount ELSE 0 END) as d61_90,
                      SUM(CASE WHEN julianday(?) - julianday(i.due_date) > 90 THEN i.outstanding_amount ELSE 0 END) as over_90,
                      SUM(i.outstanding_amount) as total
               FROM invoices i JOIN clients c ON i.client_id = c.id
               WHERE i.company_id = ? AND i.outstanding_amount > 0
               AND i.status IN ('posted','partial')
               GROUP BY c.id, c.name ORDER BY total DESC""",
            (today, today, today, today, today, company_id)
        )
        return {'rows': rows, 'as_of': today}


class AccountingRepository(BaseRepository):
    """Repository for accounting operations."""
    table = "journal_entries"

    def get_trial_balance(self, company_id: int, start_date: str, end_date: str) -> List[dict]:
        """Get trial balance data."""
        return self.db.fetchall(
            """SELECT a.account_code, a.account_name, a.account_type, a.normal_balance,
                      SUM(jl.debit) as total_debit,
                      SUM(jl.credit) as total_credit,
                      CASE WHEN a.normal_balance = 'debit'
                           THEN SUM(jl.debit) - SUM(jl.credit)
                           ELSE SUM(jl.credit) - SUM(jl.debit)
                      END as balance
               FROM journal_lines jl
               JOIN journal_entries je ON jl.journal_entry_id = je.id
               JOIN chart_of_accounts a ON jl.account_id = a.id
               WHERE je.company_id = ? AND je.status = 'posted'
               AND je.entry_date BETWEEN ? AND ?
               GROUP BY a.id, a.account_code, a.account_name, a.account_type, a.normal_balance
               HAVING (SUM(jl.debit) > 0 OR SUM(jl.credit) > 0)
               ORDER BY a.account_code""",
            (company_id, start_date, end_date)
        )

    def get_ledger(self, account_id: int, start_date: str = None, end_date: str = None) -> List[dict]:
        """Get general ledger for an account."""
        params = [account_id]
        date_filter = ""
        if start_date and end_date:
            date_filter = "AND je.entry_date BETWEEN ? AND ?"
            params.extend([start_date, end_date])

        return self.db.fetchall(
            f"""SELECT je.entry_date, je.entry_number, je.description,
                       jl.description as line_description, jl.debit, jl.credit
                FROM journal_lines jl
                JOIN journal_entries je ON jl.journal_entry_id = je.id
                WHERE jl.account_id = ? AND je.status = 'posted'
                {date_filter}
                ORDER BY je.entry_date, je.id""",
            tuple(params)
        )

    def get_account_balance(self, account_id: int, as_of_date: str = None) -> float:
        """Get account balance as of a date."""
        if not as_of_date:
            as_of_date = datetime.date.today().isoformat()

        row = self.db.fetchone(
            """SELECT a.normal_balance,
                      COALESCE(SUM(jl.debit), 0) as total_debit,
                      COALESCE(SUM(jl.credit), 0) as total_credit
               FROM chart_of_accounts a
               LEFT JOIN journal_lines jl ON jl.account_id = a.id
               LEFT JOIN journal_entries je ON jl.journal_entry_id = je.id
                   AND je.status = 'posted' AND je.entry_date <= ?
               WHERE a.id = ?
               GROUP BY a.id, a.normal_balance""",
            (as_of_date, account_id)
        )
        if not row:
            return 0.0
        if row['normal_balance'] == 'debit':
            return row['total_debit'] - row['total_credit']
        return row['total_credit'] - row['total_debit']


class LeaveRepository(BaseRepository):
    """Repository for leave management."""
    table = "leave_requests"

    def get_employee_balance(self, emp_id: int, leave_type_id: int, year: int = None) -> float:
        """Get employee leave balance."""
        if not year:
            year = datetime.date.today().year

        balance = self.db.fetchone(
            """SELECT * FROM leave_balances
               WHERE employee_id = ? AND leave_type_id = ? AND year = ?""",
            (emp_id, leave_type_id, year)
        )
        if balance:
            return balance.get('closing_balance', 0.0)
        return 0.0

    def get_employee_leaves(self, emp_id: int, year: int = None) -> List[dict]:
        """Get all leaves for an employee."""
        if not year:
            year = datetime.date.today().year
        return self.db.fetchall(
            """SELECT lr.*, lt.name as leave_type_name
               FROM leave_requests lr
               JOIN leave_types lt ON lr.leave_type_id = lt.id
               WHERE lr.employee_id = ?
               AND strftime('%Y', lr.start_date) = ?
               ORDER BY lr.start_date DESC""",
            (emp_id, str(year))
        )

    def get_pending_approvals(self, approver_id: int = None) -> List[dict]:
        """Get pending leave requests."""
        return self.db.fetchall(
            """SELECT lr.*, e.first_name, e.last_name, e.emp_number,
                      lt.name as leave_type_name
               FROM leave_requests lr
               JOIN employees e ON lr.employee_id = e.id
               JOIN leave_types lt ON lr.leave_type_id = lt.id
               WHERE lr.status = 'pending'
               ORDER BY lr.applied_date"""
        )


class AssetRepository(BaseRepository):
    """Repository for asset management."""
    table = "assets"

    def get_depreciable(self, company_id: int) -> List[dict]:
        """Get assets that can be depreciated."""
        return self.db.fetchall(
            """SELECT * FROM assets
               WHERE company_id = ? AND is_active = 1
               AND net_book_value > salvage_value
               AND disposal_date IS NULL""",
            (company_id,)
        )

    def get_depreciation_schedule(self, asset_id: int) -> List[dict]:
        """Get depreciation schedule for an asset."""
        return self.db.fetchall(
            "SELECT * FROM asset_depreciation WHERE asset_id = ? ORDER BY depreciation_date",
            (asset_id,)
        )

    def get_by_category(self, category: str, company_id: int) -> List[dict]:
        """Get assets by category."""
        return self.db.fetchall(
            "SELECT * FROM assets WHERE company_id = ? AND category = ? AND is_active = 1",
            (company_id, category)
        )


# =============================================================================
# SECTION 6: SERVICE LAYER
# =============================================================================

class BaseService:
    """Base service class providing database access and logging."""

    def __init__(self, db: DatabaseManager):
        self.db = db
        self._logger = logging.getLogger(f"service.{self.__class__.__name__}")


class AuthService(BaseService):
    """Authentication and authorization service."""

    def __init__(self, db: DatabaseManager):
        super().__init__(db)
        self.security = SecurityManager(db)

    def login(self, username: str, password: str) -> Optional[dict]:
        """Authenticate user and return user dict or None."""
        user = self.db.fetchone(
            "SELECT * FROM users WHERE username = ? AND is_active = 1",
            (username,)
        )
        if not user:
            self._logger.warning(f"Login failed: user '{username}' not found")
            return None

        # Check if account is locked
        if user.get('locked_until'):
            locked_until = datetime.datetime.fromisoformat(user['locked_until'])
            if datetime.datetime.now() < locked_until:
                self._logger.warning(f"Login failed: account locked for '{username}'")
                return None

        if not self.security.verify_password(password, user['password_hash']):
            # Increment failed attempts
            attempts = (user.get('login_attempts') or 0) + 1
            update_data = {'login_attempts': attempts}
            if attempts >= 5:
                locked_until = (datetime.datetime.now() + datetime.timedelta(minutes=30)).isoformat()
                update_data['locked_until'] = locked_until
                self._logger.warning(f"Account locked: '{username}' after {attempts} attempts")
            self.db.update('users', update_data, "id = ?", (user['id'],))
            return None

        # Successful login - reset attempts
        self.db.update('users', {
            'login_attempts': 0,
            'locked_until': None,
            'last_login': datetime.datetime.now().isoformat()
        }, "id = ?", (user['id'],))

        # Get user roles and permissions
        user_dict = dict(user)
        user_dict['roles'] = self.get_user_roles(user['id'])
        user_dict['permissions'] = self.get_user_permissions(user['id'])
        user_dict.pop('password_hash', None)  # Don't return password hash

        self.security.audit_log(user['id'], 'LOGIN', 'users', user['id'], 'User logged in')
        self._logger.info(f"User '{username}' logged in successfully")
        return user_dict

    def logout(self, user_id: int):
        """Log out a user."""
        self.security.audit_log(user_id, 'LOGOUT', 'users', user_id, 'User logged out')

    def hash_password(self, password: str) -> str:
        """Hash a password."""
        return self.security.hash_password(password)

    def change_password(self, user_id: int, old_password: str, new_password: str) -> bool:
        """Change user password."""
        user = self.db.fetchone("SELECT * FROM users WHERE id = ?", (user_id,))
        if not user:
            return False
        if not self.security.verify_password(old_password, user['password_hash']):
            return False
        new_hash = self.security.hash_password(new_password)
        self.db.update('users', {
            'password_hash': new_hash,
            'must_change_password': 0
        }, "id = ?", (user_id,))
        return True

    def reset_password(self, user_id: int, new_password: str) -> bool:
        """Reset user password (admin action)."""
        new_hash = self.security.hash_password(new_password)
        rows = self.db.update('users', {
            'password_hash': new_hash,
            'must_change_password': 1,
            'login_attempts': 0,
            'locked_until': None
        }, "id = ?", (user_id,))
        return rows > 0

    def check_permission(self, user_id: int, permission: str) -> bool:
        """Check if user has a specific permission."""
        user = self.db.fetchone("SELECT * FROM users WHERE id = ?", (user_id,))
        if not user:
            return False
        if user.get('is_admin'):
            return True
        return self.security.check_permission(user_id, permission)

    def get_user_permissions(self, user_id: int) -> List[str]:
        """Get all permissions for a user."""
        perms = self.db.fetchall(
            """SELECT DISTINCT p.code FROM permissions p
               JOIN role_permissions rp ON p.id = rp.permission_id
               JOIN user_roles ur ON rp.role_id = ur.role_id
               WHERE ur.user_id = ?""",
            (user_id,)
        )
        return [p['code'] for p in perms]

    def get_user_roles(self, user_id: int) -> List[dict]:
        """Get roles assigned to a user."""
        return self.db.fetchall(
            """SELECT r.* FROM roles r
               JOIN user_roles ur ON r.id = ur.role_id
               WHERE ur.user_id = ?""",
            (user_id,)
        )

    def assign_role(self, user_id: int, role_id: int) -> bool:
        """Assign a role to a user."""
        try:
            self.db.execute(
                "INSERT OR IGNORE INTO user_roles (user_id, role_id) VALUES (?, ?)",
                (user_id, role_id)
            )
            return True
        except Exception:
            return False

    def create_user(self, data: dict) -> int:
        """Create a new user."""
        password = data.pop('password', 'admin123')
        data['password_hash'] = self.hash_password(password)
        user_id = self.db.insert('users', data)
        return user_id


class EmployeeService(BaseService):
    """Comprehensive employee management service."""

    def __init__(self, db: DatabaseManager):
        super().__init__(db)
        self.repo = EmployeeRepository(db)

    def create_employee(self, data: dict, created_by: int = None) -> int:
        """Create a new employee with auto-generated employee number."""
        company_id = data.get('company_id', 1)

        # Auto-generate employee number if not provided
        if not data.get('emp_number'):
            data['emp_number'] = self.db.get_next_sequence(company_id, 'employee')

        # Calculate total salary
        data['total_salary'] = (
            data.get('basic_salary', 0) +
            data.get('housing_allowance', 0) +
            data.get('transport_allowance', 0) +
            data.get('food_allowance', 0) +
            data.get('phone_allowance', 0) +
            data.get('other_allowances', 0)
        )

        # Set probation end date if not provided (3 months)
        if data.get('hire_date') and not data.get('probation_end_date'):
            hire = datetime.date.fromisoformat(data['hire_date'])
            probation_end = hire + datetime.timedelta(days=90)
            data['probation_end_date'] = probation_end.isoformat()

        if created_by:
            data['created_by'] = created_by

        emp_id = self.db.insert('employees', data)

        # Initialize leave balances
        self._initialize_leave_balances(emp_id, company_id, data.get('hire_date'))

        # Log creation
        self._logger.info(f"Employee created: {data.get('emp_number')} (ID: {emp_id})")
        return emp_id

    def update_employee(self, emp_id: int, data: dict, updated_by: int = None) -> bool:
        """Update employee data and track changes."""
        old_emp = self.repo.get_by_id(emp_id)
        if not old_emp:
            return False

        # Recalculate total salary if salary fields changed
        salary_fields = ['basic_salary', 'housing_allowance', 'transport_allowance',
                        'food_allowance', 'phone_allowance', 'other_allowances']
        if any(f in data for f in salary_fields):
            data['total_salary'] = (
                data.get('basic_salary', old_emp.get('basic_salary', 0)) +
                data.get('housing_allowance', old_emp.get('housing_allowance', 0)) +
                data.get('transport_allowance', old_emp.get('transport_allowance', 0)) +
                data.get('food_allowance', old_emp.get('food_allowance', 0)) +
                data.get('phone_allowance', old_emp.get('phone_allowance', 0)) +
                data.get('other_allowances', old_emp.get('other_allowances', 0))
            )

        # Track important changes
        tracked_fields = ['basic_salary', 'job_title', 'department_id', 'branch_id']
        for field in tracked_fields:
            if field in data and data[field] != old_emp.get(field):
                self.db.insert('employee_history', {
                    'employee_id': emp_id,
                    'change_type': 'update',
                    'field_name': field,
                    'old_value': str(old_emp.get(field, '')),
                    'new_value': str(data[field]),
                    'effective_date': datetime.date.today().isoformat(),
                    'created_by': updated_by
                })

        rows = self.db.update('employees', data, "id = ?", (emp_id,))
        return rows > 0

    def terminate_employee(self, emp_id: int, termination_date: str,
                          reason: str, last_salary: bool = True) -> bool:
        """Terminate an employee."""
        emp = self.repo.get_by_id(emp_id)
        if not emp:
            return False

        rows = self.db.update('employees', {
            'is_active': 0,
            'termination_date': termination_date,
            'termination_reason': reason
        }, "id = ?", (emp_id,))

        if rows > 0:
            # Log termination
            self.db.insert('employee_history', {
                'employee_id': emp_id,
                'change_type': 'termination',
                'field_name': 'status',
                'old_value': 'active',
                'new_value': 'terminated',
                'effective_date': termination_date,
                'reason': reason
            })
            self._logger.info(f"Employee {emp_id} terminated on {termination_date}")
            return True
        return False

    def get_employee_details(self, emp_id: int) -> Optional[dict]:
        """Get full employee details including all related data."""
        return self.repo.get_with_details(emp_id)

    def get_dashboard_stats(self, company_id: int = 1) -> dict:
        """Get employee statistics for dashboard."""
        stats = {
            'total_employees': self.db.fetchscalar(
                "SELECT COUNT(*) FROM employees WHERE company_id = ? AND is_active = 1",
                (company_id,)
            ) or 0,
            'saudi_employees': self.db.fetchscalar(
                "SELECT COUNT(*) FROM employees WHERE company_id = ? AND is_active = 1 AND is_saudi = 1",
                (company_id,)
            ) or 0,
            'non_saudi_employees': self.db.fetchscalar(
                "SELECT COUNT(*) FROM employees WHERE company_id = ? AND is_active = 1 AND is_saudi = 0",
                (company_id,)
            ) or 0,
            'on_leave': self.db.fetchscalar(
                "SELECT COUNT(*) FROM employees WHERE company_id = ? AND is_on_leave = 1",
                (company_id,)
            ) or 0,
            'new_hires_this_month': self.db.fetchscalar(
                """SELECT COUNT(*) FROM employees
                   WHERE company_id = ? AND is_active = 1
                   AND strftime('%Y-%m', hire_date) = strftime('%Y-%m', 'now')""",
                (company_id,)
            ) or 0,
            'expiring_iqama_30': self.db.fetchscalar(
                """SELECT COUNT(*) FROM employees
                   WHERE company_id = ? AND is_active = 1
                   AND iqama_expiry BETWEEN date('now') AND date('now', '+30 days')""",
                (company_id,)
            ) or 0,
            'nationality_breakdown': self.db.fetchall(
                """SELECT nationality, COUNT(*) as count
                   FROM employees WHERE company_id = ? AND is_active = 1
                   GROUP BY nationality ORDER BY count DESC LIMIT 10""",
                (company_id,)
            ),
            'department_breakdown': self.db.fetchall(
                """SELECT d.name, COUNT(e.id) as count
                   FROM employees e JOIN departments d ON e.department_id = d.id
                   WHERE e.company_id = ? AND e.is_active = 1
                   GROUP BY d.id, d.name ORDER BY count DESC""",
                (company_id,)
            )
        }
        total = stats['total_employees'] or 1
        stats['saudization_rate'] = round((stats['saudi_employees'] / total) * 100, 1)
        return stats

    def _initialize_leave_balances(self, emp_id: int, company_id: int, hire_date: str = None):
        """Initialize leave balances for a new employee."""
        leave_types = self.db.fetchall(
            "SELECT * FROM leave_types WHERE company_id = ? AND is_active = 1",
            (company_id,)
        )
        year = datetime.date.today().year
        for lt in leave_types:
            # Calculate prorated balance if mid-year hire
            days = lt['days_per_year']
            if hire_date:
                try:
                    hire = datetime.date.fromisoformat(hire_date)
                    days_remaining = (datetime.date(year, 12, 31) - hire).days + 1
                    days = round(lt['days_per_year'] * days_remaining / 365, 1)
                except Exception:
                    pass

            self.db.execute(
                """INSERT OR IGNORE INTO leave_balances
                   (employee_id, leave_type_id, year, opening_balance, closing_balance)
                   VALUES (?, ?, ?, ?, ?)""",
                (emp_id, lt['id'], year, days, days)
            )


class ContractService(BaseService):
    """Contract management service."""

    def __init__(self, db: DatabaseManager):
        super().__init__(db)
        self.repo = ContractRepository(db)

    def create_contract(self, data: dict, lines: List[dict] = None) -> int:
        """Create a new contract with optional lines."""
        company_id = data.get('company_id', 1)
        if not data.get('contract_number'):
            data['contract_number'] = self.db.get_next_sequence(company_id, 'contract')

        contract_id = self.db.insert('contracts', data)

        if lines:
            for i, line in enumerate(lines):
                line['contract_id'] = contract_id
                line['sort_order'] = i
                # Calculate line totals
                qty = line.get('quantity', 1)
                price = line.get('unit_price', 0)
                disc = line.get('discount_percent', 0)
                total_price = qty * price
                discount_amt = total_price * (disc / 100)
                net = total_price - discount_amt
                vat_rate = line.get('vat_rate', 0.15)
                vat_amt = net * vat_rate
                line['total_price'] = total_price
                line['vat_amount'] = vat_amt
                line['line_total'] = net + vat_amt
                self.db.insert('contract_lines', line)

        # Update total value
        total = self.db.fetchscalar(
            "SELECT SUM(line_total) FROM contract_lines WHERE contract_id = ?",
            (contract_id,)
        ) or 0
        self.db.update('contracts', {'total_value': total}, "id = ?", (contract_id,))

        return contract_id

    def assign_employee(self, contract_id: int, emp_id: int, rate: float,
                       position: str = None, start_date: str = None) -> bool:
        """Assign an employee to a contract."""
        try:
            self.db.insert('employee_assignments', {
                'employee_id': emp_id,
                'contract_id': contract_id,
                'bill_rate': rate,
                'position': position,
                'assignment_date': start_date or datetime.date.today().isoformat(),
                'is_active': 1
            })
            return True
        except Exception as e:
            self._logger.error(f"Error assigning employee {emp_id} to contract {contract_id}: {e}")
            return False

    def get_contract_profitability(self, contract_id: int) -> dict:
        """Calculate contract profitability."""
        contract = self.repo.get_with_lines(contract_id)
        if not contract:
            return {}

        # Revenue from invoices
        revenue = self.db.fetchscalar(
            """SELECT COALESCE(SUM(subtotal), 0) FROM invoices
               WHERE contract_id = ? AND status IN ('posted','partial','paid')""",
            (contract_id,)
        ) or 0

        # Cost from employee assignments (salaries)
        cost = self.db.fetchscalar(
            """SELECT COALESCE(SUM(e.total_salary), 0)
               FROM employee_assignments ea
               JOIN employees e ON ea.employee_id = e.id
               WHERE ea.contract_id = ? AND ea.is_active = 1""",
            (contract_id,)
        ) or 0

        gross_profit = revenue - cost
        margin = (gross_profit / revenue * 100) if revenue > 0 else 0

        return {
            'contract': contract,
            'total_revenue': revenue,
            'total_cost': cost,
            'gross_profit': gross_profit,
            'margin_percent': round(margin, 2),
            'assigned_employees': len(contract.get('assignments', []))
        }


class LeaveService(BaseService):
    """Leave management service."""

    def __init__(self, db: DatabaseManager):
        super().__init__(db)
        self.repo = LeaveRepository(db)

    def request_leave(self, emp_id: int, leave_type_id: int,
                     start: str, end: str, reason: str = "") -> int:
        """Submit a leave request."""
        # Calculate days
        start_date = datetime.date.fromisoformat(start)
        end_date = datetime.date.fromisoformat(end)
        days = (end_date - start_date).days + 1

        if days <= 0:
            raise ValueError("End date must be after start date")

        # Check balance
        balance = self.repo.get_employee_balance(
            emp_id, leave_type_id, start_date.year
        )

        leave_type = self.db.fetchone("SELECT * FROM leave_types WHERE id = ?", (leave_type_id,))
        if leave_type and leave_type.get('is_paid') and balance < days:
            raise ValueError(f"Insufficient leave balance. Available: {balance} days, Requested: {days} days")

        request_id = self.db.insert('leave_requests', {
            'employee_id': emp_id,
            'leave_type_id': leave_type_id,
            'start_date': start,
            'end_date': end,
            'days_requested': days,
            'reason': reason,
            'status': 'pending',
            'applied_date': datetime.datetime.now().isoformat()
        })
        return request_id

    def approve_leave(self, request_id: int, approver_id: int) -> bool:
        """Approve a leave request."""
        request = self.db.fetchone("SELECT * FROM leave_requests WHERE id = ?", (request_id,))
        if not request or request['status'] != 'pending':
            return False

        self.db.update('leave_requests', {
            'status': 'approved',
            'approved_by': approver_id,
            'approved_at': datetime.datetime.now().isoformat()
        }, "id = ?", (request_id,))

        # Deduct from balance
        year = datetime.date.fromisoformat(request['start_date']).year
        self._deduct_balance(
            request['employee_id'],
            request['leave_type_id'],
            year,
            request['days_requested']
        )

        # Update employee leave status if current
        today = datetime.date.today().isoformat()
        if request['start_date'] <= today <= request['end_date']:
            self.db.update('employees', {'is_on_leave': 1},
                          "id = ?", (request['employee_id'],))

        return True

    def reject_leave(self, request_id: int, reason: str = "") -> bool:
        """Reject a leave request."""
        rows = self.db.update('leave_requests', {
            'status': 'rejected',
            'rejection_reason': reason
        }, "id = ?", (request_id,))
        return rows > 0

    def get_balance(self, emp_id: int, year: int = None) -> dict:
        """Get all leave balances for an employee."""
        if not year:
            year = datetime.date.today().year

        balances = self.db.fetchall(
            """SELECT lb.*, lt.name as leave_type_name, lt.code, lt.days_per_year
               FROM leave_balances lb
               JOIN leave_types lt ON lb.leave_type_id = lt.id
               WHERE lb.employee_id = ? AND lb.year = ?""",
            (emp_id, year)
        )
        return {b['code']: b for b in balances}

    def accrue_leave(self, emp_id: int, leave_type_id: int = None) -> bool:
        """Accrue monthly leave for an employee."""
        emp = self.db.fetchone("SELECT * FROM employees WHERE id = ?", (emp_id,))
        if not emp or not emp['is_active']:
            return False

        year = datetime.date.today().year
        company_id = emp['company_id']

        leave_types_query = "SELECT * FROM leave_types WHERE company_id = ? AND is_active = 1"
        params = [company_id]
        if leave_type_id:
            leave_types_query += " AND id = ?"
            params.append(leave_type_id)

        leave_types = self.db.fetchall(leave_types_query, tuple(params))

        for lt in leave_types:
            monthly_accrual = lt['days_per_year'] / 12.0

            # Get or create balance record
            balance = self.db.fetchone(
                "SELECT * FROM leave_balances WHERE employee_id = ? AND leave_type_id = ? AND year = ?",
                (emp_id, lt['id'], year)
            )

            if balance:
                new_accrued = balance['accrued'] + monthly_accrual
                new_closing = balance['opening_balance'] + new_accrued - balance['used']
                self.db.update('leave_balances', {
                    'accrued': new_accrued,
                    'closing_balance': new_closing,
                    'updated_at': datetime.datetime.now().isoformat()
                }, "id = ?", (balance['id'],))
            else:
                self.db.insert('leave_balances', {
                    'employee_id': emp_id,
                    'leave_type_id': lt['id'],
                    'year': year,
                    'opening_balance': 0,
                    'accrued': monthly_accrual,
                    'used': 0,
                    'closing_balance': monthly_accrual
                })
        return True

    def _deduct_balance(self, emp_id: int, leave_type_id: int, year: int, days: float):
        """Deduct days from leave balance."""
        balance = self.db.fetchone(
            "SELECT * FROM leave_balances WHERE employee_id = ? AND leave_type_id = ? AND year = ?",
            (emp_id, leave_type_id, year)
        )
        if balance:
            new_used = balance['used'] + days
            new_closing = balance['closing_balance'] - days
            self.db.update('leave_balances', {
                'used': new_used,
                'closing_balance': max(0, new_closing),
                'updated_at': datetime.datetime.now().isoformat()
            }, "id = ?", (balance['id'],))


class InvoiceService(BaseService):
    """Invoice and billing service."""

    def __init__(self, db: DatabaseManager):
        super().__init__(db)
        self.repo = InvoiceRepository(db)
        self.zatca = ZATCAEngine(db)
        self.accounting = AccountingService(db)

    def create_invoice(self, data: dict, lines: List[dict]) -> int:
        """Create a new invoice with line items."""
        company_id = data.get('company_id', 1)
        if not data.get('invoice_number'):
            data['invoice_number'] = self.db.get_next_sequence(company_id, 'invoice')

        # Set due date if not provided
        if not data.get('due_date') and data.get('invoice_date'):
            invoice_date = datetime.date.fromisoformat(data['invoice_date'])
            payment_terms = data.get('payment_terms', 30)
            data['due_date'] = (invoice_date + datetime.timedelta(days=payment_terms)).isoformat()

        # Calculate totals from lines
        subtotal = 0
        total_vat = 0
        for line in lines:
            qty = line.get('quantity', 1)
            price = line.get('unit_price', 0)
            disc_pct = line.get('discount_percent', 0)
            vat_rate = line.get('vat_rate', data.get('vat_rate', 0.15))

            line_subtotal = qty * price
            disc_amt = line_subtotal * (disc_pct / 100)
            net = line_subtotal - disc_amt
            vat_amt = net * vat_rate

            line['subtotal'] = round(net, 2)
            line['discount_amount'] = round(disc_amt, 2)
            line['vat_amount'] = round(vat_amt, 2)
            line['line_total'] = round(net + vat_amt, 2)

            subtotal += net
            total_vat += vat_amt

        data['subtotal'] = round(subtotal, 2)
        data['vat_amount'] = round(total_vat, 2)
        data['total_amount'] = round(subtotal + total_vat - data.get('discount_amount', 0), 2)
        data['outstanding_amount'] = data['total_amount']
        data['paid_amount'] = 0
        data['status'] = 'draft'

        invoice_id = self.db.insert('invoices', data)

        for i, line in enumerate(lines):
            line['invoice_id'] = invoice_id
            line['sort_order'] = i
            self.db.insert('invoice_lines', line)

        return invoice_id

    def post_invoice(self, invoice_id: int, posted_by: int = None) -> bool:
        """Post an invoice and create journal entry."""
        invoice = self.repo.get_with_lines(invoice_id)
        if not invoice or invoice['status'] != 'draft':
            return False

        company_id = invoice['company_id']

        # Generate ZATCA QR
        qr_code = self.zatca.generate_qr_code(invoice)

        # Get client AR account
        ar_account = self.db.fetchone(
            "SELECT * FROM chart_of_accounts WHERE company_id = ? AND account_code = '1200'",
            (company_id,)
        )
        revenue_account = self.db.fetchone(
            "SELECT * FROM chart_of_accounts WHERE company_id = ? AND account_code = '4100'",
            (company_id,)
        )
        vat_account = self.db.fetchone(
            "SELECT * FROM chart_of_accounts WHERE company_id = ? AND account_code = '2300'",
            (company_id,)
        )

        if ar_account and revenue_account:
            # Create journal entry
            je_data = {
                'company_id': company_id,
                'entry_number': self.db.get_next_sequence(company_id, 'journal'),
                'entry_date': invoice['invoice_date'],
                'description': f"Invoice {invoice['invoice_number']}",
                'source_type': 'invoice',
                'source_id': invoice_id
            }
            lines = [
                {'account_id': ar_account['id'], 'debit': invoice['total_amount'],
                 'credit': 0, 'description': f"AR - {invoice['invoice_number']}"},
                {'account_id': revenue_account['id'], 'debit': 0,
                 'credit': invoice['subtotal'], 'description': 'Service revenue'},
            ]
            if invoice['vat_amount'] > 0 and vat_account:
                lines.append({
                    'account_id': vat_account['id'], 'debit': 0,
                    'credit': invoice['vat_amount'], 'description': 'Output VAT'
                })

            je_id = self.accounting.create_journal_entry(je_data, lines)
            self.accounting.post_journal_entry(je_id)

            self.db.update('invoices', {
                'status': 'posted',
                'zatca_qr_code': qr_code,
                'zatca_uuid': str(uuid.uuid4()),
                'journal_entry_id': je_id
            }, "id = ?", (invoice_id,))
        else:
            self.db.update('invoices', {
                'status': 'posted',
                'zatca_qr_code': qr_code,
                'zatca_uuid': str(uuid.uuid4())
            }, "id = ?", (invoice_id,))

        return True

    def record_payment(self, invoice_id: int, amount: float,
                      payment_date: str, method: str = "bank_transfer",
                      reference: str = "") -> int:
        """Record a payment against an invoice."""
        invoice = self.repo.get_by_id(invoice_id)
        if not invoice:
            raise ValueError("Invoice not found")

        outstanding = invoice.get('outstanding_amount', 0)
        if amount > outstanding:
            raise ValueError(f"Payment amount ({amount}) exceeds outstanding ({outstanding})")

        company_id = invoice['company_id']
        payment_number = self.db.get_next_sequence(company_id, 'payment')

        payment_id = self.db.insert('payments', {
            'company_id': company_id,
            'payment_number': payment_number,
            'client_id': invoice['client_id'],
            'payment_date': payment_date,
            'amount': amount,
            'payment_method': method,
            'reference': reference,
            'invoice_id': invoice_id,
            'status': 'posted'
        })

        # Update invoice
        new_paid = invoice.get('paid_amount', 0) + amount
        new_outstanding = invoice['total_amount'] - new_paid
        new_status = 'paid' if new_outstanding <= 0.01 else 'partial'

        self.db.update('invoices', {
            'paid_amount': round(new_paid, 2),
            'outstanding_amount': max(0, round(new_outstanding, 2)),
            'status': new_status
        }, "id = ?", (invoice_id,))

        return payment_id

    def generate_zatca_qr(self, invoice_id: int) -> str:
        """Generate ZATCA QR code for an invoice."""
        invoice = self.repo.get_by_id(invoice_id)
        if not invoice:
            return ""
        return self.zatca.generate_qr_code(invoice)


class AccountingService(BaseService):
    """Accounting and financial management service."""

    def __init__(self, db: DatabaseManager):
        super().__init__(db)
        self.repo = AccountingRepository(db)

    def create_journal_entry(self, data: dict, lines: List[dict]) -> int:
        """Create a journal entry with lines."""
        company_id = data.get('company_id', 1)
        if not data.get('entry_number'):
            data['entry_number'] = self.db.get_next_sequence(company_id, 'journal')

        total_debit = sum(l.get('debit', 0) for l in lines)
        total_credit = sum(l.get('credit', 0) for l in lines)

        if abs(total_debit - total_credit) > 0.01:
            raise ValueError(
                f"Journal entry is not balanced: Debit {total_debit}, Credit {total_credit}"
            )

        data['total_debit'] = round(total_debit, 2)
        data['total_credit'] = round(total_credit, 2)
        data['status'] = 'draft'

        je_id = self.db.insert('journal_entries', data)

        for i, line in enumerate(lines):
            line['journal_entry_id'] = je_id
            line['sort_order'] = i
            self.db.insert('journal_lines', line)

        return je_id

    def post_journal_entry(self, entry_id: int, posted_by: int = None) -> bool:
        """Post a journal entry (mark as final)."""
        entry = self.db.fetchone("SELECT * FROM journal_entries WHERE id = ?", (entry_id,))
        if not entry or entry['status'] != 'draft':
            return False

        self.db.update('journal_entries', {
            'status': 'posted',
            'posted_by': posted_by,
            'posted_at': datetime.datetime.now().isoformat()
        }, "id = ?", (entry_id,))
        return True

    def get_trial_balance(self, company_id: int, start_date: str, end_date: str) -> List[dict]:
        """Get trial balance."""
        return self.repo.get_trial_balance(company_id, start_date, end_date)

    def get_profit_loss(self, company_id: int, start_date: str, end_date: str) -> dict:
        """Generate Profit & Loss statement."""
        rows = self.db.fetchall(
            """SELECT a.account_code, a.account_name, a.account_type, a.account_subtype,
                      a.normal_balance,
                      COALESCE(SUM(jl.debit), 0) as total_debit,
                      COALESCE(SUM(jl.credit), 0) as total_credit
               FROM chart_of_accounts a
               LEFT JOIN journal_lines jl ON jl.account_id = a.id
               LEFT JOIN journal_entries je ON jl.journal_entry_id = je.id
                   AND je.status = 'posted' AND je.company_id = ?
                   AND je.entry_date BETWEEN ? AND ?
               WHERE a.company_id = ? AND a.account_type IN ('revenue', 'expense')
               AND a.is_header = 0
               GROUP BY a.id, a.account_code, a.account_name, a.account_type, a.account_subtype, a.normal_balance
               ORDER BY a.account_code""",
            (company_id, start_date, end_date, company_id)
        )

        revenue = []
        expenses = []
        total_revenue = 0
        total_expenses = 0

        for row in rows:
            if row['normal_balance'] == 'credit':
                balance = row['total_credit'] - row['total_debit']
            else:
                balance = row['total_debit'] - row['total_credit']

            row['balance'] = round(balance, 2)

            if row['account_type'] == 'revenue':
                revenue.append(row)
                total_revenue += balance
            else:
                expenses.append(row)
                total_expenses += balance

        net_profit = total_revenue - total_expenses

        return {
            'period': f"{start_date} to {end_date}",
            'revenue': revenue,
            'expenses': expenses,
            'total_revenue': round(total_revenue, 2),
            'total_expenses': round(total_expenses, 2),
            'net_profit': round(net_profit, 2),
            'is_profit': net_profit >= 0
        }

    def get_balance_sheet(self, company_id: int, as_of_date: str) -> dict:
        """Generate Balance Sheet."""
        rows = self.db.fetchall(
            """SELECT a.account_code, a.account_name, a.account_type, a.account_subtype,
                      a.normal_balance,
                      COALESCE(SUM(jl.debit), 0) as total_debit,
                      COALESCE(SUM(jl.credit), 0) as total_credit
               FROM chart_of_accounts a
               LEFT JOIN journal_lines jl ON jl.account_id = a.id
               LEFT JOIN journal_entries je ON jl.journal_entry_id = je.id
                   AND je.status = 'posted' AND je.company_id = ?
                   AND je.entry_date <= ?
               WHERE a.company_id = ? AND a.account_type IN ('asset', 'liability', 'equity')
               AND a.is_header = 0
               GROUP BY a.id, a.account_code, a.account_name, a.account_type, a.account_subtype, a.normal_balance
               ORDER BY a.account_code""",
            (company_id, as_of_date, company_id)
        )

        assets = []
        liabilities = []
        equity = []
        total_assets = 0
        total_liabilities = 0
        total_equity = 0

        for row in rows:
            if row['normal_balance'] == 'debit':
                balance = row['total_debit'] - row['total_credit']
            else:
                balance = row['total_credit'] - row['total_debit']

            row['balance'] = round(balance, 2)

            if row['account_type'] == 'asset':
                assets.append(row)
                total_assets += balance
            elif row['account_type'] == 'liability':
                liabilities.append(row)
                total_liabilities += balance
            else:
                equity.append(row)
                total_equity += balance

        return {
            'as_of': as_of_date,
            'assets': assets,
            'liabilities': liabilities,
            'equity': equity,
            'total_assets': round(total_assets, 2),
            'total_liabilities': round(total_liabilities, 2),
            'total_equity': round(total_equity, 2),
            'is_balanced': abs(total_assets - (total_liabilities + total_equity)) < 1.0
        }

    def reconcile_bank(self, bank_account_id: int, statement_date: str,
                      statement_balance: float) -> dict:
        """Perform bank reconciliation."""
        bank_account = self.db.fetchone(
            "SELECT * FROM bank_accounts WHERE id = ?", (bank_account_id,)
        )
        if not bank_account:
            return {}

        # Get unreconciled transactions
        unreconciled = self.db.fetchall(
            """SELECT * FROM bank_transactions
               WHERE bank_account_id = ? AND is_reconciled = 0
               AND transaction_date <= ?
               ORDER BY transaction_date""",
            (bank_account_id, statement_date)
        )

        book_balance = self.db.fetchscalar(
            """SELECT COALESCE(SUM(CASE WHEN amount > 0 THEN amount ELSE 0 END) -
                               SUM(CASE WHEN amount < 0 THEN ABS(amount) ELSE 0 END), 0)
               FROM bank_transactions WHERE bank_account_id = ? AND transaction_date <= ?""",
            (bank_account_id, statement_date)
        ) or 0

        return {
            'bank_account': bank_account,
            'statement_date': statement_date,
            'statement_balance': statement_balance,
            'book_balance': book_balance,
            'difference': round(statement_balance - book_balance, 2),
            'unreconciled_transactions': unreconciled,
            'is_reconciled': abs(statement_balance - book_balance) < 0.01
        }


class AssetService(BaseService):
    """Fixed asset management service."""

    def __init__(self, db: DatabaseManager):
        super().__init__(db)
        self.repo = AssetRepository(db)

    def register_asset(self, data: dict) -> int:
        """Register a new asset."""
        company_id = data.get('company_id', 1)
        if not data.get('asset_number'):
            data['asset_number'] = self.db.get_next_sequence(company_id, 'asset')

        purchase_price = data.get('purchase_price', 0)
        data['net_book_value'] = purchase_price
        data['accumulated_depreciation'] = 0

        asset_id = self.db.insert('assets', data)
        return asset_id

    def calculate_depreciation(self, asset_id: int) -> float:
        """Calculate monthly depreciation for an asset."""
        asset = self.repo.get_by_id(asset_id)
        if not asset:
            return 0.0

        depreciable_amount = asset['purchase_price'] - asset.get('salvage_value', 0)
        useful_life = asset.get('useful_life_months', 60)

        if useful_life <= 0 or depreciable_amount <= 0:
            return 0.0

        method = asset.get('depreciation_method', 'straight_line')
        if method == 'straight_line':
            monthly_dep = depreciable_amount / useful_life
        elif method == 'declining_balance':
            rate = asset.get('depreciation_rate', 40) / 100
            monthly_dep = asset.get('net_book_value', 0) * rate / 12
        else:
            monthly_dep = depreciable_amount / useful_life

        # Don't depreciate below salvage value
        nbv = asset.get('net_book_value', 0)
        salvage = asset.get('salvage_value', 0)
        if nbv - monthly_dep < salvage:
            monthly_dep = max(0, nbv - salvage)

        return round(monthly_dep, 2)

    def run_depreciation_batch(self, company_id: int, period: str) -> dict:
        """Run depreciation for all assets in a period."""
        assets = self.repo.get_depreciable(company_id)
        results = {
            'period': period,
            'assets_processed': 0,
            'total_depreciation': 0,
            'errors': []
        }

        for asset in assets:
            try:
                amount = self.calculate_depreciation(asset['id'])
                if amount <= 0:
                    continue

                # Check if already depreciated for this period
                existing = self.db.fetchone(
                    "SELECT id FROM asset_depreciation WHERE asset_id = ? AND period = ?",
                    (asset['id'], period)
                )
                if existing:
                    continue

                new_accumulated = asset.get('accumulated_depreciation', 0) + amount
                new_nbv = asset['purchase_price'] - new_accumulated

                self.db.insert('asset_depreciation', {
                    'asset_id': asset['id'],
                    'period': period,
                    'depreciation_date': datetime.date.today().isoformat(),
                    'amount': amount,
                    'accumulated_amount': new_accumulated,
                    'net_book_value': max(0, new_nbv)
                })

                self.db.update('assets', {
                    'accumulated_depreciation': new_accumulated,
                    'net_book_value': max(0, new_nbv)
                }, "id = ?", (asset['id'],))

                results['assets_processed'] += 1
                results['total_depreciation'] += amount

            except Exception as e:
                results['errors'].append({'asset_id': asset['id'], 'error': str(e)})

        results['total_depreciation'] = round(results['total_depreciation'], 2)
        return results


class NotificationService(BaseService):
    """Notification management service."""

    def create_notification(self, user_id: int, title: str, message: str,
                           notification_type: str = 'info',
                           expires_in_days: int = 30) -> int:
        """Create a new notification."""
        expires_at = (datetime.datetime.now() + datetime.timedelta(days=expires_in_days)).isoformat()
        notification_id = self.db.insert('notifications', {
            'user_id': user_id,
            'title': title,
            'message': message,
            'notification_type': notification_type,
            'is_read': 0,
            'expires_at': expires_at
        })
        return notification_id

    def get_unread(self, user_id: int) -> List[dict]:
        """Get unread notifications for a user."""
        return self.db.fetchall(
            """SELECT * FROM notifications
               WHERE user_id = ? AND is_read = 0
               AND (expires_at IS NULL OR expires_at > datetime('now'))
               ORDER BY created_at DESC LIMIT 20""",
            (user_id,)
        )

    def mark_read(self, notification_id: int) -> bool:
        """Mark a notification as read."""
        rows = self.db.update('notifications', {
            'is_read': 1,
            'read_at': datetime.datetime.now().isoformat()
        }, "id = ?", (notification_id,))
        return rows > 0

    def mark_all_read(self, user_id: int) -> int:
        """Mark all notifications as read for a user."""
        cursor = self.db.execute(
            "UPDATE notifications SET is_read = 1, read_at = datetime('now') WHERE user_id = ? AND is_read = 0",
            (user_id,)
        )
        return cursor.rowcount

    def get_unread_count(self, user_id: int) -> int:
        """Get count of unread notifications."""
        return self.db.fetchscalar(
            "SELECT COUNT(*) FROM notifications WHERE user_id = ? AND is_read = 0",
            (user_id,)
        ) or 0


# =============================================================================
# SECTION 7: Business Engines
# =============================================================================


class GOSIEngine(BaseService):
    """Engine for General Organization for Social Insurance (GOSI) calculations.

    Handles employer and employee contribution calculations, GOSI base
    determination, and periodic GOSI reporting for Saudi and non-Saudi
    employees according to Saudi labor regulations.
    """

    def get_gosi_rates(self, is_saudi: bool) -> dict:
        """Return the applicable GOSI contribution rates.

        Args:
            is_saudi: Whether the employee holds Saudi nationality.

        Returns:
            Dict with employer_rate, employee_rate, and salary ceiling.
        """
        if is_saudi:
            return {
                'employer_rate': GOSI_SAUDI_EMPLOYER_RATE,
                'employee_rate': GOSI_SAUDI_EMPLOYEE_RATE,
                'ceiling': GOSI_SALARY_CEILING,
            }
        return {
            'employer_rate': GOSI_NON_SAUDI_EMPLOYER_RATE,
            'employee_rate': GOSI_NON_SAUDI_EMPLOYEE_RATE,
            'ceiling': GOSI_SALARY_CEILING,
        }

    def _compute_gosi_base(self, basic_salary: float,
                           housing_allowance: float) -> float:
        """Compute the GOSI-eligible base salary capped at the ceiling.

        GOSI base is the sum of basic salary and housing allowance, but
        must not exceed the regulatory ceiling.
        """
        raw = basic_salary + housing_allowance
        return min(raw, GOSI_SALARY_CEILING)

    def calculate_gosi(self, employee_id: int) -> dict:
        """Calculate GOSI contributions for an employee by looking up their record.

        Args:
            employee_id: Primary key in the employees table.

        Returns:
            Dict containing gosi_base, employer_contribution,
            employee_contribution, total, and rate metadata.

        Raises:
            ValueError: If the employee is not found.
        """
        employee = self.db.fetchone(
            "SELECT id, emp_number, first_name, last_name, basic_salary, "
            "housing_allowance, is_saudi, gosi_eligible FROM employees "
            "WHERE id = ?",
            (employee_id,),
        )
        if not employee:
            raise ValueError(f"Employee {employee_id} not found")

        return self.calculate_gosi_for_payroll(employee)

    def calculate_gosi_for_payroll(self, employee_data: dict) -> dict:
        """Calculate GOSI contributions from an already-loaded employee dict.

        Expected keys: basic_salary, housing_allowance, is_saudi,
        gosi_eligible (optional, defaults to True).

        Returns:
            Dict with gosi_base, employer_contribution,
            employee_contribution, total_contribution, and rates.
        """
        gosi_eligible = employee_data.get('gosi_eligible', 1)
        if not gosi_eligible:
            return {
                'gosi_base': 0.0,
                'employer_contribution': 0.0,
                'employee_contribution': 0.0,
                'total_contribution': 0.0,
                'employer_rate': 0.0,
                'employee_rate': 0.0,
                'is_saudi': bool(employee_data.get('is_saudi', 0)),
                'is_eligible': False,
            }

        is_saudi = bool(employee_data.get('is_saudi', 0))
        basic = float(employee_data.get('basic_salary', 0))
        housing = float(employee_data.get('housing_allowance', 0))
        rates = self.get_gosi_rates(is_saudi)
        gosi_base = self._compute_gosi_base(basic, housing)

        employer_contrib = round(gosi_base * rates['employer_rate'], 2)
        employee_contrib = round(gosi_base * rates['employee_rate'], 2)

        return {
            'gosi_base': gosi_base,
            'employer_contribution': employer_contrib,
            'employee_contribution': employee_contrib,
            'total_contribution': round(employer_contrib + employee_contrib, 2),
            'employer_rate': rates['employer_rate'],
            'employee_rate': rates['employee_rate'],
            'is_saudi': is_saudi,
            'is_eligible': True,
        }

    def generate_gosi_report(self, company_id: int, period: str) -> dict:
        """Generate an aggregate GOSI report for a company and period.

        Args:
            company_id: The company to report on.
            period: Period string in ``YYYY-MM`` format.

        Returns:
            Dict with period metadata, per-employee breakdowns, and totals.
        """
        try:
            year, month = period.split('-')
            year, month = int(year), int(month)
        except (ValueError, AttributeError):
            raise ValueError(f"Invalid period format '{period}', expected YYYY-MM")

        employees = self.db.fetchall(
            "SELECT id, emp_number, first_name, last_name, basic_salary, "
            "housing_allowance, is_saudi, gosi_eligible FROM employees "
            "WHERE company_id = ? AND is_active = 1 AND gosi_eligible = 1",
            (company_id,),
        )

        details: List[dict] = []
        total_employer = 0.0
        total_employee = 0.0
        saudi_count = 0
        non_saudi_count = 0

        for emp in employees:
            calc = self.calculate_gosi_for_payroll(emp)
            detail = {
                'employee_id': emp['id'],
                'emp_number': emp.get('emp_number', ''),
                'name': f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip(),
                'is_saudi': calc['is_saudi'],
                **calc,
            }
            details.append(detail)
            total_employer += calc['employer_contribution']
            total_employee += calc['employee_contribution']
            if calc['is_saudi']:
                saudi_count += 1
            else:
                non_saudi_count += 1

        return {
            'company_id': company_id,
            'period': period,
            'year': year,
            'month': month,
            'total_employees': len(details),
            'saudi_count': saudi_count,
            'non_saudi_count': non_saudi_count,
            'total_employer_contribution': round(total_employer, 2),
            'total_employee_contribution': round(total_employee, 2),
            'total_contribution': round(total_employer + total_employee, 2),
            'details': details,
            'generated_at': datetime.datetime.now().isoformat(),
        }


class EOSBEngine(BaseService):
    """Engine for End-of-Service Benefits (EOSB) calculations.

    Implements the Saudi Labor Law formula:
    - First 5 years of service: half-month basic salary per year.
    - Years beyond 5: full-month basic salary per year.
    Partial years are pro-rated.
    """

    def _service_years(self, hire_date_str: str,
                       termination_date_str: str = None) -> float:
        """Return total years of service as a float (fractional years)."""
        try:
            hire = datetime.datetime.strptime(hire_date_str[:10], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            raise ValueError(f"Invalid hire_date: {hire_date_str}")

        if termination_date_str:
            try:
                end = datetime.datetime.strptime(
                    termination_date_str[:10], '%Y-%m-%d'
                ).date()
            except (ValueError, TypeError):
                end = datetime.date.today()
        else:
            end = datetime.date.today()

        if end < hire:
            return 0.0

        delta = end - hire
        return round(delta.days / 365.25, 4)

    def _calculate_benefit(self, basic_salary: float,
                           years_of_service: float) -> float:
        """Core EOSB computation.

        - First 5 years: 0.5 × monthly basic per year
        - After 5 years: 1.0 × monthly basic per year
        Partial years are pro-rated proportionally.
        """
        if years_of_service <= 0 or basic_salary <= 0:
            return 0.0

        half_month_salary = basic_salary / 2.0
        full_month_salary = basic_salary

        if years_of_service <= EOSB_HALF_MONTH_YEARS:
            return round(half_month_salary * years_of_service, 2)

        first_part = half_month_salary * EOSB_HALF_MONTH_YEARS
        remaining_years = years_of_service - EOSB_HALF_MONTH_YEARS
        second_part = full_month_salary * remaining_years
        return round(first_part + second_part, 2)

    def calculate_eosb(self, employee_id: int) -> dict:
        """Calculate EOSB for an employee by ID.

        Args:
            employee_id: Primary key of the employee record.

        Returns:
            Dict with service details and computed benefit amount.

        Raises:
            ValueError: If the employee is not found.
        """
        employee = self.db.fetchone(
            "SELECT id, emp_number, first_name, last_name, basic_salary, "
            "hire_date, termination_date, is_active FROM employees WHERE id = ?",
            (employee_id,),
        )
        if not employee:
            raise ValueError(f"Employee {employee_id} not found")

        return self.calculate_eosb_from_data(
            basic_salary=float(employee.get('basic_salary', 0)),
            hire_date=employee.get('hire_date', ''),
            termination_date=employee.get('termination_date'),
        )

    def calculate_eosb_from_data(self, basic_salary: float,
                                 hire_date: str,
                                 termination_date: str = None) -> dict:
        """Calculate EOSB from raw data without a database lookup.

        Args:
            basic_salary: Monthly basic salary.
            hire_date: ISO date string of the hire date.
            termination_date: Optional ISO date string; defaults to today.

        Returns:
            Dict with years_of_service, basic_salary, and eosb_amount.
        """
        years = self._service_years(hire_date, termination_date)
        benefit = self._calculate_benefit(basic_salary, years)

        end_date = termination_date if termination_date else datetime.date.today().isoformat()

        return {
            'hire_date': hire_date,
            'end_date': end_date,
            'years_of_service': round(years, 2),
            'basic_salary': basic_salary,
            'half_month_years': min(years, EOSB_HALF_MONTH_YEARS),
            'full_month_years': max(0, round(years - EOSB_FULL_MONTH_START, 2)),
            'eosb_amount': benefit,
        }

    def generate_eosb_report(self, company_id: int) -> dict:
        """Generate a company-wide EOSB liability report.

        Calculates the current EOSB provision for every active employee,
        useful for financial reporting and accrual accounting.

        Args:
            company_id: The company to report on.

        Returns:
            Dict with total_liability, employee count, and per-employee details.
        """
        employees = self.db.fetchall(
            "SELECT id, emp_number, first_name, last_name, basic_salary, "
            "hire_date, termination_date, is_active FROM employees "
            "WHERE company_id = ? AND is_active = 1",
            (company_id,),
        )

        details: List[dict] = []
        total_liability = 0.0

        for emp in employees:
            try:
                calc = self.calculate_eosb_from_data(
                    basic_salary=float(emp.get('basic_salary', 0)),
                    hire_date=emp.get('hire_date', ''),
                    termination_date=emp.get('termination_date'),
                )
                detail = {
                    'employee_id': emp['id'],
                    'emp_number': emp.get('emp_number', ''),
                    'name': f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip(),
                    **calc,
                }
                details.append(detail)
                total_liability += calc['eosb_amount']
            except Exception as exc:
                self._logger.warning(
                    "Failed to calculate EOSB for employee %s: %s",
                    emp.get('id'), exc,
                )

        return {
            'company_id': company_id,
            'total_employees': len(details),
            'total_liability': round(total_liability, 2),
            'details': details,
            'generated_at': datetime.datetime.now().isoformat(),
        }


class WPSEngine(BaseService):
    """Engine for Wage Protection System (WPS) SIF file generation.

    The SIF (Salary Information File) format is used by Saudi banks to
    process payroll transfers in compliance with the Ministry of Human
    Resources' Wage Protection System.
    """

    # SIF record type constants
    RECORD_TYPE_HEADER = 'HDR'
    RECORD_TYPE_DETAIL = 'DTL'
    RECORD_TYPE_TRAILER = 'TRL'
    PAYMENT_TYPE_SALARY = 'SAL'

    def _pad_right(self, value: str, length: int) -> str:
        """Pad a string with spaces on the right to the given length."""
        return str(value)[:length].ljust(length)

    def _pad_left_zero(self, value: Union[int, float, str], length: int) -> str:
        """Pad a numeric value with leading zeros to the given length."""
        s = str(int(value))
        if len(s) > length:
            self._logger.warning(
                "Value %s exceeds %d digits; truncating", value, length,
            )
        return s.zfill(length)[-length:]

    def _format_amount(self, amount: float) -> str:
        """Format a monetary amount as a 15-character zero-padded string (2 decimal implied)."""
        cents = int(round(amount * 100))
        return str(max(cents, 0)).zfill(15)

    def format_sif_record(self, entry: dict) -> str:
        """Format a single payroll entry into a SIF detail record line.

        Args:
            entry: Dict with employee_id, emp_number, bank_iban,
                   bank_account, net_salary, basic_salary,
                   housing_allowance, other_allowances, total_deductions.

        Returns:
            Fixed-width string representing one SIF detail line.
        """
        parts = [
            self.RECORD_TYPE_DETAIL,
            self._pad_right(str(entry.get('emp_number', '')), 15),
            self._pad_right(str(entry.get('bank_iban', '')), 24),
            self._pad_right(str(entry.get('bank_account', '')), 20),
            self._format_amount(float(entry.get('net_salary', 0))),
            self._format_amount(float(entry.get('basic_salary', 0))),
            self._format_amount(float(entry.get('housing_allowance', 0))),
            self._format_amount(float(entry.get('other_allowances', 0))),
            self._format_amount(float(entry.get('total_deductions', 0))),
            self._pad_right(self.PAYMENT_TYPE_SALARY, 3),
        ]
        return '|'.join(parts)

    def validate_sif_data(self, payroll_entries: List[dict]) -> dict:
        """Validate payroll entries before SIF generation.

        Checks for missing IBANs, negative amounts, and other data
        integrity issues.

        Args:
            payroll_entries: List of payroll entry dicts.

        Returns:
            Dict with is_valid flag, error list, and warning list.
        """
        errors: List[str] = []
        warnings: List[str] = []

        if not payroll_entries:
            errors.append("No payroll entries provided")
            return {'is_valid': False, 'errors': errors, 'warnings': warnings}

        for idx, entry in enumerate(payroll_entries):
            emp_ref = entry.get('emp_number', entry.get('employee_id', f'row-{idx}'))

            iban = str(entry.get('bank_iban', '') or '')
            if not iban:
                errors.append(f"Employee {emp_ref}: missing bank IBAN")
            elif not iban.startswith('SA') or len(iban) != 24:
                warnings.append(f"Employee {emp_ref}: IBAN '{iban}' may be invalid")

            net = float(entry.get('net_salary', 0))
            if net <= 0:
                errors.append(f"Employee {emp_ref}: net salary is {net}")

            basic = float(entry.get('basic_salary', 0))
            if basic <= 0:
                warnings.append(f"Employee {emp_ref}: basic salary is {basic}")

        return {
            'is_valid': len(errors) == 0,
            'errors': errors,
            'warnings': warnings,
            'total_records': len(payroll_entries),
        }

    def generate_sif_file(self, payroll_run_id: int) -> str:
        """Generate a complete SIF file for a payroll run.

        Fetches entries from the database, validates them, and produces a
        multi-line SIF string including header, detail records, and trailer.

        Args:
            payroll_run_id: The ID of the payroll run.

        Returns:
            The full SIF file content as a string.

        Raises:
            ValueError: If the payroll run is not found or has no entries.
        """
        run = self.db.fetchone(
            "SELECT pr.id, pr.period_id, pr.run_number, pr.status, "
            "pp.company_id, pp.period_year, pp.period_month "
            "FROM payroll_runs pr "
            "JOIN payroll_periods pp ON pr.period_id = pp.id "
            "WHERE pr.id = ?",
            (payroll_run_id,),
        )
        if not run:
            raise ValueError(f"Payroll run {payroll_run_id} not found")

        entries = self.db.fetchall(
            "SELECT pe.*, e.emp_number, e.bank_name, e.bank_account, "
            "e.bank_iban, e.first_name, e.last_name "
            "FROM payroll_entries pe "
            "JOIN employees e ON pe.employee_id = e.id "
            "WHERE pe.payroll_run_id = ?",
            (payroll_run_id,),
        )
        if not entries:
            raise ValueError(f"No payroll entries found for run {payroll_run_id}")

        validation = self.validate_sif_data(entries)
        if not validation['is_valid']:
            self._logger.error("SIF validation failed: %s", validation['errors'])
            raise ValueError(
                f"SIF validation failed with {len(validation['errors'])} error(s): "
                + "; ".join(validation['errors'][:5])
            )

        company = self.db.fetchone(
            "SELECT name, commercial_registration FROM companies WHERE id = ?",
            (run.get('company_id'),),
        )

        now = datetime.datetime.now()
        file_ref = f"SIF{run.get('company_id', 0):04d}{now.strftime('%Y%m%d%H%M%S')}"

        # Header
        header_parts = [
            self.RECORD_TYPE_HEADER,
            self._pad_right(file_ref, 20),
            self._pad_right(str(company.get('name', '') if company else ''), 40),
            self._pad_right(
                str(company.get('commercial_registration', '') if company else ''), 15
            ),
            now.strftime('%Y%m%d'),
            now.strftime('%H%M%S'),
            self._pad_left_zero(len(entries), 6),
        ]
        header = '|'.join(header_parts)

        # Detail records
        detail_lines = [self.format_sif_record(e) for e in entries]

        # Trailer
        total_net = sum(float(e.get('net_salary', 0)) for e in entries)
        trailer_parts = [
            self.RECORD_TYPE_TRAILER,
            self._pad_left_zero(len(entries), 6),
            self._format_amount(total_net),
        ]
        trailer = '|'.join(trailer_parts)

        lines = [header] + detail_lines + [trailer]
        sif_content = '\n'.join(lines)

        self._logger.info(
            "Generated SIF file for run %s with %d records, total %.2f",
            payroll_run_id, len(entries), total_net,
        )
        return sif_content


class ZATCAEngine(BaseService):
    """Engine for ZATCA (Zakat, Tax and Customs Authority) e-invoicing.

    Implements TLV (Tag-Length-Value) encoding for generating ZATCA-compliant
    QR codes on electronic invoices as required by Saudi e-invoicing
    regulations (Fatoora).
    """

    # TLV tag assignments per ZATCA specification
    TAG_SELLER_NAME = 1
    TAG_VAT_NUMBER = 2
    TAG_TIMESTAMP = 3
    TAG_TOTAL_WITH_VAT = 4
    TAG_VAT_TOTAL = 5

    def encode_tlv(self, tag: int, value: str) -> bytes:
        """Encode a single TLV (Tag-Length-Value) element.

        Args:
            tag: Integer tag identifier (1-5 for ZATCA Phase 1).
            value: The string value to encode.

        Returns:
            Bytes representing the TLV-encoded element.
        """
        value_bytes = str(value).encode('utf-8')
        length = len(value_bytes)
        return bytes([tag, length]) + value_bytes

    def generate_qr_data(self, seller_name: str, vat_number: str,
                         timestamp: str, total: float,
                         vat_total: float) -> str:
        """Generate a ZATCA-compliant QR code data string (base64).

        Encodes the five mandatory TLV fields and returns the result as
        a base64-encoded string suitable for embedding in a QR code.

        Args:
            seller_name: The seller/company name.
            vat_number: VAT registration number.
            timestamp: ISO-8601 formatted timestamp of the invoice.
            total: Invoice total amount including VAT.
            vat_total: Total VAT amount.

        Returns:
            Base64-encoded string of the concatenated TLV data.
        """
        tlv_data = b''
        tlv_data += self.encode_tlv(self.TAG_SELLER_NAME, seller_name)
        tlv_data += self.encode_tlv(self.TAG_VAT_NUMBER, vat_number)
        tlv_data += self.encode_tlv(self.TAG_TIMESTAMP, timestamp)
        tlv_data += self.encode_tlv(self.TAG_TOTAL_WITH_VAT, f"{total:.2f}")
        tlv_data += self.encode_tlv(self.TAG_VAT_TOTAL, f"{vat_total:.2f}")
        return base64.b64encode(tlv_data).decode('ascii')

    def generate_invoice_qr(self, invoice_id: int) -> str:
        """Generate a QR code data string for a persisted invoice.

        Looks up the invoice and its associated company, then delegates
        to :meth:`generate_qr_data`.

        Args:
            invoice_id: Primary key of the invoice.

        Returns:
            Base64-encoded QR string.

        Raises:
            ValueError: If the invoice or company is not found.
        """
        invoice = self.db.fetchone(
            "SELECT id, company_id, invoice_number, invoice_date, "
            "total_amount, vat_amount FROM invoices WHERE id = ?",
            (invoice_id,),
        )
        if not invoice:
            raise ValueError(f"Invoice {invoice_id} not found")

        company = self.db.fetchone(
            "SELECT name, vat_number FROM companies WHERE id = ?",
            (invoice.get('company_id'),),
        )
        if not company:
            raise ValueError(
                f"Company {invoice.get('company_id')} not found for invoice {invoice_id}"
            )

        return self.generate_qr_data(
            seller_name=company.get('name', ''),
            vat_number=company.get('vat_number', ''),
            timestamp=invoice.get('invoice_date', datetime.datetime.now().isoformat()),
            total=float(invoice.get('total_amount', 0)),
            vat_total=float(invoice.get('vat_amount', 0)),
        )

    def generate_qr_code(self, invoice: dict) -> str:
        """Generate a ZATCA QR string from an invoice dict.

        This is the primary method used by ``InvoiceService`` when creating
        or updating invoices. It expects the invoice dict to carry either
        direct seller/VAT fields or a nested ``company`` dict.

        Args:
            invoice: Dict with invoice data. Recognised keys include
                seller_name / company.name, vat_number / company.vat_number,
                invoice_date, total_amount, and vat_amount.

        Returns:
            Base64-encoded QR data string, or empty string on failure.
        """
        try:
            company_data = invoice.get('company', {})
            seller_name = (
                invoice.get('seller_name')
                or company_data.get('name', '')
            )
            vat_number = (
                invoice.get('vat_number')
                or company_data.get('vat_number', '')
            )
            timestamp = invoice.get(
                'invoice_date', datetime.datetime.now().isoformat()
            )
            total = float(invoice.get('total_amount', 0))
            vat_total = float(invoice.get('vat_amount', 0))

            return self.generate_qr_data(
                seller_name=seller_name,
                vat_number=vat_number,
                timestamp=timestamp,
                total=total,
                vat_total=vat_total,
            )
        except Exception as exc:
            self._logger.error("Failed to generate QR code: %s", exc)
            return ''

    def validate_qr(self, qr_base64: str) -> dict:
        """Decode and validate a ZATCA QR base64 string.

        Parses the TLV-encoded data and checks that all five mandatory
        fields are present.

        Args:
            qr_base64: Base64-encoded QR data string.

        Returns:
            Dict with is_valid flag, decoded fields, and any errors.
        """
        result: Dict[str, Any] = {
            'is_valid': False,
            'fields': {},
            'errors': [],
        }

        if not qr_base64:
            result['errors'].append("Empty QR data")
            return result

        try:
            raw = base64.b64decode(qr_base64)
        except Exception:
            result['errors'].append("Invalid base64 encoding")
            return result

        tag_names = {
            self.TAG_SELLER_NAME: 'seller_name',
            self.TAG_VAT_NUMBER: 'vat_number',
            self.TAG_TIMESTAMP: 'timestamp',
            self.TAG_TOTAL_WITH_VAT: 'total_with_vat',
            self.TAG_VAT_TOTAL: 'vat_total',
        }

        offset = 0
        while offset < len(raw):
            if offset + 2 > len(raw):
                result['errors'].append("Truncated TLV data")
                break
            tag = raw[offset]
            length = raw[offset + 1]
            offset += 2

            if offset + length > len(raw):
                result['errors'].append(f"TLV tag {tag}: declared length exceeds data")
                break

            value = raw[offset:offset + length].decode('utf-8', errors='replace')
            offset += length

            field_name = tag_names.get(tag, f'unknown_tag_{tag}')
            result['fields'][field_name] = value

        required = set(tag_names.values())
        found = set(result['fields'].keys())
        missing = required - found
        if missing:
            result['errors'].append(f"Missing required fields: {', '.join(sorted(missing))}")

        result['is_valid'] = len(result['errors']) == 0
        return result


class PayrollEngine(BaseService):
    """Comprehensive payroll processing engine.

    Orchestrates the full payroll lifecycle: creating runs, calculating
    gross and net salaries, applying GOSI contributions and deductions,
    and finalising payroll for posting.
    """

    def __init__(self, db: 'DatabaseManager'):
        super().__init__(db)
        self.gosi_engine = GOSIEngine(db)

    # ------------------------------------------------------------------
    # Run lifecycle
    # ------------------------------------------------------------------

    def create_payroll_run(self, company_id: int, period_id: int,
                           run_by: int = None) -> int:
        """Create a new payroll run record.

        Args:
            company_id: Company to run payroll for.
            period_id: The payroll period to process.
            run_by: User ID who initiated the run (optional).

        Returns:
            The new payroll_run ID.

        Raises:
            ValueError: If the period is not found or already processed.
        """
        period = self.db.fetchone(
            "SELECT id, company_id, period_year, period_month, status "
            "FROM payroll_periods WHERE id = ? AND company_id = ?",
            (period_id, company_id),
        )
        if not period:
            raise ValueError(f"Payroll period {period_id} not found for company {company_id}")

        if period.get('status') == 'posted':
            raise ValueError(f"Payroll period {period_id} is already posted")

        existing_count = self.db.fetchscalar(
            "SELECT COUNT(*) FROM payroll_runs WHERE period_id = ?",
            (period_id,),
        ) or 0

        run_id = self.db.insert('payroll_runs', {
            'period_id': period_id,
            'run_number': existing_count + 1,
            'run_type': 'regular',
            'status': 'draft',
            'total_employees': 0,
            'total_gross': 0,
            'total_net': 0,
            'run_by': run_by,
            'run_at': datetime.datetime.now().isoformat(),
            'notes': '',
        })
        self._logger.info("Created payroll run %s for period %s", run_id, period_id)
        return run_id

    def process_payroll(self, run_id: int) -> dict:
        """Process payroll for all active employees in the run's company.

        Calculates salaries, GOSI, and deductions for each employee, then
        persists the results to ``payroll_entries``.

        Args:
            run_id: The payroll run to process.

        Returns:
            Summary dict with counts and totals.
        """
        run = self.db.fetchone(
            "SELECT pr.id, pr.period_id, pr.status, pp.company_id, "
            "pp.period_year, pp.period_month, pp.start_date, pp.end_date "
            "FROM payroll_runs pr "
            "JOIN payroll_periods pp ON pr.period_id = pp.id "
            "WHERE pr.id = ?",
            (run_id,),
        )
        if not run:
            raise ValueError(f"Payroll run {run_id} not found")

        if run.get('status') not in ('draft', 'processing'):
            raise ValueError(f"Payroll run {run_id} cannot be processed (status: {run['status']})")

        self.db.execute(
            "UPDATE payroll_runs SET status = 'processing' WHERE id = ?",
            (run_id,),
        )

        employees = self.db.fetchall(
            "SELECT * FROM employees WHERE company_id = ? AND is_active = 1",
            (run['company_id'],),
        )

        period_data = {
            'period_id': run['period_id'],
            'period_year': run.get('period_year'),
            'period_month': run.get('period_month'),
            'start_date': run.get('start_date'),
            'end_date': run.get('end_date'),
        }

        processed = 0
        total_gross = 0.0
        total_net = 0.0
        errors: List[str] = []

        for emp in employees:
            try:
                entry = self.calculate_employee_payroll(emp, period_data)
                entry['payroll_run_id'] = run_id
                entry['employee_id'] = emp['id']
                entry['period_id'] = period_data['period_id']
                entry['period_year'] = period_data.get('period_year')
                entry['period_month'] = period_data.get('period_month')
                entry['status'] = 'calculated'

                self.db.insert('payroll_entries', entry)
                processed += 1
                total_gross += entry.get('gross_salary', 0)
                total_net += entry.get('net_salary', 0)
            except Exception as exc:
                self._logger.error(
                    "Payroll error for employee %s: %s", emp.get('id'), exc,
                )
                errors.append(f"Employee {emp.get('emp_number', emp['id'])}: {exc}")

        new_status = 'calculated' if not errors else 'partial'
        self.db.execute(
            "UPDATE payroll_runs SET status = ?, total_employees = ?, "
            "total_gross = ?, total_net = ? WHERE id = ?",
            (new_status, processed, round(total_gross, 2),
             round(total_net, 2), run_id),
        )

        return {
            'run_id': run_id,
            'status': new_status,
            'total_employees': processed,
            'total_gross': round(total_gross, 2),
            'total_net': round(total_net, 2),
            'errors': errors,
        }

    # ------------------------------------------------------------------
    # Per-employee calculations
    # ------------------------------------------------------------------

    def calculate_gross_salary(self, employee: dict) -> float:
        """Sum all salary components to obtain the gross salary.

        Args:
            employee: Dict with basic_salary, housing_allowance,
                transport_allowance, food_allowance, other_allowances.

        Returns:
            The gross salary as a float.
        """
        components = [
            float(employee.get('basic_salary', 0)),
            float(employee.get('housing_allowance', 0)),
            float(employee.get('transport_allowance', 0)),
            float(employee.get('food_allowance', 0)),
            float(employee.get('other_allowances', 0)),
        ]
        return round(sum(components), 2)

    def calculate_employee_payroll(self, employee: dict,
                                   period_data: dict) -> dict:
        """Calculate payroll for a single employee.

        Computes gross salary, GOSI contributions, and deductions to
        arrive at the net salary.

        Args:
            employee: Employee record dict from the database.
            period_data: Period metadata (period_id, year, month, dates).

        Returns:
            Dict of all payroll entry fields ready for insertion.
        """
        basic = float(employee.get('basic_salary', 0))
        housing = float(employee.get('housing_allowance', 0))
        transport = float(employee.get('transport_allowance', 0))
        food = float(employee.get('food_allowance', 0))
        other = float(employee.get('other_allowances', 0))
        gross = self.calculate_gross_salary(employee)

        gosi = self.gosi_engine.calculate_gosi_for_payroll(employee)

        entry_data = {
            'basic_salary': basic,
            'housing_allowance': housing,
            'transport_allowance': transport,
            'food_allowance': food,
            'other_allowances': other,
            'overtime_amount': 0.0,
            'bonus_amount': 0.0,
            'gross_salary': gross,
            'gosi_employee': gosi.get('employee_contribution', 0),
            'gosi_employer': gosi.get('employer_contribution', 0),
            'gosi_base': gosi.get('gosi_base', 0),
            'working_days': 30,
            'actual_days': 30,
            'payment_method': 'bank_transfer',
            'bank_iban': employee.get('bank_iban', ''),
        }

        entry_data = self.apply_deductions(entry_data, employee)
        return entry_data

    def apply_deductions(self, entry_data: dict, employee: dict) -> dict:
        """Apply loan, advance, absence, and other deductions.

        Queries outstanding loans and advances from the database, then
        calculates the total deductions and net salary.

        Args:
            entry_data: The in-progress payroll entry dict.
            employee: The employee record dict.

        Returns:
            Updated entry_data with deduction fields and net_salary.
        """
        employee_id = employee.get('id')
        loan_deduction = 0.0
        advance_deduction = 0.0

        if employee_id:
            active_loans = self.db.fetchall(
                "SELECT id, installment_amount, outstanding_amount "
                "FROM employee_loans "
                "WHERE employee_id = ? AND status = 'active' "
                "AND outstanding_amount > 0",
                (employee_id,),
            )
            for loan in active_loans:
                installment = float(loan.get('installment_amount', 0))
                outstanding = float(loan.get('outstanding_amount', 0))
                amount = min(installment, outstanding)
                loan_deduction += amount

        entry_data['loan_deduction'] = round(loan_deduction, 2)
        entry_data['advance_deduction'] = round(advance_deduction, 2)
        entry_data['absence_deduction'] = 0.0
        entry_data['other_deductions'] = 0.0

        total_deductions = (
            entry_data.get('gosi_employee', 0)
            + entry_data['loan_deduction']
            + entry_data['advance_deduction']
            + entry_data['absence_deduction']
            + entry_data['other_deductions']
        )
        entry_data['total_deductions'] = round(total_deductions, 2)
        entry_data['net_salary'] = round(
            entry_data.get('gross_salary', 0) - total_deductions, 2
        )

        return entry_data

    # ------------------------------------------------------------------
    # Finalisation & reporting
    # ------------------------------------------------------------------

    def finalize_payroll_run(self, run_id: int) -> bool:
        """Finalise a payroll run, marking it ready for posting.

        Updates the run status and locks all associated entries.

        Args:
            run_id: The payroll run to finalise.

        Returns:
            True if finalisation succeeded.

        Raises:
            ValueError: If the run is in an invalid state.
        """
        run = self.db.fetchone(
            "SELECT id, status FROM payroll_runs WHERE id = ?", (run_id,),
        )
        if not run:
            raise ValueError(f"Payroll run {run_id} not found")

        if run.get('status') not in ('calculated', 'partial'):
            raise ValueError(
                f"Payroll run {run_id} cannot be finalised (status: {run['status']})"
            )

        with self.db.transaction():
            self.db.execute(
                "UPDATE payroll_runs SET status = 'finalized', "
                "posted_at = ? WHERE id = ?",
                (datetime.datetime.now().isoformat(), run_id),
            )
            self.db.execute(
                "UPDATE payroll_entries SET status = 'finalized' "
                "WHERE payroll_run_id = ? AND status = 'calculated'",
                (run_id,),
            )

            # Deduct loan installments from outstanding balances
            entries = self.db.fetchall(
                "SELECT employee_id, loan_deduction FROM payroll_entries "
                "WHERE payroll_run_id = ? AND loan_deduction > 0",
                (run_id,),
            )
            for entry in entries:
                if entry['loan_deduction'] > 0:
                    self.db.execute(
                        "UPDATE employee_loans SET outstanding_amount = "
                        "MAX(outstanding_amount - ?, 0) "
                        "WHERE employee_id = ? AND status = 'active'",
                        (entry['loan_deduction'], entry['employee_id']),
                    )

        self._logger.info("Finalised payroll run %s", run_id)
        return True

    def get_payroll_summary(self, run_id: int) -> dict:
        """Retrieve a summary of a payroll run including aggregated figures.

        Args:
            run_id: The payroll run to summarise.

        Returns:
            Dict with run metadata, totals, and per-entry details.

        Raises:
            ValueError: If the run is not found.
        """
        run = self.db.fetchone(
            "SELECT pr.*, pp.period_year, pp.period_month, pp.period_name "
            "FROM payroll_runs pr "
            "JOIN payroll_periods pp ON pr.period_id = pp.id "
            "WHERE pr.id = ?",
            (run_id,),
        )
        if not run:
            raise ValueError(f"Payroll run {run_id} not found")

        entries = self.db.fetchall(
            "SELECT pe.*, e.emp_number, e.first_name, e.last_name, e.is_saudi "
            "FROM payroll_entries pe "
            "JOIN employees e ON pe.employee_id = e.id "
            "WHERE pe.payroll_run_id = ?",
            (run_id,),
        )

        total_basic = sum(float(e.get('basic_salary', 0)) for e in entries)
        total_gross = sum(float(e.get('gross_salary', 0)) for e in entries)
        total_gosi_ee = sum(float(e.get('gosi_employee', 0)) for e in entries)
        total_gosi_er = sum(float(e.get('gosi_employer', 0)) for e in entries)
        total_deductions = sum(float(e.get('total_deductions', 0)) for e in entries)
        total_net = sum(float(e.get('net_salary', 0)) for e in entries)

        return {
            'run_id': run_id,
            'period_name': run.get('period_name', ''),
            'period_year': run.get('period_year'),
            'period_month': run.get('period_month'),
            'status': run.get('status'),
            'run_at': run.get('run_at'),
            'total_employees': len(entries),
            'total_basic': round(total_basic, 2),
            'total_gross': round(total_gross, 2),
            'total_gosi_employee': round(total_gosi_ee, 2),
            'total_gosi_employer': round(total_gosi_er, 2),
            'total_deductions': round(total_deductions, 2),
            'total_net': round(total_net, 2),
            'entries': entries,
        }


class SecurityManager:
    """Authentication and authorisation manager.

    Handles password hashing (bcrypt preferred, SHA-256 fallback), session
    management, and role-based access control. This class does **not**
    extend ``BaseService``; it takes a ``DatabaseManager`` directly.
    """

    SESSION_EXPIRY_HOURS = 24

    def __init__(self, db: 'DatabaseManager'):
        self.db = db
        self._logger = logging.getLogger('security.SecurityManager')

    # ------------------------------------------------------------------
    # Password hashing
    # ------------------------------------------------------------------

    def hash_password(self, password: str) -> str:
        """Hash a plaintext password.

        Uses bcrypt when available, otherwise falls back to salted SHA-256.

        Args:
            password: The plaintext password to hash.

        Returns:
            The hashed password string.
        """
        if not password:
            raise ValueError("Password cannot be empty")

        if HAS_BCRYPT:
            try:
                import bcrypt as _bcrypt
                hashed = _bcrypt.hashpw(
                    password.encode('utf-8'), _bcrypt.gensalt()
                )
                return hashed.decode('utf-8')
            except Exception as exc:
                self._logger.warning("bcrypt hashing failed, using fallback: %s", exc)

        salt = uuid.uuid4().hex
        hashed = hashlib.sha256((salt + password).encode('utf-8')).hexdigest()
        return f"sha256${salt}${hashed}"

    def verify_password(self, password: str, hashed: str) -> bool:
        """Verify a plaintext password against a stored hash.

        Supports both bcrypt and the sha256$salt$hash fallback format.

        Args:
            password: Plaintext password to verify.
            hashed: Stored hash to compare against.

        Returns:
            True if the password matches.
        """
        if not password or not hashed:
            return False

        if hashed.startswith('sha256$'):
            parts = hashed.split('$', 2)
            if len(parts) != 3:
                return False
            _, salt, stored_hash = parts
            check = hashlib.sha256((salt + password).encode('utf-8')).hexdigest()
            return check == stored_hash

        if HAS_BCRYPT:
            try:
                import bcrypt as _bcrypt
                return _bcrypt.checkpw(
                    password.encode('utf-8'), hashed.encode('utf-8')
                )
            except Exception:
                return False

        return False

    # ------------------------------------------------------------------
    # Session management
    # ------------------------------------------------------------------

    def create_session(self, user_id: int) -> str:
        """Create a new authenticated session for a user.

        Generates a cryptographically random session token and stores it
        in the database with an expiry time.

        Args:
            user_id: The user to create a session for.

        Returns:
            The session token string.
        """
        token = uuid.uuid4().hex + uuid.uuid4().hex
        now = datetime.datetime.now()
        expires = now + datetime.timedelta(hours=self.SESSION_EXPIRY_HOURS)

        self.db.insert('user_sessions', {
            'user_id': user_id,
            'session_token': token,
            'created_at': now.isoformat(),
            'expires_at': expires.isoformat(),
            'is_active': 1,
        })
        self._logger.info("Created session for user %s", user_id)
        return token

    def validate_session(self, token: str) -> Optional[dict]:
        """Validate a session token and return the associated user.

        Checks the token exists, is active, and has not expired.

        Args:
            token: The session token to validate.

        Returns:
            User dict if valid, None otherwise.
        """
        if not token:
            return None

        session = self.db.fetchone(
            "SELECT us.*, u.username, u.email, u.full_name, u.company_id, "
            "u.is_admin FROM user_sessions us "
            "JOIN users u ON us.user_id = u.id "
            "WHERE us.session_token = ? AND us.is_active = 1",
            (token,),
        )
        if not session:
            return None

        expires_at = session.get('expires_at', '')
        try:
            expires = datetime.datetime.fromisoformat(expires_at)
            if datetime.datetime.now() > expires:
                self.db.execute(
                    "UPDATE user_sessions SET is_active = 0 WHERE session_token = ?",
                    (token,),
                )
                return None
        except (ValueError, TypeError):
            return None

        return {
            'user_id': session.get('user_id'),
            'username': session.get('username'),
            'email': session.get('email'),
            'full_name': session.get('full_name'),
            'company_id': session.get('company_id'),
            'is_admin': bool(session.get('is_admin', 0)),
            'session_token': token,
            'expires_at': expires_at,
        }

    def invalidate_session(self, token: str) -> bool:
        """Invalidate (log out) a session.

        Args:
            token: The session token to invalidate.

        Returns:
            True if a session was found and invalidated.
        """
        rows = self.db.execute(
            "UPDATE user_sessions SET is_active = 0 WHERE session_token = ? AND is_active = 1",
            (token,),
        )
        return getattr(rows, 'rowcount', 0) > 0

    # ------------------------------------------------------------------
    # Role-based access control
    # ------------------------------------------------------------------

    def check_permission(self, user_id: int, permission: str) -> bool:
        """Check whether a user has a specific permission.

        Looks through all roles assigned to the user and checks if any
        of them grant the requested permission code.

        Args:
            user_id: The user to check.
            permission: The permission code string (e.g., ``payroll.run``).

        Returns:
            True if the user has the permission.
        """
        user = self.db.fetchone(
            "SELECT id, is_admin FROM users WHERE id = ? AND is_active = 1",
            (user_id,),
        )
        if not user:
            return False

        if user.get('is_admin'):
            return True

        count = self.db.fetchscalar(
            "SELECT COUNT(*) FROM user_roles ur "
            "JOIN role_permissions rp ON ur.role_id = rp.role_id "
            "JOIN permissions p ON rp.permission_id = p.id "
            "WHERE ur.user_id = ? AND p.code = ?",
            (user_id, permission),
        )
        return (count or 0) > 0

    def assign_role(self, user_id: int, role_id: int) -> bool:
        """Assign a role to a user.

        If the user already has the role, no duplicate is created.

        Args:
            user_id: The user to assign the role to.
            role_id: The role to assign.

        Returns:
            True if the role was assigned (or already existed).
        """
        existing = self.db.fetchone(
            "SELECT id FROM user_roles WHERE user_id = ? AND role_id = ?",
            (user_id, role_id),
        )
        if existing:
            self._logger.debug(
                "User %s already has role %s", user_id, role_id,
            )
            return True

        try:
            self.db.insert('user_roles', {
                'user_id': user_id,
                'role_id': role_id,
            })
            self._logger.info("Assigned role %s to user %s", role_id, user_id)
            return True
        except Exception as exc:
            self._logger.error(
                "Failed to assign role %s to user %s: %s",
                role_id, user_id, exc,
            )
            return False

    def remove_role(self, user_id: int, role_id: int) -> bool:
        """Remove a role from a user.

        Args:
            user_id: The user to remove the role from.
            role_id: The role to remove.

        Returns:
            True if the role was removed.
        """
        rows = self.db.delete('user_roles', "user_id = ? AND role_id = ?",
                              (user_id, role_id))
        return rows > 0

    def get_user_permissions(self, user_id: int) -> List[str]:
        """Retrieve all permission codes granted to a user via their roles.

        Args:
            user_id: The user whose permissions to retrieve.

        Returns:
            Sorted list of unique permission code strings.
        """
        rows = self.db.fetchall(
            "SELECT DISTINCT p.code FROM permissions p "
            "JOIN role_permissions rp ON p.id = rp.permission_id "
            "JOIN user_roles ur ON rp.role_id = ur.role_id "
            "WHERE ur.user_id = ?",
            (user_id,),
        )
        return sorted(row['code'] for row in rows if row.get('code'))

    def get_user_roles(self, user_id: int) -> List[dict]:
        """Retrieve all roles assigned to a user.

        Args:
            user_id: The user whose roles to retrieve.

        Returns:
            List of role dicts with id, name, and display_name.
        """
        return self.db.fetchall(
            "SELECT r.id, r.name, r.display_name, r.description "
            "FROM roles r "
            "JOIN user_roles ur ON r.id = ur.role_id "
            "WHERE ur.user_id = ?",
            (user_id,),
        )


# =============================================================================
# SECTION 8: REPORTING ENGINE
# =============================================================================


class ReportEngine(BaseService):
    """Comprehensive reporting engine for generating payroll, employee,
    invoice, and financial reports in PDF, Excel, and CSV formats.

    Supports ZATCA-compliant invoice rendering and standard Saudi
    financial statements (Trial Balance, Profit & Loss, Balance Sheet).
    """

    # Report output directory prefix
    _REPORT_PREFIX = "erp_report_"

    def __init__(self, db: DatabaseManager):
        super().__init__(db)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate_payroll_report(
        self, run_id: int, format: str = "pdf"
    ) -> str:
        """Generate a payroll report for a specific payroll run.

        Args:
            run_id: ID of the payroll run.
            format: Output format – 'pdf', 'excel', or 'csv'.

        Returns:
            Absolute path to the generated report file.

        Raises:
            ValueError: If the payroll run does not exist.
        """
        run = self.db.fetchone(
            "SELECT pr.id, pr.period_id, pr.status, pr.total_employees, "
            "pr.total_gross, pr.total_net "
            "FROM payroll_runs pr WHERE pr.id = ?",
            (run_id,),
        )
        if not run:
            raise ValueError(f"Payroll run {run_id} not found")

        period = self.db.fetchone(
            "SELECT period_name, period_year, period_month "
            "FROM payroll_periods WHERE id = ?",
            (run["period_id"],),
        )
        period_label = (
            period["period_name"]
            if period
            else f"Period {run['period_id']}"
        )

        entries = self.db.fetchall(
            "SELECT pe.employee_id, e.emp_number, "
            "e.first_name || ' ' || e.last_name AS employee_name, "
            "pe.basic_salary, pe.gross_salary, pe.net_salary, "
            "pe.gosi_employee, pe.gosi_employer, pe.total_deductions, "
            "COALESCE(pe.gross_salary - pe.basic_salary, 0) AS allowances "
            "FROM payroll_entries pe "
            "JOIN employees e ON pe.employee_id = e.id "
            "WHERE pe.payroll_run_id = ? "
            "ORDER BY e.emp_number",
            (run_id,),
        )

        headers = [
            "Emp #", "Employee Name", "Basic Salary", "Allowances",
            "Gross Salary", "GOSI (Emp)", "Deductions", "Net Salary",
        ]
        rows: List[list] = []
        totals = [Decimal(0)] * 6
        for e in entries:
            basic = Decimal(str(e["basic_salary"] or 0))
            allowances = Decimal(str(e["allowances"] or 0))
            gross = Decimal(str(e["gross_salary"] or 0))
            gosi_emp = Decimal(str(e["gosi_employee"] or 0))
            deductions = Decimal(str(e["total_deductions"] or 0))
            net = Decimal(str(e["net_salary"] or 0))
            rows.append([
                e["emp_number"],
                e["employee_name"],
                f"{basic:,.2f}",
                f"{allowances:,.2f}",
                f"{gross:,.2f}",
                f"{gosi_emp:,.2f}",
                f"{deductions:,.2f}",
                f"{net:,.2f}",
            ])
            totals[0] += basic
            totals[1] += allowances
            totals[2] += gross
            totals[3] += gosi_emp
            totals[4] += deductions
            totals[5] += net

        rows.append([
            "", "TOTALS",
            f"{totals[0]:,.2f}", f"{totals[1]:,.2f}",
            f"{totals[2]:,.2f}", f"{totals[3]:,.2f}",
            f"{totals[4]:,.2f}", f"{totals[5]:,.2f}",
        ])

        title = f"Payroll Report – {period_label}"
        filename = f"payroll_run_{run_id}"

        return self._render_report(
            headers, rows, title, filename, format,
            subtitle=f"Run #{run_id} | Status: {run['status']} | "
                     f"Employees: {run['total_employees']}",
        )

    def generate_employee_report(
        self, company_id: int, format: str = "pdf"
    ) -> str:
        """Generate an active-employee roster report for a company.

        Args:
            company_id: ID of the company.
            format: Output format – 'pdf', 'excel', or 'csv'.

        Returns:
            Absolute path to the generated report file.
        """
        employees = self.db.fetchall(
            "SELECT e.emp_number, "
            "e.first_name || ' ' || e.last_name AS employee_name, "
            "CASE WHEN e.is_saudi = 1 THEN 'Saudi' ELSE 'Non-Saudi' END "
            "  AS nationality, "
            "e.id_number, e.job_title, "
            "COALESCE(d.name, '') AS department, "
            "e.hire_date, e.basic_salary, "
            "COALESCE(e.total_salary, e.basic_salary) AS total_salary "
            "FROM employees e "
            "LEFT JOIN departments d ON e.department_id = d.id "
            "WHERE e.company_id = ? AND e.is_active = 1 "
            "ORDER BY e.emp_number",
            (company_id,),
        )

        headers = [
            "Emp #", "Name", "Nationality", "ID Number",
            "Job Title", "Department", "Hire Date",
            "Basic Salary", "Total Salary",
        ]
        rows: List[list] = []
        for emp in employees:
            rows.append([
                emp["emp_number"],
                emp["employee_name"],
                emp["nationality"],
                emp.get("id_number", ""),
                emp.get("job_title", ""),
                emp["department"],
                emp.get("hire_date", ""),
                f"{Decimal(str(emp['basic_salary'] or 0)):,.2f}",
                f"{Decimal(str(emp['total_salary'] or 0)):,.2f}",
            ])

        company = self.db.fetchone(
            "SELECT name FROM companies WHERE id = ?", (company_id,)
        )
        company_name = company["name"] if company else "Company"
        title = f"Employee Roster – {company_name}"
        filename = f"employees_company_{company_id}"

        return self._render_report(headers, rows, title, filename, format)

    def generate_invoice_report(
        self, invoice_id: int, format: str = "pdf"
    ) -> str:
        """Generate a formatted invoice document.

        Args:
            invoice_id: ID of the invoice.
            format: Output format – 'pdf', 'excel', or 'csv'.

        Returns:
            Absolute path to the generated report file.

        Raises:
            ValueError: If the invoice does not exist.
        """
        invoice = self.db.fetchone(
            "SELECT i.id, i.invoice_number, i.invoice_date, i.client_id, "
            "i.subtotal, i.vat_amount, i.total_amount, i.status, "
            "i.company_id "
            "FROM invoices i WHERE i.id = ?",
            (invoice_id,),
        )
        if not invoice:
            raise ValueError(f"Invoice {invoice_id} not found")

        client = self.db.fetchone(
            "SELECT name FROM clients WHERE id = ?",
            (invoice["client_id"],),
        )
        company = self.db.fetchone(
            "SELECT name, name_ar, vat_number, commercial_registration "
            "FROM companies WHERE id = ?",
            (invoice.get("company_id", 0),),
        )
        lines = self.db.fetchall(
            "SELECT description, quantity, unit_price, total "
            "FROM invoice_lines WHERE invoice_id = ?",
            (invoice_id,),
        )

        headers = ["#", "Description", "Qty", "Unit Price", "Total"]
        rows: List[list] = []
        for idx, line in enumerate(lines, 1):
            rows.append([
                str(idx),
                line["description"],
                str(line["quantity"]),
                f"{Decimal(str(line['unit_price'] or 0)):,.2f}",
                f"{Decimal(str(line['total'] or 0)):,.2f}",
            ])

        subtotal = Decimal(str(invoice["subtotal"] or 0))
        vat = Decimal(str(invoice["vat_amount"] or 0))
        total = Decimal(str(invoice["total_amount"] or 0))

        rows.append(["", "", "", "Subtotal", f"{subtotal:,.2f}"])
        rows.append(["", "", "", "VAT (15%)", f"{vat:,.2f}"])
        rows.append(["", "", "", "Total", f"{total:,.2f}"])

        company_name = company["name"] if company else "Company"
        client_name = client["name"] if client else "Client"

        if format == "pdf" and HAS_REPORTLAB:
            return self._generate_invoice_pdf(
                invoice, company, client_name, lines,
                subtotal, vat, total,
            )

        title = (
            f"Invoice {invoice['invoice_number']} – {company_name}"
        )
        filename = f"invoice_{invoice['invoice_number']}"
        return self._render_report(
            headers, rows, title, filename, format,
            subtitle=f"Client: {client_name} | Date: "
                     f"{invoice['invoice_date']} | "
                     f"Status: {invoice['status']}",
        )

    def generate_financial_report(
        self,
        company_id: int,
        report_type: str,
        start_date: str,
        end_date: str,
        format: str = "pdf",
    ) -> str:
        """Generate a financial statement.

        Args:
            company_id: ID of the company.
            report_type: One of 'trial_balance', 'profit_loss',
                         or 'balance_sheet'.
            start_date: Start of reporting period (YYYY-MM-DD).
            end_date: End of reporting period (YYYY-MM-DD).
            format: Output format – 'pdf', 'excel', or 'csv'.

        Returns:
            Absolute path to the generated report file.

        Raises:
            ValueError: If *report_type* is unrecognised.
        """
        valid_types = ("trial_balance", "profit_loss", "balance_sheet")
        if report_type not in valid_types:
            raise ValueError(
                f"Invalid report_type '{report_type}'. "
                f"Must be one of {valid_types}"
            )

        company = self.db.fetchone(
            "SELECT name FROM companies WHERE id = ?", (company_id,)
        )
        company_name = company["name"] if company else "Company"

        if report_type == "trial_balance":
            return self._generate_trial_balance(
                company_id, company_name, start_date, end_date, format,
            )
        elif report_type == "profit_loss":
            return self._generate_profit_loss(
                company_id, company_name, start_date, end_date, format,
            )
        else:
            return self._generate_balance_sheet(
                company_id, company_name, start_date, end_date, format,
            )

    # ------------------------------------------------------------------
    # Financial sub-reports
    # ------------------------------------------------------------------

    def _generate_trial_balance(
        self,
        company_id: int,
        company_name: str,
        start_date: str,
        end_date: str,
        format: str,
    ) -> str:
        """Build a trial-balance report."""
        rows_data = self.db.fetchall(
            "SELECT coa.account_code, coa.account_name, "
            "COALESCE(SUM(jel.debit), 0) AS total_debit, "
            "COALESCE(SUM(jel.credit), 0) AS total_credit "
            "FROM journal_entry_lines jel "
            "JOIN chart_of_accounts coa ON jel.account_id = coa.id "
            "JOIN journal_entries je ON jel.entry_id = je.id "
            "WHERE je.company_id = ? "
            "AND je.entry_date BETWEEN ? AND ? "
            "AND je.status = 'posted' "
            "GROUP BY coa.account_code, coa.account_name "
            "ORDER BY coa.account_code",
            (company_id, start_date, end_date),
        )

        headers = ["Account Code", "Account Name", "Debit", "Credit", "Balance"]
        rows: List[list] = []
        total_debit = Decimal(0)
        total_credit = Decimal(0)
        for r in rows_data:
            debit = Decimal(str(r["total_debit"]))
            credit = Decimal(str(r["total_credit"]))
            balance = debit - credit
            rows.append([
                r["account_code"],
                r["account_name"],
                f"{debit:,.2f}",
                f"{credit:,.2f}",
                f"{balance:,.2f}",
            ])
            total_debit += debit
            total_credit += credit

        rows.append([
            "", "TOTALS",
            f"{total_debit:,.2f}",
            f"{total_credit:,.2f}",
            f"{total_debit - total_credit:,.2f}",
        ])

        title = f"Trial Balance – {company_name}"
        filename = f"trial_balance_{company_id}_{start_date}_{end_date}"
        return self._render_report(
            headers, rows, title, filename, format,
            subtitle=f"Period: {start_date} to {end_date}",
        )

    def _generate_profit_loss(
        self,
        company_id: int,
        company_name: str,
        start_date: str,
        end_date: str,
        format: str,
    ) -> str:
        """Build a profit-and-loss report."""
        revenue_rows = self.db.fetchall(
            "SELECT coa.account_code, coa.account_name, "
            "COALESCE(SUM(jel.credit), 0) - COALESCE(SUM(jel.debit), 0) "
            "  AS net_amount "
            "FROM journal_entry_lines jel "
            "JOIN chart_of_accounts coa ON jel.account_id = coa.id "
            "JOIN journal_entries je ON jel.entry_id = je.id "
            "WHERE je.company_id = ? "
            "AND je.entry_date BETWEEN ? AND ? "
            "AND je.status = 'posted' "
            "AND coa.account_type = 'revenue' "
            "GROUP BY coa.account_code, coa.account_name "
            "ORDER BY coa.account_code",
            (company_id, start_date, end_date),
        )
        expense_rows = self.db.fetchall(
            "SELECT coa.account_code, coa.account_name, "
            "COALESCE(SUM(jel.debit), 0) - COALESCE(SUM(jel.credit), 0) "
            "  AS net_amount "
            "FROM journal_entry_lines jel "
            "JOIN chart_of_accounts coa ON jel.account_id = coa.id "
            "JOIN journal_entries je ON jel.entry_id = je.id "
            "WHERE je.company_id = ? "
            "AND je.entry_date BETWEEN ? AND ? "
            "AND je.status = 'posted' "
            "AND coa.account_type = 'expense' "
            "GROUP BY coa.account_code, coa.account_name "
            "ORDER BY coa.account_code",
            (company_id, start_date, end_date),
        )

        headers = ["Account Code", "Account Name", "Amount (SAR)"]
        rows: List[list] = []

        rows.append(["", "--- REVENUE ---", ""])
        total_revenue = Decimal(0)
        for r in revenue_rows:
            amt = Decimal(str(r["net_amount"]))
            rows.append([r["account_code"], r["account_name"], f"{amt:,.2f}"])
            total_revenue += amt
        rows.append(["", "Total Revenue", f"{total_revenue:,.2f}"])

        rows.append(["", "--- EXPENSES ---", ""])
        total_expenses = Decimal(0)
        for r in expense_rows:
            amt = Decimal(str(r["net_amount"]))
            rows.append([r["account_code"], r["account_name"], f"{amt:,.2f}"])
            total_expenses += amt
        rows.append(["", "Total Expenses", f"{total_expenses:,.2f}"])

        net_income = total_revenue - total_expenses
        rows.append(["", "NET INCOME", f"{net_income:,.2f}"])

        title = f"Profit & Loss – {company_name}"
        filename = f"profit_loss_{company_id}_{start_date}_{end_date}"
        return self._render_report(
            headers, rows, title, filename, format,
            subtitle=f"Period: {start_date} to {end_date}",
        )

    def _generate_balance_sheet(
        self,
        company_id: int,
        company_name: str,
        start_date: str,
        end_date: str,
        format: str,
    ) -> str:
        """Build a balance-sheet report."""
        account_data = self.db.fetchall(
            "SELECT coa.account_code, coa.account_name, "
            "coa.account_type, "
            "COALESCE(SUM(jel.debit), 0) AS total_debit, "
            "COALESCE(SUM(jel.credit), 0) AS total_credit "
            "FROM journal_entry_lines jel "
            "JOIN chart_of_accounts coa ON jel.account_id = coa.id "
            "JOIN journal_entries je ON jel.entry_id = je.id "
            "WHERE je.company_id = ? "
            "AND je.entry_date <= ? "
            "AND je.status = 'posted' "
            "GROUP BY coa.account_code, coa.account_name, "
            "         coa.account_type "
            "ORDER BY coa.account_code",
            (company_id, end_date),
        )

        headers = ["Account Code", "Account Name", "Amount (SAR)"]
        rows: List[list] = []
        section_totals: Dict[str, Decimal] = defaultdict(Decimal)

        sections = {
            "asset": "ASSETS",
            "liability": "LIABILITIES",
            "equity": "EQUITY",
        }
        grouped: Dict[str, list] = defaultdict(list)
        for acct in account_data:
            atype = acct["account_type"]
            grouped[atype].append(acct)

        for atype, label in sections.items():
            rows.append(["", f"--- {label} ---", ""])
            for acct in grouped.get(atype, []):
                debit = Decimal(str(acct["total_debit"]))
                credit = Decimal(str(acct["total_credit"]))
                if atype == "asset":
                    balance = debit - credit
                else:
                    balance = credit - debit
                rows.append([
                    acct["account_code"],
                    acct["account_name"],
                    f"{balance:,.2f}",
                ])
                section_totals[atype] += balance
            rows.append(["", f"Total {label}", f"{section_totals[atype]:,.2f}"])

        le_total = section_totals["liability"] + section_totals["equity"]
        rows.append(["", "TOTAL LIABILITIES + EQUITY", f"{le_total:,.2f}"])

        title = f"Balance Sheet – {company_name}"
        filename = f"balance_sheet_{company_id}_{end_date}"
        return self._render_report(
            headers, rows, title, filename, format,
            subtitle=f"As of {end_date}",
        )

    # ------------------------------------------------------------------
    # Invoice-specific PDF generation
    # ------------------------------------------------------------------

    def _generate_invoice_pdf(
        self,
        invoice: dict,
        company: Optional[dict],
        client_name: str,
        lines: List[dict],
        subtotal: Decimal,
        vat: Decimal,
        total: Decimal,
    ) -> str:
        """Create a professional invoice PDF with optional ZATCA QR."""
        filepath = os.path.join(
            tempfile.gettempdir(),
            f"{self._REPORT_PREFIX}invoice_{invoice['invoice_number']}.pdf",
        )
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        elements: List[Any] = []

        # Company header
        company_name = company["name"] if company else "Company"
        company_name_ar = company.get("name_ar", "") if company else ""
        vat_number = company.get("vat_number", "") if company else ""
        cr_number = (
            company.get("commercial_registration", "") if company else ""
        )

        header_style = ParagraphStyle(
            "InvoiceHeader", parent=styles["Heading1"],
            fontSize=18, spaceAfter=6,
        )
        elements.append(Paragraph(company_name, header_style))
        if company_name_ar:
            elements.append(Paragraph(company_name_ar, styles["Heading2"]))
        if vat_number:
            elements.append(
                Paragraph(f"VAT #: {vat_number}", styles["Normal"])
            )
        if cr_number:
            elements.append(
                Paragraph(f"CR #: {cr_number}", styles["Normal"])
            )
        elements.append(Spacer(1, 12))

        # Invoice meta
        meta_style = ParagraphStyle(
            "InvoiceMeta", parent=styles["Normal"], fontSize=10,
        )
        elements.append(
            Paragraph(
                f"<b>Invoice:</b> {invoice['invoice_number']}", meta_style,
            )
        )
        elements.append(
            Paragraph(
                f"<b>Date:</b> {invoice['invoice_date']}", meta_style,
            )
        )
        elements.append(
            Paragraph(f"<b>Client:</b> {client_name}", meta_style)
        )
        elements.append(
            Paragraph(
                f"<b>Status:</b> {invoice['status']}", meta_style,
            )
        )
        elements.append(Spacer(1, 12))

        # Line items table
        table_data = [["#", "Description", "Qty", "Unit Price", "Total"]]
        for idx, line in enumerate(lines, 1):
            table_data.append([
                str(idx),
                line["description"],
                str(line["quantity"]),
                f"{Decimal(str(line['unit_price'] or 0)):,.2f}",
                f"{Decimal(str(line['total'] or 0)):,.2f}",
            ])

        table_data.append(["", "", "", "Subtotal", f"{subtotal:,.2f}"])
        table_data.append(["", "", "", "VAT (15%)", f"{vat:,.2f}"])
        table_data.append(["", "", "", "Total (SAR)", f"{total:,.2f}"])

        col_widths = [0.5 * inch, 3.0 * inch, 0.7 * inch, 1.2 * inch, 1.2 * inch]
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (1, 1), (1, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -4), 0.5, colors.grey),
            ("LINEABOVE", (3, -3), (-1, -3), 1, colors.black),
            ("LINEABOVE", (3, -1), (-1, -1), 1.5, colors.black),
            ("FONTNAME", (0, -3), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (-1, -4), colors.HexColor("#ecf0f1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -4),
             [colors.white, colors.HexColor("#ecf0f1")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 18))

        # ZATCA TLV QR code
        qr_data = self._build_zatca_tlv(
            company_name, vat_number,
            str(invoice["invoice_date"]),
            str(total), str(vat),
        )
        if qr_data and HAS_QRCODE:
            try:
                qr_img = qrcode.make(qr_data)
                buf = io.BytesIO()
                qr_img.save(buf, format="PNG")
                buf.seek(0)
                elements.append(
                    Paragraph("ZATCA QR Code:", styles["Heading4"])
                )
                elements.append(RLImage(buf, width=1.5 * inch, height=1.5 * inch))
            except Exception as exc:
                self._logger.warning("QR generation failed: %s", exc)

        try:
            doc.build(elements)
        except Exception as exc:
            self._logger.error("PDF build failed: %s", exc)
            raise
        self._logger.info("Invoice PDF generated: %s", filepath)
        return filepath

    # ------------------------------------------------------------------
    # ZATCA TLV helper
    # ------------------------------------------------------------------

    @staticmethod
    def _build_zatca_tlv(
        seller: str,
        vat_number: str,
        timestamp: str,
        total: str,
        vat_amount: str,
    ) -> Optional[str]:
        """Build a Base64-encoded ZATCA TLV string for QR codes.

        Returns:
            Base64 string or ``None`` when inputs are insufficient.
        """
        if not seller or not vat_number:
            return None
        try:
            tlv = b""
            for tag, value in [
                (1, seller),
                (2, vat_number),
                (3, timestamp),
                (4, total),
                (5, vat_amount),
            ]:
                encoded = value.encode("utf-8")
                tlv += bytes([tag, len(encoded)]) + encoded
            return base64.b64encode(tlv).decode("ascii")
        except Exception:
            return None

    # ------------------------------------------------------------------
    # PDF header helper
    # ------------------------------------------------------------------

    def _create_pdf_header(
        self, canvas_obj: Any, title: str, company_name: str
    ) -> None:
        """Draw a standard header on a reportlab canvas page.

        Args:
            canvas_obj: A ``reportlab.pdfgen.canvas.Canvas`` instance.
            title: The report title.
            company_name: Company name rendered above the title.
        """
        canvas_obj.saveState()
        canvas_obj.setFont("Helvetica-Bold", 14)
        page_w, _ = A4
        canvas_obj.drawCentredString(
            page_w / 2, 800, company_name,
        )
        canvas_obj.setFont("Helvetica", 11)
        canvas_obj.drawCentredString(page_w / 2, 780, title)
        canvas_obj.setFont("Helvetica", 8)
        canvas_obj.drawRightString(
            page_w - 40, 800,
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        )
        canvas_obj.line(40, 775, page_w - 40, 775)
        canvas_obj.restoreState()

    # ------------------------------------------------------------------
    # Unified rendering dispatcher
    # ------------------------------------------------------------------

    def _render_report(
        self,
        headers: List[str],
        rows: List[list],
        title: str,
        filename: str,
        format: str,
        subtitle: str = "",
    ) -> str:
        """Route report generation to the appropriate format handler.

        Args:
            headers: Column header strings.
            rows: Row data (list of lists).
            title: Report title.
            filename: Base filename (without extension).
            format: 'pdf', 'excel', or 'csv'.
            subtitle: Optional subtitle line.

        Returns:
            Absolute path to the generated file.
        """
        fmt = format.lower().strip()
        if fmt == "pdf" and HAS_REPORTLAB:
            return self._generate_pdf(headers, rows, title, filename, subtitle)
        elif fmt == "excel" and HAS_OPENPYXL:
            return self._generate_excel(headers, rows, title, filename)
        else:
            if fmt not in ("csv",) and fmt == "pdf" and not HAS_REPORTLAB:
                self._logger.warning(
                    "reportlab unavailable; falling back to CSV"
                )
            elif fmt not in ("csv",) and fmt == "excel" and not HAS_OPENPYXL:
                self._logger.warning(
                    "openpyxl unavailable; falling back to CSV"
                )
            return self._generate_csv_fallback(headers, rows, filename)

    # ------------------------------------------------------------------
    # PDF generation (reportlab)
    # ------------------------------------------------------------------

    def _generate_pdf(
        self,
        headers: List[str],
        rows: List[list],
        title: str,
        filename: str,
        subtitle: str = "",
    ) -> str:
        """Generate a PDF report using reportlab.

        Args:
            headers: Column headers.
            rows: Data rows (list of lists).
            title: Report title.
            filename: Base filename.
            subtitle: Optional subtitle.

        Returns:
            Absolute path to the PDF file.
        """
        filepath = os.path.join(
            tempfile.gettempdir(),
            f"{self._REPORT_PREFIX}{filename}.pdf",
        )
        page_size = landscape(A4) if len(headers) > 6 else A4
        doc = SimpleDocTemplate(
            filepath, pagesize=page_size,
            leftMargin=36, rightMargin=36,
            topMargin=50, bottomMargin=36,
        )
        styles = getSampleStyleSheet()
        elements: List[Any] = []

        title_style = ParagraphStyle(
            "ReportTitle", parent=styles["Heading1"],
            fontSize=16, spaceAfter=4, alignment=1,
        )
        elements.append(Paragraph(title, title_style))

        if subtitle:
            sub_style = ParagraphStyle(
                "ReportSubtitle", parent=styles["Normal"],
                fontSize=9, alignment=1, spaceAfter=10,
                textColor=colors.HexColor("#555555"),
            )
            elements.append(Paragraph(subtitle, sub_style))

        elements.append(Spacer(1, 8))

        # Build table
        table_data = [headers] + rows
        num_cols = len(headers)
        avail_width = (
            page_size[0] - doc.leftMargin - doc.rightMargin
        )
        col_width = avail_width / num_cols
        col_widths = [col_width] * num_cols

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2),
             [colors.white, colors.HexColor("#f5f6fa")]),
        ]
        if rows:
            style_commands.append(
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")
            )
            style_commands.append(
                ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black)
            )
        table.setStyle(TableStyle(style_commands))
        elements.append(table)

        # Footer with generation timestamp
        elements.append(Spacer(1, 14))
        footer_style = ParagraphStyle(
            "ReportFooter", parent=styles["Normal"],
            fontSize=7, textColor=colors.HexColor("#999999"),
        )
        elements.append(Paragraph(
            f"Generated on "
            f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            footer_style,
        ))

        try:
            doc.build(elements)
        except Exception as exc:
            self._logger.error("PDF generation failed: %s", exc)
            raise
        self._logger.info("PDF report generated: %s", filepath)
        return filepath

    # ------------------------------------------------------------------
    # Excel generation (openpyxl)
    # ------------------------------------------------------------------

    def _generate_excel(
        self,
        headers: List[str],
        rows: List[list],
        title: str,
        filename: str,
    ) -> str:
        """Generate an Excel workbook using openpyxl.

        Args:
            headers: Column headers.
            rows: Data rows.
            title: Worksheet title (truncated to 31 chars for Excel).
            filename: Base filename.

        Returns:
            Absolute path to the Excel file.
        """
        filepath = os.path.join(
            tempfile.gettempdir(),
            f"{self._REPORT_PREFIX}{filename}.xlsx",
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]

        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(
            start_color="2C3E50", end_color="2C3E50", fill_type="solid",
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True,
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title row
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1, end_column=len(headers),
        )
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # Timestamp row
        ws.merge_cells(
            start_row=2, start_column=1,
            end_row=2, end_column=len(headers),
        )
        ts_cell = ws.cell(
            row=2, column=1,
            value=f"Generated: "
                  f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        )
        ts_cell.font = Font(italic=True, size=8, color="999999")
        ts_cell.alignment = Alignment(horizontal="center")

        header_row = 4
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        alt_fill = PatternFill(
            start_color="F5F6FA", end_color="F5F6FA", fill_type="solid",
        )
        data_alignment = Alignment(horizontal="center", vertical="center")
        for row_idx, row in enumerate(rows, header_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = data_alignment
                if (row_idx - header_row) % 2 == 0:
                    cell.fill = alt_fill

        # Bold last row (totals)
        if rows:
            last_row = header_row + len(rows)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=last_row, column=col_idx)
                cell.font = Font(bold=True, size=10)

        # Auto-column widths
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1]))
            for row in rows:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            adjusted = min(max_len + 4, 50)
            ws.column_dimensions[
                get_column_letter(col_idx)
            ].width = adjusted

        # Freeze header
        ws.freeze_panes = f"A{header_row + 1}"

        try:
            wb.save(filepath)
        except Exception as exc:
            self._logger.error("Excel generation failed: %s", exc)
            raise
        self._logger.info("Excel report generated: %s", filepath)
        return filepath

    # ------------------------------------------------------------------
    # CSV fallback
    # ------------------------------------------------------------------

    def _generate_csv_fallback(
        self,
        headers: List[str],
        rows: List[list],
        filename: str,
    ) -> str:
        """Generate a CSV file as a universal fallback.

        Args:
            headers: Column headers.
            rows: Data rows.
            filename: Base filename.

        Returns:
            Absolute path to the CSV file.
        """
        filepath = os.path.join(
            tempfile.gettempdir(),
            f"{self._REPORT_PREFIX}{filename}.csv",
        )
        try:
            with open(filepath, "w", newline="", encoding="utf-8-sig") as fh:
                writer = csv.writer(fh)
                writer.writerow(headers)
                writer.writerows(rows)
        except Exception as exc:
            self._logger.error("CSV generation failed: %s", exc)
            raise
        self._logger.info("CSV report generated: %s", filepath)
        return filepath

